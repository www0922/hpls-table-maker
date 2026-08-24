"""
统一格式化引擎
=============
PE100/PE150 的 format_font.py 核心逻辑提取。

设计思路：完全相同的引擎 + 配置参数驱动差异。
"""

import openpyxl
from openpyxl.styles import Font

DEFAULT_FONT = Font(name='Times New Roman', size=10)


def auto_column_width(ws, max_width=60, sample_rows=200):
    """自动计算并设置工作表的列宽。

    中文字符按 2 倍宽度补偿（与西文等宽字体不同）。

    Args:
        ws: openpyxl Worksheet
        max_width: 列宽上限（默认 60）
        sample_rows: 采样的最大行数（默认 200）
    """
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        limit = min(ws.max_row, sample_rows)
        for row in range(1, limit + 1):
            try:
                val = str(ws.cell(row=row, column=col_idx).value or '')
                char_len = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, char_len)
            except (AttributeError, TypeError):
                pass
        if max_len > 0:
            letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[letter].width = min(max_len + 4, max_width)


def apply_standard_format(
    wb,
    *,
    sheet_order,
    data_sheet_predicate=None,
    header_height_map=None,
    data_header_height=90,
    data_row_height=30,
    font=None,
    number_formats=None,
    max_col_width=60,
    verbose=True,
):
    """对 workbook 统一应用标准格式。

    处理流程（每个 sheet）：
    1. 设置表头行高
    2. 设置数据行行高
    3. 设置字体 + 数字格式
    4. 自动列宽
    5. 按 sheet_order 调整 sheet 顺序

    Args:
        wb: openpyxl Workbook 对象
        sheet_order: sheet 排列顺序，如 ``['T7+制备']``。
                     不在列表中的 sheet 追加到末尾。
        data_sheet_predicate: 判断是否为数据 sheet 的函数，
                              如 ``lambda n: is_data_sheet(n)``。
                              用于表头行高判断。
        header_height_map: 特殊 sheet 的表头行高，
                           ``{sheet_name: height}``。
                           数据 sheet 统一用 data_header_height。
        data_header_height: 数据 sheet 表头行高（默认 90）
        data_row_height: 数据行行高（默认 30）
        font: openpyxl Font 对象（默认 Times New Roman 10pt）
        number_formats: 列数字格式，``{col_letter: format_str}``，
                        如 ``{'D': '0.00', 'E': '0.000'}``
        max_col_width: 自动列宽上限（默认 60）
        verbose: 是否打印每个 sheet 的处理进度（默认 True）
    """
    if font is None:
        font = DEFAULT_FONT
    if header_height_map is None:
        header_height_map = {}
    if number_formats is None:
        number_formats = {}
    if data_sheet_predicate is None:
        data_sheet_predicate = lambda n: False

    for sn in wb.sheetnames:
        ws = wb[sn]

        # ── 1. 表头行高 ──
        if data_sheet_predicate(sn):
            ws.row_dimensions[1].height = data_header_height
        elif sn in header_height_map:
            ws.row_dimensions[1].height = header_height_map[sn]

        # ── 2. 数据行行高 ──
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = data_row_height

        # ── 3. 字体 + 数字格式 ──
        is_offline_stats = sn == '下机数据统计模版'
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                try:
                    cell.font = font
                    if not is_offline_stats:
                        col_letter = cell.column_letter
                        if col_letter in number_formats:
                            cell.number_format = number_formats[col_letter]
                except AttributeError:
                    pass  # 跳过 MergedCell

        # ── 4. 自动列宽 ──
        auto_column_width(ws, max_width=max_col_width)

        if verbose:
            print(f'{sn}: 格式已设置')

    # ── 5. Sheet 顺序 ──
    ordered = [s for s in sheet_order if s in wb.sheetnames]
    remainder = [s for s in wb.sheetnames if s not in sheet_order]
    wb._sheets = [wb[s] for s in ordered + remainder]
