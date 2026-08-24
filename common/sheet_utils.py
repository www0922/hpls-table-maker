"""
Sheet 工具函数集
===============
PE100/PE150 共用的 sheet 判定、映射、同步逻辑。

差异通过参数化处理：
- sheet_map_a 的匹配优先级 → ``priority`` 参数
- sync_b_table_sheets 的额外表头补全 → ``ensure_columns`` 参数
"""

import openpyxl


def is_data_sheet(name):
    """判断是否为数据 sheet：单个大写字母 A-Z。

    三个项目 (PE100/PE150/XP) 完全相同。
    """
    return (
        len(name) == 1
        and name.isalpha()
        and name.isascii()
        and name.isupper()
    )


def sheet_map_a(src_path, *, priority=('number', 'letter')):
    """扫描 A 表 sheet 名，返回 [(源sheet名, 目标sheet字母)]。

    映射规则：
        number: "1"→A, "2"→B, "3"→C ...
        letter: "A"→A, "B"→B, "C"→C ...（直接使用，需大写）

    Args:
        src_path: A 表文件路径
        priority: 匹配优先级顺序。
                  PE100 用 ('letter', 'number')
                  PE150 用 ('number', 'letter')
                  默认值为 PE150 顺序（先尝试数字解析更自然）

    Returns:
        list of (str, str): [(源sheet名, 目标sheet字母), ...]

    Raises:
        ValueError: A 表中没有有效 sheet 名
    """
    wb = openpyxl.load_workbook(src_path, data_only=True)
    src_sheets = list(wb.sheetnames)
    wb.close()

    mapping = []

    for sn in src_sheets:
        matched = False
        for rule in priority:
            if matched:
                break
            if rule == 'number':
                try:
                    num = int(sn)
                    letter = chr(ord('A') + num - 1)
                    mapping.append((sn, letter))
                    matched = True
                except ValueError:
                    pass
            elif rule == 'letter':
                if len(sn) == 1 and sn.isalpha() and sn.isascii() and sn.isupper():
                    mapping.append((sn, sn.upper()))
                    matched = True

    if not mapping:
        raise ValueError(f'A表中没有找到有效的sheet名: {src_sheets}')
    return mapping


def get_target_sheets(src_path, *, priority=('number', 'letter')):
    """从 A 表获取目标 B 表数据 sheet 列表。

    Args:
        src_path: A 表文件路径
        priority: 传给 sheet_map_a 的匹配优先级

    Returns:
        list of str: 排序后的目标 sheet 字母列表，如 ['C', 'D']
    """
    return sorted(set(dst for _, dst in sheet_map_a(src_path, priority=priority)))


def sync_b_table_sheets(wb, target_sheets, *, ensure_columns=None):
    """根据 A 表动态调整 B 表的数据 sheet。

    三个操作（按优先级）：
    1. **重命名匹配**：多余的 sheet 重命名为缺失的 sheet
    2. **删除多余**：A 表无对应数据的 sheet 直接删除
    3. **复制创建**：仍缺的 sheet 从现有数据 sheet 复制创建

    非数据 sheet（如 T7+制备、文库环化）保留不动。

    Args:
        wb: openpyxl Workbook 对象（B 表）
        target_sheets: 目标数据 sheet 列表，如 ['C', 'D']
        ensure_columns: 需要确保表头的列，dict[int, str] 格式。
                        PE150: {18: '人工检测结果', 19: '人工备注'}
                        PE100: None（跳过）

    Raises:
        ValueError: B 表中没有可用的数据 sheet 作为复制模板
    """
    existing = sorted(sn for sn in wb.sheetnames if is_data_sheet(sn))

    to_create = [t for t in target_sheets if t not in existing]
    to_delete = [e for e in existing if e not in target_sheets]

    # ── 1. 重命名匹配 ──
    while to_create and to_delete:
        new_name = to_create.pop(0)
        old_name = to_delete.pop(0)
        wb[old_name].title = new_name
        print(f'  [重命名] {old_name} → {new_name}')

    # ── 2. 删除多余 ──
    for sn in to_delete:
        del wb[sn]
        print(f'  [删除] 无数据sheet: {sn}')

    # ── 3. 复制创建 ──
    for sn in to_create:
        existing_data = [s for s in wb.sheetnames if is_data_sheet(s)]
        if not existing_data:
            raise ValueError('B表中没有可用的数据sheet作为复制模板')
        template_src = existing_data[0]
        ws_src = wb[template_src]
        ws_new = wb.copy_worksheet(ws_src)
        ws_new.title = sn
        # 调整新 sheet 位置：插入到数据 sheet 区域末尾
        data_indices = [
            i for i, s in enumerate(wb.sheetnames) if is_data_sheet(s)
        ]
        if data_indices:
            target_idx = max(data_indices)
            wb.move_sheet(
                ws_new, offset=target_idx - wb.sheetnames.index(sn)
            )
        print(f'  [创建] 复制 {template_src} → {sn}')

    # ── 4. 确保特定列表头（PE150: R/S 列） ──
    if ensure_columns:
        for sn in [s for s in wb.sheetnames if is_data_sheet(s)]:
            others = [
                s for s in wb.sheetnames if is_data_sheet(s) and s != sn
            ]
            src_sn = others[0] if others else sn
            ws = wb[sn]
            for col_idx, default_label in ensure_columns.items():
                if ws.cell(row=1, column=col_idx).value is None:
                    ws.cell(row=1, column=col_idx).value = (
                        wb[src_sn].cell(row=1, column=col_idx).value
                        or default_label
                    )
