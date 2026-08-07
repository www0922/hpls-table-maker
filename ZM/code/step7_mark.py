"""步骤七：标记“纯化”和“已定量”。

优先级：U+P时已定量优先 → 纯化 → 单文库已定量。

* 任一数据行 U 列含“纯化”，或 W 列人工备注出现独立标记 B：
  组内第一行 G 列写“纯化”；
* 单文库组 P 列已有 qPCR 结果：G 列写“已定量”；
* 多文库组不再无条件标记“已定量”。
"""

from __future__ import annotations

import re

import openpyxl

from config import DST_POOL
from pooling_utils import CENTER, normalized, read_groups


from config import get_target_sheets
SHEETS = get_target_sheets()
COL_G = 7
COL_P = 16
COL_U = 21
COL_W = 23


def contains_b_marker(value):
    text = normalized(value)
    return "B" in text.upper()


def process_sheet(ws, name):
    groups = read_groups(ws)
    pure_count = quantified_count = 0

    for group in groups:
        data_rows = group["data_rows"]
        if not data_rows:
            continue
        first_row = data_rows[0]
        has_u = any("纯化" in normalized(ws.cell(row=row, column=COL_U).value) for row in data_rows)
        has_p = ws.cell(row=first_row, column=COL_P).value is not None
        needs_purification = has_u or any(
            contains_b_marker(ws.cell(row=row, column=COL_W).value) for row in data_rows
        )

        status = None
        # U+P都有 → 已定量 (最高优先)
        if has_u and has_p:
            status = "已定量"
            quantified_count += 1
        elif needs_purification:
            status = "纯化"
            pure_count += 1
        elif len(data_rows) == 1 and has_p:
            status = "已定量"
            quantified_count += 1

        if status:
            cell = ws.cell(row=first_row, column=COL_G)
            cell.value = status
            cell.alignment = CENTER

    print(
        f"  Sheet {name}: {len(groups)}组, 纯化{pure_count}组, 已定量{quantified_count}组"
    )


def main():
    workbook = openpyxl.load_workbook(DST_POOL)
    for name in SHEETS:
        print(f"\n处理 [{name}]")
        process_sheet(workbook[name], name)
    workbook.save(DST_POOL)
    print(f"\n[DONE] 步骤七完成 -> {DST_POOL}")


if __name__ == "__main__":
    main()
