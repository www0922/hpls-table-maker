"""统一 Pooling 输出字体、行高和列宽。"""

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill

from config import DST_POOL, get_target_sheets
from pooling_utils import last_business_row


FONT = Font(name="Times New Roman", size=10)
RED_BOLD = Font(name="Times New Roman", size=10, bold=True, color="FF0000")


def main(pool_wb=None, pcr_wb=None):
    workbook_local = pool_wb if pool_wb is not None else openpyxl.load_workbook(DST_POOL)

    for sheet_name in workbook_local.sheetnames:
        ws = workbook_local[sheet_name]
        from config import get_target_sheets
        is_dilution = ws.title not in set(get_target_sheets())
        max_col = 19 if is_dilution else 23
        max_row = last_business_row(ws)

        ws.row_dimensions[1].height = 40
        for row_index in range(2, max_row + 1):
            ws.row_dimensions[row_index].height = 30

        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            for cell in row:
                if is_dilution and cell.column == 17 and cell.row >= 3:
                    cell.font = RED_BOLD
                else:
                    cell.font = FONT
                # 静态浮点数超过3位小数则截断
                v = cell.value
                if isinstance(v, float) and v != int(v) and not (isinstance(v, str) and v.startswith("=")):
                    s = f"{v:.10f}".rstrip("0")
                    if "." in s and len(s.split(".")[1]) > 3:
                        cell.value = round(v, 3)

        for col_index in range(1, max_col + 1):
            max_length = 0
            for row_index in range(1, min(max_row + 1, 200)):
                value = str(ws.cell(row=row_index, column=col_index).value or "")
                char_length = sum(2 if ord(char) > 127 else 1 for char in value)
                max_length = max(max_length, char_length)
            if max_length > 0:
                letter = openpyxl.utils.get_column_letter(col_index)
                ws.column_dimensions[letter].width = min(max_length + 4, 60)

        # E/F列条件格式（排除汇总行：B列为纯数字）
        if not is_dilution:
            skip_summary = 'NOT(ISNUMBER(B2+0))'
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            light_red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            light_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_font = Font(color="FF0000")

            col_e = openpyxl.utils.get_column_letter(5)
            col_f = openpyxl.utils.get_column_letter(6)
            e_range = f"{col_e}2:{col_e}{max_row}"
            f_range = f"{col_f}2:{col_f}{max_row}"

            # E列有数值: 浅绿底红字
            ws.conditional_formatting.add(
                e_range,
                FormulaRule(formula=[f'AND({col_e}2>0,{skip_summary})'],
                            fill=light_green_fill, font=red_font))
            # F列有数值 0.25~2.5: 黄底红字
            ws.conditional_formatting.add(
                f_range,
                FormulaRule(formula=[f'AND({col_f}2<>\"\",{col_f}2>=0.25,{col_f}2<=2.5,{skip_summary})'],
                            fill=yellow_fill, font=red_font))
            # F列有数值但超出范围: 浅红底红字
            ws.conditional_formatting.add(
                f_range,
                FormulaRule(formula=[f'AND({col_f}2<>\"\",OR({col_f}2<0.25,{col_f}2>2.5),{skip_summary})'],
                            fill=light_red_fill, font=red_font))

        print(f"{sheet_name}: 格式已设置")

    # 子表排序: ABCD数据面 → 文库稀释计算表 → 下机数据统计模版
    desired = list(get_target_sheets()) + ["文库稀释计算表", "下机数据统计模版"]
    current = workbook_local.sheetnames
    new_order = [s for s in desired if s in current] + [s for s in current if s not in desired]
    workbook_local._sheets = [workbook_local[s] for s in new_order]

    if pool_wb is None:
        workbook_local.save(DST_POOL)
    print(f"\n[DONE] Pooling -> {DST_POOL}")

    # PCR 定量表字体及条件格式
    from config import DST_PCR
    pcr_wb_local = pcr_wb if pcr_wb is not None else openpyxl.load_workbook(DST_PCR)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for ws in pcr_wb_local.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=18):
            for cell in row:
                if cell.value is not None:
                    cell.font = FONT

        # 序号列有数据且片段列为空时，B列和H列填充黄色（仅主体数据区）
        # G3 = 最后序号+1, 主体数据起始行=6, last_main_row = G3 + 4
        g3_val = ws.cell(row=3, column=7).value
        last_main = g3_val + 4 if isinstance(g3_val, (int, float)) else ws.max_row
        ws.conditional_formatting.add(
            f"B6:B{last_main}",
            FormulaRule(formula=['AND(A6<>"",H6="")'], fill=yellow_fill))
        ws.conditional_formatting.add(
            f"H6:H{last_main}",
            FormulaRule(formula=['AND(A6<>"",H6="")'], fill=yellow_fill))

        # 子表格表头：加粗12pt（被上方全局字体覆盖，需补设）
        sub_header_row = last_main + 2
        sub_header_font = Font(name="Times New Roman", size=12, bold=True)
        for col in (3, 4, 5, 6):
            cell = ws.cell(row=sub_header_row, column=col)
            if cell.value is not None:
                cell.font = sub_header_font

        print(f"PCR [{ws.title}]: 字体已设置")
    if pcr_wb is None:
        pcr_wb_local.save(DST_PCR)
    print(f"[DONE] PCR -> {DST_PCR}")


if __name__ == "__main__":
    main()
