"""步骤九：填充 qPCR 定量表。"""

from __future__ import annotations

from copy import copy
import re

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.styles import Border, Font, Side

from config import DST_PCR, DST_POOL, get_src_a
from pooling_utils import CENTER, normalized, read_groups, safe_float
from step3_lookup import build_external_lookup, build_self_lookup, is_external_library
from step5_summary import is_bare_library, calculated_e_f


RED_BOLD = Font(bold=True, color="FF0000")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
DATA_START = 6


def is_group_name(sample_id):
    return bool(re.match(r"^[AB]\d+(?:-\d+(?:\.\d+)?(?:-\d+)?)?$", normalized(sample_id), re.IGNORECASE))


def is_hgc_prefix(sample_id):
    s = normalized(sample_id).upper()
    return s.startswith("HGC-") or s.startswith("HGC-LIB") or s.startswith("HGC-POOL")


def build_a_table_c_lookup():
    """构建 {HaploX编号(C列) → 样本名称(D列)} 从A表"""
    wb = openpyxl.load_workbook(get_src_a(), data_only=True)
    lookup = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in range(2, ws.max_row + 1):
            c = normalized(ws.cell(row=row, column=3).value)
            d = normalized(ws.cell(row=row, column=4).value)
            if c and d:
                lookup[c] = d
    wb.close()
    print(f"  A表C→D查找: {len(lookup)}条")
    return lookup


def collect_pooling_groups(workbook):
    result = {}
    from config import get_target_sheets
    for face in get_target_sheets():
        ws = workbook[face]
        for group in read_groups(ws):
            rows = group["data_rows"]
            if not rows:
                continue
            first = rows[0]
            sample_id = normalized(ws.cell(row=first, column=1).value)
            if not sample_id:
                continue
            # 从E/F公式反算M/N静态值(稀释后浓度)
            evaluated = [calculated_e_f(ws, row) for row in rows]
            e_sum = sum(e for e, f in evaluated)
            f_sum = sum(f for e, f in evaluated)
            conc = round(e_sum / f_sum, 3) if f_sum > 0 else None
            # M列: 稀释后(>15→12.5)
            m_raw = ws.cell(row=first, column=13).value
            if isinstance(m_raw, str) and m_raw.startswith("="):
                m_val = 12.5 if (conc and conc > 15) else conc
            else:
                m_val = safe_float(m_raw, default=None)
            # N列: 稀释后(>50→45)
            n_raw = ws.cell(row=first, column=14).value
            if isinstance(n_raw, str) and n_raw.startswith("="):
                n_val = 45.0 if (conc and conc > 50) else conc
            else:
                n_val = safe_float(n_raw, default=None)
            result[sample_id] = {
                "o": ws.cell(row=first, column=15).value,
                "k": ws.cell(row=first, column=11).value,
                "m": m_val,
                "n": n_val,
                "data_amount": sum(
                    safe_float(ws.cell(row=row, column=4).value) for row in rows
                ),
            }
    return result


def collect_pending_dilution_rows(workbook):
    ws = workbook['文库稀释计算表']
    rows = []
    for row in range(3, ws.max_row + 1):
        status = normalized(ws.cell(row=row, column=1).value)
        sample_id = normalized(ws.cell(row=row, column=2).value)
        if not sample_id or status == "已定量":
            continue
        rows.append({
            "status": status,
            "sample_id": sample_id,
            "data_amount": ws.cell(row=row, column=8).value,
            "library_type": ws.cell(row=row, column=7).value,
            "contract": ws.cell(row=row, column=18).value,
            "remark": ws.cell(row=row, column=17).value,
            "customer": ws.cell(row=row, column=19).value,
        })
    return rows


def source_values(sample_id, pool_record, external_lookup, self_hybrid, self_sample):
    n_value = None; o_value = None
    if is_group_name(sample_id):
        n_value = pool_record.get("m")
        o_value = pool_record.get("n")
    elif is_external_library(sample_id):
        external = external_lookup.get(sample_id)
        if external:
            n_value = external.get("conc_l")
    else:
        self_record = self_hybrid.get(sample_id) or self_sample.get(sample_id)
        if self_record:
            n_value = self_record.get("qubit")
    return n_value, o_value


def clear_pcr_targets(ws):
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= DATA_START:
            ws.unmerge_cells(str(merged))
    for row in range(DATA_START, ws.max_row + 1):
        for col in range(1, 19):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            cell.value = None


def write_record(ws, row, seq, record, pool_record, a_lookup, external_lookup, self_hybrid, self_sample):
    sample_id = record["sample_id"]
    n_value, o_value = source_values(sample_id, pool_record, external_lookup, self_hybrid, self_sample)

    # A列: 序号
    cell = ws.cell(row=row, column=1)
    cell.value = seq
    cell.alignment = CENTER

    # B列: SampleID
    cell = ws.cell(row=row, column=2)
    cell.value = sample_id
    cell.alignment = CENTER

    # C列: HGC开头→A表D列(样本名称); 非HGC→复制B列
    c_value = a_lookup.get(sample_id, sample_id) if is_hgc_prefix(sample_id) else sample_id
    cell = ws.cell(row=row, column=3)
    cell.value = c_value
    cell.alignment = CENTER

    # D列: 数据量
    cell = ws.cell(row=row, column=4)
    cell.value = record["data_amount"] or pool_record.get("data_amount")
    cell.alignment = CENTER

    # E列: 文库类型 (稀释表G列)
    cell = ws.cell(row=row, column=5)
    cell.value = record["library_type"]
    cell.alignment = CENTER

    # F列: 合同编号
    cell = ws.cell(row=row, column=6)
    cell.value = record["contract"]
    cell.alignment = CENTER

    # H列: 片段
    if is_group_name(sample_id):
        h_value = pool_record.get("o")
    elif is_external_library(sample_id) or is_bare_library(sample_id):
        h_value = pool_record.get("k")
    else:
        h_value = pool_record.get("o")
    cell = ws.cell(row=row, column=8)
    cell.value = h_value
    cell.alignment = CENTER

    # I列: =G/H/650×1000000
    cell = ws.cell(row=row, column=9)
    cell.value = f"=G{row}/H{row}/650*1000000"
    cell.alignment = CENTER

    # J列: 纯化标记
    if record["status"] == "纯化":
        cell = ws.cell(row=row, column=10)
        cell.value = "纯化"
        cell.alignment = CENTER
        cell.font = RED_BOLD

    # M列: =N/G
    cell = ws.cell(row=row, column=13)
    cell.value = f"=N{row}/G{row}"
    cell.alignment = CENTER

    # N列: 浓度
    cell = ws.cell(row=row, column=14)
    cell.value = n_value
    cell.alignment = CENTER

    # O列
    cell = ws.cell(row=row, column=15)
    cell.value = o_value
    cell.alignment = CENTER

    # P列: =O/I
    cell = ws.cell(row=row, column=16)
    cell.value = f"=O{row}/I{row}"
    cell.alignment = CENTER

    # Q列: 备注
    cell = ws.cell(row=row, column=17)
    cell.value = record["remark"]
    cell.alignment = CENTER

    # R列: 客户单位
    cell = ws.cell(row=row, column=18)
    cell.value = record["customer"]
    cell.alignment = CENTER

    # 框线 A-I
    for col in range(1, 10):
        ws.cell(row=row, column=col).border = THIN_BORDER


def main(pool_wb=None, pcr_wb=None):
    pool_wb_local = pool_wb if pool_wb is not None else openpyxl.load_workbook(DST_POOL, data_only=False)
    pool_groups = collect_pooling_groups(pool_wb_local)
    pending = collect_pending_dilution_rows(pool_wb_local)
    a_lookup = build_a_table_c_lookup()
    external_lookup = build_external_lookup()
    self_hybrid, self_sample = build_self_lookup()

    from datetime import datetime
    pcr_wb_local = pcr_wb if pcr_wb is not None else openpyxl.load_workbook(DST_PCR)
    base_title = datetime.now().strftime("%m%d")
    pcr_wb_local.worksheets[0].title = base_title

    ws = pcr_wb_local.worksheets[0]
    clear_pcr_targets(ws)

    for index, record in enumerate(pending):
        row = DATA_START + index
        seq = index + 1
        pool_record = pool_groups.get(record["sample_id"], {})
        write_record(ws, row, seq, record, pool_record, a_lookup, external_lookup, self_hybrid, self_sample)

    # G3: 最后序号+1
    ws.cell(row=3, column=7).value = len(pending) + 1

    # --- 子表格（每条2行，纵向合并）---
    last_main_row = ws.max_row
    sub_start = last_main_row + 2  # 空1行

    # 表头行：C=样本名称, D=浓度ng/ul, E=平行检测人, F=XCL, G=空
    sub_headers = {3: "样本名称", 4: "浓度ng/ul", 5: "平行检测人", 6: "XCL"}
    for col, title in sub_headers.items():
        cell = ws.cell(row=sub_start, column=col)
        cell.value = title
        cell.alignment = CENTER
        cell.font = Font(bold=True, size=12)
        cell.border = THIN_BORDER
    # G列表头: 空（与主表浓度ng/ul列对齐）
    ws.cell(row=sub_start, column=7).border = THIN_BORDER

    for idx, record in enumerate(pending):
        r = sub_start + 1 + idx * 2  # 每条占2行, r和r+1
        seq = idx + 1
        sample_id = record["sample_id"]

        # A列: 序号 (合并2行, 无框线)
        ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=1)
        cell_a = ws.cell(row=r, column=1)
        cell_a.value = seq
        cell_a.alignment = CENTER

        # C列: 样本名称 (合并2行)
        ws.merge_cells(start_row=r, start_column=3, end_row=r + 1, end_column=3)
        cell_c = ws.cell(row=r, column=3)
        cell_c.value = sample_id
        cell_c.alignment = CENTER
        ws.cell(row=r, column=3).border = THIN_BORDER
        ws.cell(row=r + 1, column=3).border = THIN_BORDER

        # D列: 浓度ng/ul (不合并, 各行独立, 用户手动填入, 保留两位小数)
        for d_row in (r, r + 1):
            cell_d = ws.cell(row=d_row, column=4)
            cell_d.alignment = CENTER
            cell_d.border = THIN_BORDER
            cell_d.number_format = '0.00'

        # E列: 平行检测人 (合并2行, =AVERAGE(D{r},D{r+1}))
        ws.merge_cells(start_row=r, start_column=5, end_row=r + 1, end_column=5)
        cell_e = ws.cell(row=r, column=5)
        cell_e.value = f'=AVERAGE(D{r},D{r+1})'
        cell_e.alignment = CENTER
        cell_e.number_format = '0.00'
        cell_e.border = THIN_BORDER
        ws.cell(row=r + 1, column=5).border = THIN_BORDER

        # F列: XCL (合并2行, =STDEV(D{r},D{r+1}))
        ws.merge_cells(start_row=r, start_column=6, end_row=r + 1, end_column=6)
        cell_f = ws.cell(row=r, column=6)
        cell_f.value = f'=STDEV(D{r},D{r+1})'
        cell_f.alignment = CENTER
        cell_f.number_format = '0.00'
        cell_f.border = THIN_BORDER
        ws.cell(row=r + 1, column=6).border = THIN_BORDER

        # G列: 与主表浓度ng/ul列对齐 (合并2行, =F/E)
        ws.merge_cells(start_row=r, start_column=7, end_row=r + 1, end_column=7)
        cell_g = ws.cell(row=r, column=7)
        cell_g.value = f'=F{r}/E{r}'
        cell_g.alignment = CENTER
        cell_g.number_format = '0.00'
        cell_g.border = THIN_BORDER
        ws.cell(row=r + 1, column=7).border = THIN_BORDER

    if pool_wb is None:
        pool_wb_local.close()
    if pcr_wb is None:
        pcr_wb_local.save(DST_PCR)
    print(f"  qPCR定量表: 写入{len(pending)}条, 子表格起始行={sub_start}")
    print(f"\n[DONE] 步骤九完成 -> {DST_PCR}")


if __name__ == "__main__":
    main()
