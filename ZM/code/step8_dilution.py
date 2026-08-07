"""步骤八：制作文库稀释计算表。"""

from __future__ import annotations

from copy import copy
from datetime import datetime
import re

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.styles import Border, Font, Side

from config import DST_POOL
from pooling_utils import CENTER, normalized, read_groups, safe_float


from config import get_target_sheets
SHEETS = get_target_sheets()
RED_BOLD = Font(bold=True, color="FF0000")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

P_A = 1; P_D = 4; P_G = 7; P_H = 8
P_P = 16; P_Q = 17; P_R = 18; P_S = 19; P_T = 20

D_C = 3; D_D = 4; D_E = 5; D_F = 6; D_H = 8
D_I = 9; D_J = 10; D_K = 11; D_L = 12

TARGET_COLUMNS = {1, 2, D_C, D_D, D_E, D_H, D_I, D_J, D_K, D_L, 17, 18, 19}


def collect_groups(ws):
    result = []
    for group in read_groups(ws):
        rows = group["data_rows"]
        if not rows:
            continue
        first = rows[0]
        status = normalized(ws.cell(row=first, column=P_G).value)
        if status not in {"纯化", "已定量"}:
            status = "需要定量"
        lane_id = normalized(ws.cell(row=first, column=P_T).value)
        result.append({
            "face": ws.title,
            "lane_id": lane_id,
            "lane_label": f"{ws.title}{lane_id}",
            "status": status,
            "sample_id": normalized(ws.cell(row=first, column=P_A).value),
            "qpcr": ws.cell(row=first, column=P_P).value,
            "library_type": ws.cell(row=first, column=P_H).value,
            "data_amount": sum(safe_float(ws.cell(row=row, column=P_D).value) for row in rows),
            "contract": ws.cell(row=first, column=P_Q).value,
            "customer": ws.cell(row=first, column=P_R).value,
            "remark": ws.cell(row=first, column=P_S).value,
        })
    return result


def clear_target_data(ws):
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= 3:
            ws.unmerge_cells(str(merged))
    for row in range(3, ws.max_row + 1):
        for col in TARGET_COLUMNS:
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            cell.value = None


def write_record(ws, row, record):
    mapping = {
        1: record["status"],
        2: record["sample_id"],
        D_C: record["qpcr"],                      # qPCR浓度
        7: record["library_type"],                # 文库类型
        D_H: record["data_amount"],               # 数据量G
        D_I: record["lane_label"],                # laneID (A1, B2...)
        17: record["remark"],                     # 备注
        18: record["contract"],                   # 合同编号
        19: record["customer"],                   # 客户单位
    }
    for col, value in mapping.items():
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.alignment = CENTER

    # C/D/E/F/H 保留两位小数
    for col in (D_C, D_D, D_E, D_F, D_H):
        ws.cell(row=row, column=col).number_format = '0.00'

    # D列: 体积=2
    cell = ws.cell(row=row, column=D_D)
    cell.value = 2
    cell.alignment = CENTER

    # E列: =C×D/F-D
    cell = ws.cell(row=row, column=D_E)
    cell.value = f"=C{row}*D{row}/F{row}-D{row}"
    cell.alignment = CENTER

    # J列: =H/10
    cell = ws.cell(row=row, column=D_J)
    cell.value = f"=H{row}/10"
    cell.alignment = CENTER

    # K列: =J/H×10
    cell = ws.cell(row=row, column=D_K)
    cell.value = f"=J{row}/H{row}*10"
    cell.alignment = CENTER

    # L列: =E
    cell = ws.cell(row=row, column=D_L)
    cell.value = f"=E{row}"
    cell.alignment = CENTER

    # 框线 A-I
    for col in range(1, 10):
        ws.cell(row=row, column=col).border = THIN_BORDER

    ws.cell(row=row, column=17).font = RED_BOLD


def is_hgc_or_ex(sample_id):
    """以 HGC 开头或以 E/X 结尾 → 排在组末尾。"""
    s = str(sample_id or "")
    return bool(re.match(r"^HGC", s, re.IGNORECASE) or re.search(r"[EX]$", s, re.IGNORECASE))


def main():
    workbook = openpyxl.load_workbook(DST_POOL)
    dilution_sheet = workbook['文库稀释计算表']
    dilution_sheet.cell(row=1, column=1).value = f"{datetime.now():%Y%m%d}-LH00"
    dilution_sheet.cell(row=2, column=2).value = "SampleID"

    records = []
    for name in SHEETS:
        sheet_records = collect_groups(workbook[name])
        print(f"  Sheet {name}: {len(sheet_records)}组")
        records.extend(sheet_records)

    # 同 lane 内 HGC/E/X 条目排在末尾
    records.sort(key=lambda r: (r["face"], r["lane_id"], is_hgc_or_ex(r["sample_id"])))

    clear_target_data(dilution_sheet)
    current_row = 3
    previous_lane = None

    for record in records:
        lane_key = (record["face"], record["lane_id"])
        if previous_lane is not None and lane_key != previous_lane:
            for col in range(1, 20):
                cell = dilution_sheet.cell(row=current_row, column=col)
                cell.value = None
                cell.border = Border()
            current_row += 1
        previous_lane = lane_key

        write_record(dilution_sheet, current_row, record)
        current_row += 1

    # 每组I列合并单元格
    current_start = None
    for row in range(3, current_row):
        has_data = dilution_sheet.cell(row=row, column=2).value is not None
        if has_data and current_start is None:
            current_start = row
        elif not has_data and current_start is not None:
            if row - 1 > current_start:
                dilution_sheet.merge_cells(start_row=current_start, start_column=D_I,
                                           end_row=row - 1, end_column=D_I)
            current_start = None
    if current_start is not None and current_row - 1 > current_start:
        dilution_sheet.merge_cells(start_row=current_start, start_column=D_I,
                                   end_row=current_row - 1, end_column=D_I)

    workbook.save(DST_POOL)
    print(f"  稀释表: {len(records)}行")
    print(f"\n[DONE] 步骤八完成 -> {DST_POOL}")


if __name__ == "__main__":
    main()
