"""步骤三：按文库编号填充浓度、片段、板号、状态及质检信息。

已确认的列规则：

* HGC-Lib/HGC-POOL：外包表 M 为空时取 L；M 有值时取 P，
  同时写入 Pooling C/I；M→K、O→L、Z→V、AA→W。
* HGC-数字、数字E、数字X及其他非外包编号：自建库
  K→C、N→K（采用不覆盖组浓度的安全规则）、O→L、S→G、T→L。
* qPCR 历史 K 列结果→Pooling P；纯化合同命中→U 列“纯化”。
"""

from __future__ import annotations

import openpyxl

from config import DST_POOL, get_src_c, get_src_d, get_src_purify
from pooling_utils import CENTER, is_summary_row, normalized
from source_lookups import build_qpcr_lookup


from config import get_target_sheets
SHEETS = get_target_sheets()

P_C = 3
P_G = 7
P_I = 9
P_K = 11
P_L = 12
P_P = 16
P_Q = 17
P_U = 21
P_V = 22
P_W = 23


def is_external_library(lib_id):
    value = normalized(lib_id).upper()
    return value.startswith("HGC-LIB") or value.startswith("HGC-POOL")


def build_external_lookup():
    workbook = openpyxl.load_workbook(get_src_c(), read_only=True, data_only=True)
    lookup = {}
    for worksheet in workbook.worksheets:
        for values in worksheet.iter_rows(
            min_row=2, max_col=27, values_only=True
        ):
            key = normalized(values[5])  # F: HGC编号
            if not key:
                continue
            lookup[key] = {
                "conc_l": values[11],
                "frag": values[12],
                "plate": values[14],
                "conc_p": values[15],
                "detection": values[25],
                "manual_remark": values[26],
            }
    workbook.close()
    print(f"  外包质检表: {len(lookup)}条")
    return lookup


def build_self_lookup():
    workbook = openpyxl.load_workbook(get_src_d(), read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    by_hybrid, by_sample = {}, {}
    for values in worksheet.iter_rows(min_row=2, max_col=20, values_only=True):
        record = {
            "qubit": values[10],
            "frag": values[13],
            "evaluation": values[14],
            "well": values[18],
            "plate": values[19],
        }
        hybrid = normalized(values[2])
        sample = normalized(values[3])
        if hybrid:
            by_hybrid[hybrid] = record
        if sample:
            by_sample[sample] = record
    workbook.close()
    print(f"  自建库表: 杂交编号{len(by_hybrid)}条, 样本编号{len(by_sample)}条")
    return by_hybrid, by_sample


def build_purify_lookup():
    workbook = openpyxl.load_workbook(get_src_purify(), read_only=True, data_only=True)
    contracts = set()
    ws = workbook['1-12月纯化']
    for values in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        contract = normalized(values[6])
        if contract:
            contracts.add(contract)
    workbook.close()
    print(f"  纯化表(1-12月纯化): {len(contracts)}个合同编号")
    return contracts


def write_value(ws, row, col, value):
    if value is None:
        return
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.alignment = CENTER


def has_value(value):
    return value is not None and normalized(value) != ""


def fill_sheet(
    ws,
    external_lookup,
    self_hybrid_lookup,
    self_sample_lookup,
    purify_contracts,
    qpcr_lookup,
    name,
):
    matched = missed = 0
    for row in range(2, ws.max_row + 1):
        if is_summary_row(ws, row):
            continue
        lib_id = normalized(ws.cell(row=row, column=2).value)
        if not lib_id:
            continue

        if is_external_library(lib_id):
            record = external_lookup.get(lib_id)
            if record:
                concentration = record["conc_p"] if has_value(record["frag"]) else record["conc_l"]
                write_value(ws, row, P_C, concentration)
                write_value(ws, row, P_I, concentration)
                write_value(ws, row, P_K, record["frag"])
                write_value(ws, row, P_L, record["plate"])
                write_value(ws, row, P_V, record["detection"])
                write_value(ws, row, P_W, record["manual_remark"])
                matched += 1
            else:
                missed += 1
        else:
            record = self_sample_lookup.get(lib_id) or self_hybrid_lookup.get(lib_id)
            if record:
                write_value(ws, row, P_C, record["qubit"])
                # K/M 列冲突尚未确认：片段保留在 K，M 专用于组浓度。
                write_value(ws, row, P_K, record["frag"])
                write_value(ws, row, P_L, record["evaluation"])
                write_value(ws, row, P_G, record["well"])
                if has_value(record["plate"]):
                    write_value(ws, row, P_L, record["plate"])
                matched += 1
            else:
                missed += 1

        qpcr_value = qpcr_lookup.get(lib_id)
        if qpcr_value is not None:
            write_value(ws, row, P_P, qpcr_value)

        contract = normalized(ws.cell(row=row, column=P_Q).value)
        if contract and contract in purify_contracts:
            write_value(ws, row, P_U, "纯化")

    print(f"  Sheet {name}: 匹配={matched}, 未匹配={missed}")


def main():
    external_lookup = build_external_lookup()
    self_hybrid_lookup, self_sample_lookup = build_self_lookup()
    purify_contracts = build_purify_lookup()
    qpcr_lookup = build_qpcr_lookup()

    workbook = openpyxl.load_workbook(DST_POOL)
    for name in SHEETS:
        print(f"\n处理 [{name}]")
        fill_sheet(
            workbook[name],
            external_lookup,
            self_hybrid_lookup,
            self_sample_lookup,
            purify_contracts,
            qpcr_lookup,
            name,
        )
    workbook.save(DST_POOL)
    print(f"\n[DONE] 步骤三完成 -> {DST_POOL}")


if __name__ == "__main__":
    main()
