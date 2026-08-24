"""清空输出副本中的旧业务数据，保留模板公式和格式。"""

import openpyxl

from config import reset_output, get_src_a, SHEET_RENAME
from pooling_utils import clear_business_rows


def main():
    pool_dst, pcr_dst = reset_output()

    # 根据A表确定需要的sheet, 映射为pooling表sheet名 (A→S, B→P)
    src_wb = openpyxl.load_workbook(get_src_a(), data_only=True)
    src_sheets = sorted(sn for sn in src_wb.sheetnames
                        if len(sn) == 1 and sn.isalpha() and sn.isascii() and sn.isupper())
    src_wb.close()
    target_sheets = [SHEET_RENAME.get(s, s) for s in src_sheets]
    print(f"上机表sheets: {src_sheets} → Pooling: {target_sheets}")

    pool_workbook = openpyxl.load_workbook(pool_dst)
    for name in target_sheets:
        if name not in pool_workbook.sheetnames:
            ws_a = pool_workbook["A"]
            ws_new = pool_workbook.copy_worksheet(ws_a)
            ws_new.title = name
            print(f"Pooling [{name}]: 从A面复制创建")
    # 删除多余的数据sheet
    for sn in list(pool_workbook.sheetnames):
        if len(sn) == 1 and sn.isalpha() and sn.isascii() and sn.isupper() and sn not in target_sheets:
            del pool_workbook[sn]
            print(f"Pooling [{sn}]: 删除(无对应上机数据)")
    for name in target_sheets:
        clear_business_rows(pool_workbook[name])
        print(f"Pooling [{name}]: 已清空")
    pool_workbook.save(pool_dst)

    pcr_workbook = openpyxl.load_workbook(pcr_dst)
    for ws in pcr_workbook.worksheets:
        for merged in list(ws.merged_cells.ranges):
            if merged.min_row >= 6:
                ws.unmerge_cells(str(merged))
        for row in range(6, ws.max_row + 1):
            for col in range(1, min(ws.max_column, 18) + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    continue
                cell.value = None
        print(f"PCR [{ws.title}]: 已清空")
    pcr_workbook.save(pcr_dst)

    print(f"\n[DONE] Pooling -> {pool_dst}")
    print(f"[DONE] PCR    -> {pcr_dst}")


if __name__ == "__main__":
    main()
