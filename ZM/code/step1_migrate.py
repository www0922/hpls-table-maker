"""
步骤一：A表 → Pooling表 数据迁移
================================
Sheet映射: A→A, B→B

列映射 (A表 → Pooling表):
  C列(HaploX编号) → B列(文库编号)
  C列为空 → D列(样本名称) → B列
  D列含"HGC"/"Lib"等多文库标记 → 只取第一行信息
  B列合并格(260480E/260480X等) → 取B列第一个值 → B列, J列求和 → D列
  J列(预分配数据量G) → D列
  H列(文库类型) → H列
  K列(合同编号) → Q列
  L列(客户单位) → R列
  M列(备注) → S列
  O列(laneID) → T列
"""
import openpyxl
from openpyxl.styles import Alignment
from collections import defaultdict

from config import DST_POOL, SHEET_RENAME, get_src_a

SRC = get_src_a()
_wb = openpyxl.load_workbook(SRC, data_only=True)
SHEET_MAP = [(sn, SHEET_RENAME.get(sn, sn)) for sn in _wb.sheetnames if len(sn) == 1 and sn.isalpha() and sn.isascii() and sn.isupper()]
_wb.close()
CENTER = Alignment(horizontal='center', vertical='center')

# 列索引
COL_B = 2    # B列: 文库编号
COL_C = 3    # C列: HaploX编号
COL_D = 4    # D列: 样本名称 / 数据量(目标)
COL_G = 7    # -
COL_H = 8    # H列: 文库类型
COL_J = 10   # J列: 预分配数据量G
COL_L = 12   # L列: 客户单位
COL_M = 13   # M列: 备注
COL_O = 15   # O列: Lane_ID
SRC_COL_K = 11   # 来源K列: 合同编号
COL_Q = 17   # 目标Q列: 合同编号
COL_R = 18   # R列: 客户单位(目标)
COL_S = 19   # S列: 备注(目标)
COL_T = 20   # T列: laneID(目标)


def find_merge_groups(ws):
    """扫描B列和C列合并单元格, 返回 {row: {group_id, sum_range}}"""
    groups = defaultdict(lambda: {'parent_id': None, 'sum_range': None, 'first_range': None})
    for mr in ws.merged_cells.ranges:
        if mr.min_col <= COL_B <= mr.max_col:
            parent_id = ws.cell(row=mr.min_row, column=COL_B).value
            for r in range(mr.min_row, mr.max_row + 1):
                groups[r]['parent_id'] = parent_id
                groups[r]['sum_range'] = (mr.min_row, mr.max_row)
                groups[r]['first_range'] = mr.min_row
        if mr.min_col <= COL_C <= mr.max_col:
            parent_id = ws.cell(row=mr.min_row, column=COL_C).value
            if parent_id and 'POOL' in str(parent_id).upper():
                for r in range(mr.min_row, mr.max_row + 1):
                    if groups[r]['parent_id'] is None:
                        groups[r]['parent_id'] = parent_id
                        groups[r]['sum_range'] = (mr.min_row, mr.max_row)
                        groups[r]['first_range'] = mr.min_row
    return dict(groups)


def resolve_lib_id(row, ws, groups):
    """文库编号: 合并格parent_id > HaploX > 样本名称"""
    g = groups.get(row, {})
    if g.get('parent_id') is not None:
        return str(g['parent_id'])

    c = ws.cell(row=row, column=COL_C).value
    if c and str(c).strip() and str(c) != 'None':
        return str(c).strip()

    d = ws.cell(row=row, column=COL_D).value
    if d and str(d).strip() and str(d) != 'None':
        return str(d).strip()

    return None


def resolve_data_amount(row, ws, groups, cache):
    """数据量: 合并格求和, 否则J列值"""
    g = groups.get(row, {})
    if g.get('parent_id') is not None:
        key = g['sum_range']
        if key not in cache:
            total = 0
            for r in range(key[0], key[1] + 1):
                v = ws.cell(row=r, column=COL_J).value
                if isinstance(v, (int, float)):
                    total += v
            cache[key] = total
        return cache[key]

    v = ws.cell(row=row, column=COL_J).value
    return v if isinstance(v, (int, float)) else 0


def get_first_row_info(row, ws, groups):
    """合并格内的信息取第一行(非J列)"""
    g = groups.get(row, {})
    src_row = g.get('first_range', row) if g.get('parent_id') else row
    return {
        'lib_type': ws.cell(row=src_row, column=COL_H).value,   # H列
        'contract': ws.cell(row=src_row, column=SRC_COL_K).value,   # K列
        'customer': ws.cell(row=src_row, column=COL_L).value,   # L列
        'remark': ws.cell(row=src_row, column=COL_M).value,     # M列
        'lane_id': ws.cell(row=src_row, column=COL_O).value,    # O列
    }


def is_multi_lib_in_d(d_val):
    """D列是否包含多个文库编号"""
    s = str(d_val or '')
    return 'HGC' in s or 'Lib' in s


def migrate_sheet(ws_src, ws_dst, name):
    """迁移单个工作表"""
    groups = find_merge_groups(ws_src)
    sum_cache = {}

    merge_count = len(set(g.get('sum_range') for g in groups.values() if g.get('sum_range')))
    print(f'  Sheet {name}: {ws_src.max_row - 1}条数据, {merge_count}个B列合并组')

    # 标题文字(精确匹配, 排除A表中混入的表头行)
    HEADER_TEXTS = {'HaploX编号', '文库名称', '文库类型', '客户单位', '杂交文库编号', 'Lane_ID', 'Pooling文库编号', '备注', 'Index编号', 'i7序列', 'i5序列'}

    # 收集数据
    rows_data = []
    for row in range(2, ws_src.max_row + 1):
        lib_id = resolve_lib_id(row, ws_src, groups)
        if not lib_id:
            continue
        if str(lib_id).strip() in HEADER_TEXTS:
            continue  # 跳过标题行
        data_amt = resolve_data_amount(row, ws_src, groups, sum_cache)
        info = get_first_row_info(row, ws_src, groups)

        # D列含多文库 → 只取第一行(已通过groups处理合并格)
        rows_data.append({
            'lib_id': lib_id,
            'data_amt': data_amt,
            'lib_type': info['lib_type'],
            'contract': info['contract'],
            'customer': info['customer'],
            'remark': info['remark'],
            'lane_id': info['lane_id'],
        })

    # 去重
    seen = set()
    deduped = []
    for rd in rows_data:
        lib_key = str(rd['lib_id'] or '').strip()
        lane_key = str(rd['lane_id'] or '').strip()
        if not lib_key:
            continue
        # 只去除同一 lane 内的重复池文库/合并整体；同一文库在不同
        # lane 出现时必须保留。
        key = (lib_key, lane_key)
        if key not in seen:
            seen.add(key)
            deduped.append(rd)

    dup_count = len(rows_data) - len(deduped)
    if dup_count > 0:
        print(f'    去重: {len(rows_data)} -> {len(deduped)} (去除{dup_count}条)')

    # 写入 Pooling 表
    for i, rd in enumerate(deduped):
        row = i + 2
        targets = [
            (COL_B, rd['lib_id']),
            (COL_D, rd['data_amt']),
            (COL_H, rd['lib_type']),
            (COL_Q, rd['contract']),
            (COL_R, rd['customer']),
            (COL_S, rd['remark']),
            (COL_T, rd['lane_id']),
        ]
        for col, val in targets:
            cell = ws_dst.cell(row=row, column=col)
            cell.value = val
            cell.alignment = CENTER


def main(pool_wb=None):
    wb_src = openpyxl.load_workbook(SRC)
    wb_dst = pool_wb if pool_wb is not None else openpyxl.load_workbook(DST_POOL)

    for src_name, dst_name in SHEET_MAP:
        print(f'\n{"─"*50}\n处理: A表[{src_name}] -> Pooling[{dst_name}]')
        migrate_sheet(wb_src[src_name], wb_dst[dst_name], dst_name)

    if pool_wb is None:
        wb_dst.save(DST_POOL)
    print(f'\n{"="*50}\n[DONE] 步骤一完成 -> {DST_POOL}')


if __name__ == '__main__':
    main()
