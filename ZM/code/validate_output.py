"""对生成结果执行不依赖 Excel 公式重算的业务校验。"""

from __future__ import annotations

import re

import openpyxl

from config import DST_PCR, DST_POOL
from pooling_utils import normalized, read_groups, safe_float
from step3_lookup import is_external_library
from step5_summary import is_bare_library
from step7_mark import contains_b_marker


def calculated_volume(ws, row):
    concentration = safe_float(ws.cell(row=row, column=3).value)
    data_amount = safe_float(ws.cell(row=row, column=4).value)
    formula = normalized(ws.cell(row=row, column=5).value)
    match = re.search(r"\*([0-9.]+),3\)", formula)
    if concentration <= 0 or data_amount <= 0 or not match:
        return None
    e_value = round(data_amount * float(match.group(1)), 3)
    return round(e_value / concentration, 3)


def validate_pooling(workbook):
    errors = []
    warnings = []
    group_count = 0
    from config import get_target_sheets
    for face in get_target_sheets():
        ws = workbook[face]
        for group in read_groups(ws):
            group_count += 1
            rows = group["data_rows"]
            first = rows[0]
            lane_ids = {normalized(ws.cell(row=row, column=20).value) for row in rows}
            if len(lane_ids) != 1:
                errors.append(f"{face}!{first}: 同组存在多个laneID {sorted(lane_ids)}")

            for row in rows:
                lib_id = normalized(ws.cell(row=row, column=2).value)
                if is_external_library(lib_id):
                    if ws.cell(row=row, column=3).value != ws.cell(row=row, column=9).value:
                        errors.append(f"{face}!{row}: 外包浓度C/I不一致")
                volume = calculated_volume(ws, row)
                if volume is None:
                    warnings.append(f"{face}!{row}: 缺少有效浓度或数据量，未计算取样体积")
                elif not 0.5 <= volume <= 5:
                    errors.append(f"{face}!{row}: 取样体积{volume}不在0.5～5")

            p_rows = [row for row in rows if ws.cell(row=row, column=16).value is not None]
            if p_rows and len(rows) != 1:
                errors.append(f"{face}!{first}: 已有qPCR结果的文库未单独成组")

            group_name = normalized(ws.cell(row=first, column=1).value)
            if len(rows) == 1:
                if group_name != normalized(ws.cell(row=first, column=2).value):
                    errors.append(f"{face}!{first}: 单文库组A列未使用文库编号")
            else:
                expected_prefix = f"{face}{normalized(ws.cell(row=first, column=20).value)}-"
                if not group_name.startswith(expected_prefix):
                    errors.append(f"{face}!{first}: 组名未使用face+laneID")

            fragments = [
                safe_float(ws.cell(row=row, column=10).value, default=None)
                for row in rows
            ]
            fragments = [value for value in fragments if value is not None and value > 0]
            if fragments:
                expected = round(sum(fragments) / len(fragments), 2)
                if safe_float(ws.cell(row=first, column=15).value, default=None) != expected:
                    errors.append(f"{face}!{first}: O列平均片段不正确")

            has_u = any("纯化" in normalized(ws.cell(row=row, column=21).value) for row in rows)
            has_w_b = any(contains_b_marker(ws.cell(row=row, column=23).value) for row in rows)
            pure = has_u or has_w_b
            status = normalized(ws.cell(row=first, column=7).value)
            # 单文库组 + P列有值 → 已定量（无条件最高优先）
            if len(rows) == 1 and p_rows:
                if status != "已定量":
                    errors.append(f"{face}!{first}: 单文库组已定量应标记已定量但G列为{status!r}")
            elif pure and status != "纯化":
                errors.append(f"{face}!{first}: 应标记纯化但G列为{status!r}")

            if group["summary_row"] is None:
                continue
            special = any(is_bare_library(ws.cell(row=row, column=2).value) for row in rows)
            has_fragment = bool(fragments)
            m_cell = str(ws.cell(row=first, column=13).value or "")
            n_cell = str(ws.cell(row=first, column=14).value or "")
            m_has = m_cell.startswith("=") or safe_float(ws.cell(row=first, column=13).value, default=None) is not None
            n_has = n_cell.startswith("=") or safe_float(ws.cell(row=first, column=14).value, default=None) is not None
            if special or not has_fragment:
                if not m_has:
                    errors.append(f"{face}!{first}: 组浓度应填M列")
            elif not n_has:
                errors.append(f"{face}!{first}: 组浓度应填N列")

            # M/N改为公式后稀释校验由Excel处理

    return group_count, errors, warnings


def validate_downstream(pool_workbook):
    errors = []
    dilution = pool_workbook['文库稀释计算表']
    dilution_rows = []  # (sample_id, status)，含重复：同一文库上多个 lane 各占一条
    quantified_ids = set()
    for row in range(3, dilution.max_row + 1):
        sample_id = normalized(dilution.cell(row=row, column=2).value)
        if sample_id:
            status = normalized(dilution.cell(row=row, column=1).value)
            dilution_rows.append((sample_id, status))
            if status == "已定量":
                quantified_ids.add(sample_id)

    pcr_workbook = openpyxl.load_workbook(DST_PCR, data_only=False)
    pcr_ids = []
    for ws in pcr_workbook.worksheets:
        # ZM 连续结构：主表从第 6 行写起，读到 sheet 末尾（子表格 B 列为空，不会误读）
        for row in range(6, ws.max_row + 1):
            sample_id = normalized(ws.cell(row=row, column=2).value)
            if sample_id:
                pcr_ids.append(sample_id)
                if sample_id in quantified_ids:
                    errors.append(f"{ws.title}!{row}: 已定量记录仍进入qPCR表")
    pcr_workbook.close()

    expected = sum(1 for _sample_id, status in dilution_rows if status != "已定量")
    if len(pcr_ids) != expected:
        errors.append(f"qPCR写入{len(pcr_ids)}条，预期{expected}条")
    return len(dilution_rows), len(pcr_ids), errors


def main(pool_wb=None):
    pool_workbook = openpyxl.load_workbook(DST_POOL, data_only=False)
    group_count, pooling_errors, warnings = validate_pooling(pool_workbook)
    dilution_count, pcr_count, downstream_errors = validate_downstream(pool_workbook)
    if pool_wb is None:
        pool_workbook.close()

    errors = pooling_errors + downstream_errors
    print(f"校验统计: Pooling {group_count}组, 稀释表{dilution_count}条, qPCR {pcr_count}条")
    if errors:
        print(f"[FAILED] {len(errors)}项错误")
        for error in errors[:50]:
            print(f"  - {error}")
        raise SystemExit(1)
    if warnings:
        print(f"[WARNING] {len(warnings)}项来源数据需人工复核")
        for warning in warnings[:50]:
            print(f"  - {warning}")
    print("[PASSED] 关键业务规则校验通过")


if __name__ == "__main__":
    main()
