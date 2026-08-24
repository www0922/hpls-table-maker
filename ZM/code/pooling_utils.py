"""Pooling 表通用工具。

所有步骤只处理 A:W（1:23）业务列，避免模板中的格式化空白列导致
``max_column`` 达到一万多列。组结构约定：

* 多文库组：连续数据行 + 1 个汇总行（B 列为组内数量）；
* 单文库组：1 个数据行，不要求额外间隔行；
* 不同 lane：使用空白行分隔；
* 同一 lane 末尾连续的无汇总数据行，各自视为单文库组。
"""

from __future__ import annotations

from copy import copy
from typing import Iterable

import openpyxl
from openpyxl.styles import Alignment


POOL_MAX_COL = 23
CENTER = Alignment(horizontal="center", vertical="center")


def safe_float(value, default=0.0):
    """将 Excel 值转换为 float；公式或空值返回 default。"""
    if value is None or isinstance(value, str) and value.startswith("="):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def normalized(value):
    return str(value or "").strip()


def is_summary_row(ws, row):
    """识别多文库组汇总行。"""
    b_value = normalized(ws.cell(row=row, column=2).value)
    d_value = ws.cell(row=row, column=4).value
    return b_value.isdigit() and d_value is not None


def is_blank_business_row(ws, row):
    return all(ws.cell(row=row, column=col).value is None for col in range(1, POOL_MAX_COL + 1))


def last_business_row(ws):
    """返回 A:W 范围内最后一个有值的行。"""
    for row in range(ws.max_row, 1, -1):
        if not is_blank_business_row(ws, row):
            return row
    return 1


def read_groups(ws):
    """按统一规则读取 Pooling 组。

    返回结构：
    ``{"data_rows": [...], "summary_row": int|None, "lane_id": str}``
    """
    groups = []
    pending = []

    def flush_single_tail():
        nonlocal pending
        for data_row in pending:
            groups.append(
                {
                    "data_rows": [data_row],
                    "summary_row": None,
                    "lane_id": normalized(ws.cell(row=data_row, column=20).value),
                }
            )
        pending = []

    for row in range(2, last_business_row(ws) + 1):
        if is_blank_business_row(ws, row):
            flush_single_tail()
            continue

        if is_summary_row(ws, row):
            if pending:
                groups.append(
                    {
                        "data_rows": list(pending),
                        "summary_row": row,
                        "lane_id": normalized(ws.cell(row=pending[0], column=20).value),
                    }
                )
            pending = []
            continue

        if ws.cell(row=row, column=2).value is not None:
            pending.append(row)

    flush_single_tail()
    return groups


def snapshot_row(ws, row, max_col=POOL_MAX_COL):
    return {col: ws.cell(row=row, column=col).value for col in range(1, max_col + 1)}


def copy_cell_style(source, target):
    """复制单元格样式，不复制值。"""
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def write_snapshot(ws, row, values, max_col=POOL_MAX_COL):
    for col in range(1, max_col + 1):
        v = values.get(col)
        if isinstance(v, float) and v != int(v):
            v = round(v, 3)
        cell = ws.cell(row=row, column=col)
        cell.value = v
        cell.alignment = CENTER
        cell.font = openpyxl.styles.Font()
        cell.fill = openpyxl.styles.PatternFill()


def clear_business_rows(ws):
    """清除第 2 行起业务数据，解除合并格，并重置所有业务区样式。"""
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= 2 and merged.min_col <= POOL_MAX_COL:
            ws.unmerge_cells(str(merged))
    last_row = last_business_row(ws)
    if last_row >= 2:
        ws.delete_rows(2, last_row - 1)
    # 清除所有残留样式(字体颜色、填充)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=POOL_MAX_COL):
        for cell in row:
            cell.font = openpyxl.styles.Font()
            cell.fill = openpyxl.styles.PatternFill()


def unique_preserving_order(values: Iterable):
    seen = set()
    result = []
    for value in values:
        if value not in seen:
            seen.add(value)
            result.append(value)
    return result
