"""
步骤二：从C表/D表查找并填充 B表(Pooling表)
=============================================
匹配规则:
  HGC-Lib- / HGC-POOL- 开头 → C表 (所有sheet, F列HGC编号匹配)
  其他 (HGC-数字, 数字E, 数字X, 数字R等) → D表 (先Col3杂交编号, 再Col4样本编号)

HGC类型 (C表 → B表):
  L列(Qubit浓度) → C列 + N列
  M列(平均片段)  → O列
  V列(文库结构)  → O列(覆盖平均片段? 不, 文库结构也是O列)
  W列(磷酸化*)   → P列
  X列(环化*)     → Q列

非HGC类型 (D表 → B表):
  K列(Qubit浓度) → C列 + N列
  N列(平均片段)  → O列
"""
import openpyxl
from openpyxl.styles import Alignment
import re

from config import DST, get_src_c, get_src_d, get_target_sheets

CENTER = Alignment(horizontal='center', vertical='center')

# B表目标列 (1-based)
B_C = 3     # Qubit浓度
B_M = 13    # Qubit浓度 (同C列)
B_N = 14    # 片段
B_O = 15    # 人工检测结果
B_P = 16    # 人工备注


def is_hgc_type(lib_id):
    """判断文库编号的HGC类型: 返回 'C' 或 'D'"""
    if lib_id is None:
        return 'D'
    s = str(lib_id).strip()
    if s.startswith('HGC-Lib-') or s.startswith('HGC-POOL-'):
        return 'C'
    return 'D'


def find_header(ws, targets, max_search=5):
    """在ws前max_search行中搜索目标列名, 返回 {col_name: col_index}"""
    found = {t: None for t in targets}
    for row in range(1, min(max_search + 1, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            v = str(ws.cell(row=row, column=col).value or '')
            for t in targets:
                if found[t] is None and t in v:
                    found[t] = col
        if all(v is not None for v in found.values()):
            return found, row
    return found, 1


def build_c_lookup(src):
    """构建C表查找字典: {HGC编号: row_data}, 搜所有sheet"""
    wb = openpyxl.load_workbook(src, data_only=True)
    lookup = {}

    targets = ['HGC编号', 'Qubit浓度', '平均片段', '人工检测结果', '人工备注']

    for sn in wb.sheetnames:
        ws = wb[sn]
        found, header_row = find_header(ws, targets)
        col_f = found['HGC编号']       # F列
        col_l = found['Qubit浓度']     # L列
        col_m = found['平均片段']      # M列 → B表 N列(片段)
        col_z = found['人工检测结果']   # Z列 → B表 O列
        col_aa = found['人工备注']      # AA列 → B表 P列

        if not col_f:
            continue

        print(f'  C表[{sn}]: F(HGC编号)={col_f}, L(Qubit)={col_l}, M(平均片段)={col_m}, Z(人工检测)={col_z}, AA(人工备注)={col_aa}')

        for row in range(header_row + 1, ws.max_row + 1):
            key = ws.cell(row=row, column=col_f).value
            if key is None:
                continue
            key = str(key).strip()
            if not key:
                continue

            def gv(c):
                return ws.cell(row=row, column=c).value if c else None

            lookup[key] = {
                'qubit':  gv(col_l),
                'frag':   gv(col_m),     # → N列(片段)
                'check':  gv(col_z),     # → O列(人工检测)
                'remark': gv(col_aa),    # → P列(人工备注)
            }

    wb.close()
    print(f'  C表共加载 {len(lookup)} 条记录')
    return lookup


def build_d_lookup(src):
    """构建D表查找字典: 双索引 {杂交编号: data, 样本编号: data}"""
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb[wb.sheetnames[0]]  # 自建库出库报告总表

    targets = ['杂交编号', '样本编号', 'Qubit浓度', '平均片段']
    found, header_row = find_header(ws, targets, max_search=3)

    col_d3 = found['杂交编号']    # Col3: 杂交编号/自建库Pooling编号
    col_d4 = found['样本编号']    # Col4: 样本编号
    col_k = found['Qubit浓度']    # K列
    col_n = found['平均片段']     # N列

    print(f'  D表: 杂交编号={col_d3}, 样本编号={col_d4}, Qubit(K)={col_k}, 平均片段(N)={col_n}')

    lookup_jj = {}   # 杂交编号索引
    lookup_yb = {}   # 样本编号索引

    for row in range(header_row + 1, ws.max_row + 1):
        def gv(c):
            return ws.cell(row=row, column=c).value if c else None

        rd = {
            'qubit': gv(col_k),
            'frag':  gv(col_n),
        }

        jj = gv(col_d3)
        if jj:
            lookup_jj[str(jj).strip()] = rd

        yb = gv(col_d4)
        if yb:
            lookup_yb[str(yb).strip()] = rd

    wb.close()
    print(f'  D表共加载: 杂交编号 {len(lookup_jj)} 条, 样本编号 {len(lookup_yb)} 条')
    return lookup_jj, lookup_yb


def fill_sheet(ws_b, c_lookup, d_jj, d_yb, name):
    """填充单个工作表"""
    filled = 0
    missed = 0

    for row in range(2, ws_b.max_row + 1):
        lib_id = ws_b.cell(row=row, column=2).value
        if lib_id is None:
            continue
        lib_id = str(lib_id).strip()

        if is_hgc_type(lib_id) == 'C':
            data = c_lookup.get(lib_id)
            if data:
                write_cell(ws_b, row, B_C, data.get('qubit'))
                write_cell(ws_b, row, B_M, data.get('qubit'))     # M列 = C列
                write_cell(ws_b, row, B_N, data.get('frag'))      # 片段 → N列
                write_cell(ws_b, row, B_O, data.get('check'))     # 人工检测 → O列
                write_cell(ws_b, row, B_P, data.get('remark'))    # 人工备注 → P列
                filled += 1
            else:
                missed += 1
        else:
            # 非HGC: D表, 先杂交编号(Col3)再样本编号(Col4)
            data = d_jj.get(lib_id) or d_yb.get(lib_id)
            if data:
                write_cell(ws_b, row, B_C, data.get('qubit'))
                write_cell(ws_b, row, B_M, data.get('qubit'))     # M列 = C列
                write_cell(ws_b, row, B_N, data.get('frag'))      # 片段 → N列
                filled += 1
            else:
                missed += 1

    print(f'  Sheet {name}面: 匹配={filled}, 未匹配={missed}')


def write_cell(ws, row, col, value):
    """写入单元格并设置居中"""
    if col is None:
        return
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.alignment = CENTER


def main():
    print('构建C表查找字典...')
    c_lookup = build_c_lookup(get_src_c())

    print('构建D表查找字典...')
    d_jj, d_yb = build_d_lookup(get_src_d())

    wb_dst = openpyxl.load_workbook(DST)

    for name in get_target_sheets():
        print(f'\n处理工作表 [{name}]')
        fill_sheet(wb_dst[name], c_lookup, d_jj, d_yb, name)

    wb_dst.save(DST)
    print(f'\n{"="*50}\n[DONE] 步骤二完成 -> {DST}')


if __name__ == '__main__':
    main()
