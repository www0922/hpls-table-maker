"""
步骤四：Lane编号、合并、框线
============================
1. C/D 工作表:
   - A列写入 Lane编号 (C1,C2... / D1,D2...)
   - 组内样本>1时合并A列 (单样本不合并)
   - 数据行 A~H列加框线 (汇总行不加)
   - TE Buffer 汇总行保持湖蓝色底色
"""
import openpyxl
from openpyxl.styles import Alignment, Border, Side
from config import DST, get_target_sheets

CENTER = Alignment(horizontal='center', vertical='center')
ALL_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

COL_A = 1
COL_D = 4
COL_H = 8     # 客户单位
COL_H = 8    # 框线终点 (H列客户单位)


def read_groups_by_summary(ws):
    """读取工作表, 按汇总行识别组: 汇总行B列为数字(条数)且D列有值"""
    groups = []
    current = []
    for row in range(2, ws.max_row + 1):
        b = ws.cell(row=row, column=2).value
        d = ws.cell(row=row, column=4).value

        # 汇总行: B列为数字(数据条数)且D列有值(合计)
        if isinstance(b, (int, float)) and d is not None:
            current.append(row)
            groups.append(current)
            current = []
        else:
            current.append(row)

    if current:
        groups.append(current)
    return groups


def process_sheet(ws, name):
    """处理单个工作表: Lane编号、合并、框线"""
    groups = read_groups_by_summary(ws)

    lane_info = []  # [(lane_name, customer, summary_d), ...]

    for gi, rows in enumerate(groups):
        lane_name = f'{name}{gi + 1}'

        # 数据行范围 (不含汇总行)
        data_rows = rows[:-1] if len(rows) > 1 else rows
        data_first = data_rows[0]
        data_last = data_rows[-1]

        # 取组内第一行的客户单位(H列)
        first_h = ws.cell(row=data_first, column=COL_H).value
        # 汇总D: 对数据行D列求和 (或读汇总行D列)
        summary_d = None
        if len(rows) > 1:
            summary_row = rows[-1]
            summary_d = ws.cell(row=summary_row, column=COL_D).value

        lane_info.append((lane_name, first_h, summary_d))

        # 写入 Lane编号到数据行 A列
        for row in data_rows:
            cell = ws.cell(row=row, column=COL_A)
            cell.value = lane_name
            cell.alignment = CENTER

        # 仅当组内样本数>1时合并A列
        if len(data_rows) > 1:
            ws.merge_cells(
                start_row=data_first, start_column=COL_A,
                end_row=data_last, end_column=COL_A
            )

        # 数据行 A~K列加框线 (汇总行不加)
        for row in data_rows:
            for col in range(COL_A, COL_H + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = ALL_BORDER

    print(f'  Sheet {name}面: {len(groups)} 组, Lane={name}1~{name}{len(groups)}')
    return lane_info


def main():
    wb = openpyxl.load_workbook(DST)

    all_lane_info = []

    for name in get_target_sheets():
        print(f'\n{"─"*50}')
        print(f'处理工作表 [{name}面]')
        print(f'{"─"*50}')
        info = process_sheet(wb[name], name)
        all_lane_info.extend(info)

    wb.save(DST)
    print(f'\n{"="*50}')
    print(f'[DONE] 步骤四完成 -> {DST}')


if __name__ == '__main__':
    main()
