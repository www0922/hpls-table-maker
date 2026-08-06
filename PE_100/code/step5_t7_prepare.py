"""
步骤五：T7+制备 数据填充与公式
===============================
1. 从Pooling表(C面/D面)传递数据到T7+制备
2. A列 laneID = 当天日期 + 面号 (如 0715C)
3. 相同面号为一组, C面/D面之间空一行
4. 填入公式: F, G, I, K, L, M, Q
5. 每面汇总行: H_sum, K_sum, M_sum, P=96-M_sum
"""
import openpyxl
import re
from datetime import datetime
from openpyxl.styles import Alignment, Border, Side, PatternFill
from config import DST, get_target_sheets

CENTER = Alignment(horizontal='center', vertical='center')
ALL_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
LIGHT_BLUE = PatternFill(start_color='99CCFF', end_color='99CCFF', fill_type='solid')

# T7+制备 列索引
COL_A = 1    # laneID
COL_B = 2    # 文库ssDNA编号
COL_C = 3    # SSDNA浓度 (用户填)
COL_D = 4    # 平均片段 (用户填)
COL_E = 5    # 偏差判断
COL_F = 6    # 150fmol对应质量
COL_G = 7    # 投入摩尔数
COL_H = 8    # 数据量
COL_I = 9    # 理论数据量占比
COL_J = 10   # 调整比例 (用户填)
COL_K = 11   # K列
COL_L = 12   # 实际投入占比
COL_M = 13   # ssDNA取样体积
COL_N = 14   # 文库类型
COL_O = 15   # 客户单位
COL_P = 16   # TE补充体积
COL_Q = 17   # 实际调整比例
COL_R = 18   # 转换效率
COL_S = 19   # 环化投入量


def read_pooling_data(wb):
    """从各数据sheet读取Lane分组信息和汇总数据"""
    all_data = []

    for sn in get_target_sheets():
        ws = wb[sn]
        groups = []
        current = []

        for row in range(2, ws.max_row + 1):
            b = ws.cell(row=row, column=2).value
            d = ws.cell(row=row, column=4).value

            # 汇总行: B列为数字(数据条数)且D列有值(合计)
            if isinstance(b, (int, float)) and d is not None:
                current.append(row)
                groups.append(current)
                current = []
            else:
                current.append(row)

        if current:
            groups.append(current)

        for rows in groups:
            data_rows = rows[:-1] if len(rows) > 1 else rows
            summary_row = rows[-1] if len(rows) > 1 else rows[0]

            lane = ws.cell(row=data_rows[0], column=1).value  # A列 Lane编号
            data_g = ws.cell(row=summary_row, column=4).value  # D列 汇总数据量
            lib_type = ws.cell(row=data_rows[0], column=7).value  # G列 文库类型
            customer = ws.cell(row=data_rows[0], column=8).value  # H列 客户单位
            # E列现在是公式, 解析投入量: =D{row}/D{summary_row}*{input}
            e_formula = str(ws.cell(row=data_rows[0], column=5).value or '')
            m = re.search(r'\*(\d+)$', e_formula)
            e_sum = int(m.group(1)) if m else 0

            if lane:
                all_data.append({
                    'lane': str(lane).strip(),
                    'data_g': data_g,
                    'e_sum': e_sum,
                    'lib_type': lib_type,
                    'customer': customer,
                    'sheet': sn,
                })

    return all_data


def fill_t7_sheet(ws_t7, pool_data):
    """填充 T7+制备 工作表"""
    # 清空数据行
    if ws_t7.max_row > 1:
        ws_t7.delete_rows(2, ws_t7.max_row - 1)

    date_str = datetime.now().strftime('%m%d')

    # 按面号分组
    groups = {}  # {sheet_letter: [row_data, ...]}
    for rd in pool_data:
        letter = rd['sheet']  # C or D
        if letter not in groups:
            groups[letter] = []
        groups[letter].append(rd)

    current_row = 2

    for sheet_letter in get_target_sheets():
        if sheet_letter not in groups:
            continue

        # 面间空一行 (第一个面不空)
        if current_row > 2:
            current_row += 1

        lane_id = f'{date_str}{sheet_letter}'
        rows = groups[sheet_letter]
        group_start = current_row
        group_end = current_row + len(rows) - 1

        for rd in rows:
            # A列: laneID
            c = ws_t7.cell(row=current_row, column=COL_A)
            c.value = lane_id
            c.alignment = CENTER

            # B列: 文库ssDNA编号 (Lane编号)
            c = ws_t7.cell(row=current_row, column=COL_B)
            c.value = rd['lane']
            c.alignment = CENTER

            # H列: 数据量
            c = ws_t7.cell(row=current_row, column=COL_H)
            c.value = rd['data_g']
            c.alignment = CENTER

            # N列: 文库类型
            c = ws_t7.cell(row=current_row, column=COL_N)
            c.value = rd['lib_type']
            c.alignment = CENTER

            # O列: 客户单位
            c = ws_t7.cell(row=current_row, column=COL_O)
            c.value = rd['customer']
            c.alignment = CENTER

            # S列: 环化投入量 = 对应组汇总行E列(组文库摩尔质量合计)
            c = ws_t7.cell(row=current_row, column=COL_S)
            c.value = rd.get('e_sum')
            c.alignment = CENTER

            # E列: D与中位值比较
            c = ws_t7.cell(row=current_row, column=COL_E)
            c.value = f'=IF(ABS(D{current_row}-D{group_end + 1})<=100,"正常",IF(D{current_row}>D{group_end + 1},"偏大"&INT(D{current_row}-D{group_end + 1})&"bp","偏小"&INT(D{group_end + 1}-D{current_row})&"bp"))'
            c.alignment = CENTER

            # F列: =D*0.33*G*0.001
            c = ws_t7.cell(row=current_row, column=COL_F)
            c.value = f'=D{current_row}*0.33*G{current_row}*0.001'
            c.alignment = CENTER

            # G列: =900*L
            c = ws_t7.cell(row=current_row, column=COL_G)
            c.value = f'=900*L{current_row}'
            c.alignment = CENTER

            # I列: =ROUND(H/H总,4)
            c = ws_t7.cell(row=current_row, column=COL_I)
            c.value = f'=ROUND(H{current_row}/H{group_end + 1},4)'
            c.alignment = CENTER

            # K列: =I*J
            c = ws_t7.cell(row=current_row, column=COL_K)
            c.value = f'=I{current_row}*J{current_row}'
            c.alignment = CENTER

            # L列: =K/K总
            c = ws_t7.cell(row=current_row, column=COL_L)
            c.value = f'=K{current_row}/K{group_end + 1}'
            c.alignment = CENTER

            # M列: =F/C
            c = ws_t7.cell(row=current_row, column=COL_M)
            c.value = f'=F{current_row}/C{current_row}'
            c.alignment = CENTER

            # Q列: =I*J
            c = ws_t7.cell(row=current_row, column=COL_Q)
            c.value = f'=I{current_row}*J{current_row}'
            c.alignment = CENTER

            # R列: =22*C/S (模板公式, 转换效率)
            c = ws_t7.cell(row=current_row, column=COL_R)
            c.value = f'=22*C{current_row}/S{current_row}'
            c.alignment = CENTER

            current_row += 1

        # 汇总行
        summary_row = current_row

        c = ws_t7.cell(row=summary_row, column=COL_D)
        c.value = f'=MEDIAN(D{group_start}:D{group_end})'
        c.alignment = CENTER

        c = ws_t7.cell(row=summary_row, column=COL_H)
        c.value = f'=SUM(H{group_start}:H{group_end})'
        c.alignment = CENTER

        c = ws_t7.cell(row=summary_row, column=COL_K)
        c.value = f'=SUM(K{group_start}:K{group_end})'
        c.alignment = CENTER

        c = ws_t7.cell(row=summary_row, column=COL_M)
        c.value = f'=SUM(M{group_start}:M{group_end})'
        c.alignment = CENTER

        # P列: =96 - M汇总 (湖蓝色)
        c = ws_t7.cell(row=summary_row, column=COL_P)
        c.value = f'=96-M{summary_row}'
        c.alignment = CENTER
        c.fill = LIGHT_BLUE

        # 数据行加框线
        for r in range(group_start, group_end + 1):
            for col in range(1, 18):
                ws_t7.cell(row=r, column=col).border = ALL_BORDER

        # 数据行行高
        for r in range(group_start, current_row):
            ws_t7.row_dimensions[r].height = 30

        current_row += 1

    print(f'  T7+制备: {len(groups)} 面, 共 {current_row - 2} 行')


def main():
    wb = openpyxl.load_workbook(DST)
    pool_data = read_pooling_data(wb)
    print(f'Pooling数据: {len(pool_data)} 组')

    print(f'\n{"─"*50}')
    print('填充 [T7+制备]')
    print(f'{"─"*50}')
    fill_t7_sheet(wb['T7+制备'], pool_data)

    wb.save(DST)
    print(f'\n{"="*50}')
    print(f'[DONE] 步骤五完成 -> {DST}')


if __name__ == '__main__':
    main()
