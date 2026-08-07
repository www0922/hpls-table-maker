"""
步骤三：B表 分组、排序、公式计算、汇总
=====================================
对B表(PE100_pooling表.xlsx)的C面/D面工作表:

分组规则:
  1. 按文库类型(G列)+客户单位(H列)预分组
  2. 类型判断:
     - 含cDNA → CDNA组: D_sum≤600G, ≤10样本, F max/min≤5
     - 含oligo/Oligo → Oligo组: 最多12样本/组
     - 不满足条件 → 单样本成组
  3. 投入量=450ng; F>8ul → 降为300ng
  4. 公式: E = D/组D合计×投入量, F = E/C
  5. 汇总行: D/E/F合计, TE Buffer=40-F合计(湖蓝色)
"""
import openpyxl
from openpyxl.styles import Alignment, PatternFill
from config import DST, get_target_sheets

CENTER = Alignment(horizontal='center', vertical='center')
LIGHT_BLUE = PatternFill(start_color='99CCFF', end_color='99CCFF', fill_type='solid')

# 列索引 (1-based)
COL_B = 2    # 文库编号
COL_C = 3    # Qubit浓度
COL_D = 4    # 数据量
COL_E = 5    # 文库摩尔质量
COL_F = 6    # 最终取样体积
COL_G = 7    # 文库类型
COL_H = 8    # 客户单位
COL_I = 9    # TE Buffer
COL_O = 14   # 平均片段 (N列)


def safe_float(v):
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def read_data(ws):
    """读取工作表所有数据行, 返回 [{row_data}, ...]"""
    max_col = ws.max_column
    rows = []
    for row in range(2, ws.max_row + 1):
        cells = {}
        for col in range(1, max_col + 1):
            cells[col] = ws.cell(row=row, column=col).value
        rows.append({
            'B': cells.get(COL_B),
            'C': safe_float(cells.get(COL_C)),
            'D': safe_float(cells.get(COL_D)),
            'G': cells.get(COL_G),         # 文库类型
            'H': cells.get(COL_H),         # 客户单位
            'O': cells.get(COL_O),         # 平均片段
            'cells': cells,
            'E': 0.0, 'F': 0.0,            # 计算值
        })
    return rows


def classify_type(lib_type=None, frag=None):
    """分类: M列片段≥300→cdna; 100≤M<300→oligo; M列为空→看G列文本
    只有cdna和oligo两类, G列不含oligo则默认cdna"""
    f = None
    if frag is not None:
        try:
            f = float(frag)
        except (ValueError, TypeError):
            pass
    if f is not None:
        if f >= 300:
            return 'cdna'
        elif f >= 100:
            return 'oligo'
        else:
            return 'cdna'  # <100 归为cdna
    # 无片段数据, 回退到文库类型文本
    lt = str(lib_type or '')
    if 'oligo' in lt.lower():
        return 'oligo'
    return 'cdna'  # 默认cdna


def normalize_lib_type(lib_type):
    """归一化文库类型名称，合并大小写/横线/修饰词等同义变体
    如 '肿瘤-C4-oligo' / '肿瘤c4杂交Oligo文库' / '肿瘤C4-oligo' → 归为同一类
    策略：去掉所有非 ASCII 字符（中文、横线等），只保留字母和数字"""
    import re
    s = str(lib_type or '').strip().lower()
    s = re.sub(r'[^a-z0-9]', '', s)
    return s


def normalize_customer(customer):
    """归一化客户单位名称，合并同属一个单位的别名
    广州基迪奥科技服务有限公司 与 广州奥智生物科技有限公司 视为同一客户"""
    s = str(customer or '').strip()
    if '基迪奥' in s or '奥智' in s:
        return '广州基迪奥/奥智'
    return s


def calc_ef(rows, input_ng):
    """计算组内每行的 E列 和 F列: E = D/合计×投入量, F = E/C"""
    d_sum = sum(r['D'] for r in rows)
    if d_sum == 0:
        for r in rows:
            r['E'] = 0.0
            r['F'] = 0.0
            r['_input'] = 0
        return

    for r in rows:
        r['E'] = round(r['D'] / d_sum * input_ng, 3)
        r['F'] = round(r['E'] / r['C'], 3) if r['C'] > 0 else 0.0
        r['_input'] = input_ng


def has_large_f(rows):
    """检查组内是否有 F > 8 的样本 (触发降投入量到300ng)"""
    return any(r['F'] > 8.0 for r in rows)


def pack_cdna_groups(rows, max_d_sum=600, max_count=10):
    """贪婪打包CDNA组: 每10个一组, D_sum≤600, F ratio≤5
    F ratio = D/C比值 (与D_sum/投入量无关)"""
    if not rows:
        return []

    MAX_F_RATIO = 5
    MIN_C = 0.001

    # 计算D/C比值
    for r in rows:
        r['_dc_ratio'] = r['D'] / r['C'] if r['C'] >= MIN_C else float('inf')

    # 分离C值异常样本 → 单样本成组
    solo = [r for r in rows if r['_dc_ratio'] == float('inf')]
    normal = [r for r in rows if r['_dc_ratio'] != float('inf')]

    # 按D列升序排列
    normal = sorted(normal, key=lambda r: r['D'])

    groups = []
    remaining = list(normal)

    while remaining:
        current = [remaining.pop(0)]
        cur_d_sum = current[0]['D']
        cur_dc_min = current[0]['_dc_ratio']
        cur_dc_max = current[0]['_dc_ratio']

        i = 0
        while i < len(remaining) and len(current) < max_count:
            r = remaining[i]
            new_d_sum = cur_d_sum + r['D']
            new_dc_min = min(cur_dc_min, r['_dc_ratio'])
            new_dc_max = max(cur_dc_max, r['_dc_ratio'])
            # 同时检查 D_sum 和 F ratio
            if new_d_sum <= max_d_sum and new_dc_max / new_dc_min <= MAX_F_RATIO:
                current.append(r)
                cur_d_sum = new_d_sum
                cur_dc_min = new_dc_min
                cur_dc_max = new_dc_max
                remaining.pop(i)
            else:
                i += 1

        calc_ef(current, 450)
        if has_large_f(current):
            calc_ef(current, 300)
        groups.append(current)

    # C值异常的样本单样本成组
    for r in solo:
        calc_ef([r], 450)
        groups.append([r])

    return groups


def try_oligo_group(rows):
    """Oligo分组: 按D/C比值排序, 每≤12个一组, F max/min≤5
    若F>8ul则降投入量至300ng; C≈0或无法满足条件则单样本成组"""
    if not rows:
        return []

    MAX_COUNT = 12
    MAX_F_RATIO = 5
    MIN_C = 0.001  # C值下限, 低于此值无法计算F, 单样本成组

    # 分离C值异常样本 (C≈0), 直接单样本成组
    solo = [r for r in rows if r['C'] < MIN_C]
    normal = [r for r in rows if r['C'] >= MIN_C]

    # 计算D/C比值, 按比值排序(相似比值的放一组以控制F ratio)
    for r in normal:
        r['_dc_ratio'] = r['D'] / r['C']
    normal = sorted(normal, key=lambda r: r['_dc_ratio'])

    groups = []
    remaining = list(normal)

    while remaining:
        current = [remaining.pop(0)]
        cur_dc_min = current[0]['_dc_ratio']
        cur_dc_max = current[0]['_dc_ratio']

        # 贪婪添加: F ratio≤5, 个数≤12
        i = 0
        while i < len(remaining) and len(current) < MAX_COUNT:
            r = remaining[i]
            new_dc_min = min(cur_dc_min, r['_dc_ratio'])
            new_dc_max = max(cur_dc_max, r['_dc_ratio'])
            # F ratio = D/C ratio (D_sum和投入量约掉)
            if new_dc_max / new_dc_min <= MAX_F_RATIO:
                current.append(r)
                cur_dc_min = new_dc_min
                cur_dc_max = new_dc_max
                remaining.pop(i)
            else:
                i += 1

        # 计算E/F, 检查F>8ul
        calc_ef(current, 450)
        if has_large_f(current):
            calc_ef(current, 300)
        groups.append(current)

    # C值异常的样本单样本成组
    for r in solo:
        calc_ef([r], 450)
        groups.append([r])

    return groups


def process_sheet(ws, name):
    """处理单个工作表"""
    all_rows = read_data(ws)
    print(f'  Sheet {name}面: 共 {len(all_rows)} 条数据')

    # 1. 分离 CDNA / Oligo (只有两类), 其余规则不变
    cdna_rows = [r for r in all_rows if classify_type(r['G'], r['O']) == 'cdna']
    oligo_rows = [r for r in all_rows if classify_type(r['G'], r['O']) == 'oligo']

    final_groups = []

    # CDNA: 全部统一类型; C≤15按客户分层; D≤10跨客户合并; 其余按客户分组
    cdna_pre = {}
    for r in cdna_rows:
        if r['D'] <= 10:
            key = '_cross_customer_'
        elif r['C'] <= 15:
            key = (normalize_customer(r['H']), 'C_low')
        else:
            key = normalize_customer(r['H'])
        if key not in cdna_pre:
            cdna_pre[key] = []
        cdna_pre[key].append(r)
    for key, rows in cdna_pre.items():
        sub = pack_cdna_groups(rows)
        final_groups.extend(sub)

    # Oligo: 所有Oligo类型统一分组 (不区分子类型/客户), ≤12个/组, F ratio≤5
    sub = try_oligo_group(oligo_rows)
    final_groups.extend(sub)

    # 每组内按文库编号排序
    for g in final_groups:
        g.sort(key=lambda r: str(r.get('B') or ''))

    # CDNA在前, Oligo在后
    final_groups.sort(key=lambda g: classify_type(g[0].get('G'), g[0].get('O')) == 'oligo')

    oligo_group_count = sum(1 for g in final_groups if classify_type(g[0].get('G'), g[0].get('O')) == 'oligo')
    print(f'    CDNA {len(cdna_pre)}预分组, Oligo {oligo_group_count}组 -> {len(final_groups)} 组')

    # 3. 清空并写回
    max_col = ws.max_column
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    current_row = 2

    for gi, group in enumerate(final_groups):
        d_first = current_row
        d_last = current_row + len(group) - 1
        summary_row = d_last + 1

        # 写入数据行
        for rd in group:
            for col in range(1, max_col + 1):
                cell = ws.cell(row=current_row, column=col)
                if col == COL_E:
                    input_ng = rd.get('_input', 450)
                    cell.value = f'=D{current_row}/D{summary_row}*{input_ng}'  # Excel公式
                elif col == COL_F:
                    cell.value = f'=ROUND(E{current_row}/C{current_row},3)'  # Excel公式
                else:
                    cell.value = rd['cells'].get(col)
                cell.alignment = CENTER
            current_row += 1

        # 汇总行
        d_sum = sum(r['D'] for r in group)
        e_sum = round(sum(r['E'] for r in group), 3)
        data_count = len(group)  # 本组数据条数

        # B列: 当前分组数据条数
        c = ws.cell(row=summary_row, column=COL_B)
        c.value = data_count
        c.alignment = CENTER

        # D列: 组数据量合计 (静态值)
        c = ws.cell(row=summary_row, column=COL_D)
        c.value = d_sum
        c.alignment = CENTER

        # E列: 组文库摩尔质量合计 (Excel公式)
        c = ws.cell(row=summary_row, column=COL_E)
        c.value = f'=ROUND(SUM(E{d_first}:E{d_last}),3)'
        c.alignment = CENTER

        # F列: 组最终取样体积合计 (Excel公式)
        c = ws.cell(row=summary_row, column=COL_F)
        c.value = f'=ROUND(SUM(F{d_first}:F{d_last}),3)'
        c.alignment = CENTER

        # I列: TE Buffer = 40 - F合计 (Excel公式, 湖蓝色)
        c = ws.cell(row=summary_row, column=COL_I)
        c.value = f'=ROUND(40-F{summary_row},3)'
        c.alignment = CENTER
        c.fill = LIGHT_BLUE

        current_row = summary_row + 1

    print(f'    写入完成: 共 {current_row - 2} 行 (含间隔和汇总行)')


def main(pool_wb=None):
    wb = pool_wb if pool_wb is not None else openpyxl.load_workbook(DST)

    for name in get_target_sheets():
        print(f'\n{"─"*50}')
        print(f'处理工作表 [{name}面]')
        print(f'{"─"*50}')
        process_sheet(wb[name], name)

    if pool_wb is None:
        wb.save(DST)
    print(f'\n{"="*50}')
    print(f'[DONE] 步骤三完成 -> {DST}')


if __name__ == '__main__':
    main()
