"""
步骤一：A表 → B表 数据迁移
============================
A表: 江西华大T7+上机-PE100-20260710.xlsx (源)
B表: PE100_pooling表.xlsx (目标, output_result/)

规则:
  Sheet: A表数字sheet→字母 (3→C, 4→D)
  文库编号(B列): 池文库编号(合并格) > HaploX编号(C列) > 样本名称(D列)
  数据量(D列):   J列 × 0.13 (M→G转换); 池文库→合并范围内J列求和×0.13
  文库类型(G列) ← H列
  客户单位(H列) ← L列
  备注(K列)     ← M列
"""
import openpyxl
from openpyxl.styles import Alignment
from collections import defaultdict

from config import DST, get_src_a, sheet_map_a, sync_b_table_sheets

CENTER = Alignment(horizontal='center', vertical='center')

# 列索引 (1-based)
COL_B = 2    # 杂交文库编号 / 文库编号(目标)
COL_C = 3    # HaploX编号
COL_D = 4    # 样本名称 / 数据量G(目标)
COL_G = 7    # 文库类型(目标)
COL_H = 8    # 文库类型(源) / 客户单位(目标)
COL_J = 10   # 预分配数据量M
COL_K = 11   # 备注(目标)
COL_L = 12   # 客户单位(源)
COL_M = 13   # 备注(源)

M_TO_G = 0.13  # M换算成G的系数


def find_pool_ranges(ws):
    """扫描A表工作表中B列和C列的合并单元格, 返回 {row: pool_info}"""
    pools = defaultdict(lambda: {'pool_id': None, 'sum_range': None})

    for mr in ws.merged_cells.ranges:
        for col in (COL_B, COL_C):
            if mr.min_col <= col <= mr.max_col:
                pool_id = ws.cell(row=mr.min_row, column=col).value
                for r in range(mr.min_row, mr.max_row + 1):
                    pools[r]['pool_id'] = pool_id
                    pools[r]['sum_range'] = (mr.min_row, mr.max_row)
    return dict(pools)


def resolve_lib_id(row, ws, pools):
    """解析文库编号: 池文库编号 > HaploX编号 > 样本名称"""
    pool = pools.get(row, {})
    pool_id = pool.get('pool_id')

    if pool_id is not None:
        return str(pool_id)

    haplox = ws.cell(row=row, column=COL_C).value
    if haplox is not None and str(haplox).strip():
        return str(haplox)

    sample = ws.cell(row=row, column=COL_D).value
    if sample is not None and str(sample).strip():
        return str(sample)

    return None


def resolve_data_amount(row, ws, pools, cache):
    """解析数据量: 池文库→求和×0.13, 普通行→J值×0.13"""
    pool = pools.get(row, {})

    if pool.get('pool_id') is not None:
        key = pool['sum_range']
        if key not in cache:
            total = 0
            for r in range(key[0], key[1] + 1):
                v = ws.cell(row=r, column=COL_J).value
                if isinstance(v, (int, float)):
                    total += v
            cache[key] = total
        return round(cache[key] * M_TO_G, 2)

    v = ws.cell(row=row, column=COL_J).value
    raw = v if isinstance(v, (int, float)) else 0
    return round(raw * M_TO_G, 2)


def migrate_sheet(ws_src, ws_dst, name):
    """迁移单个工作表"""
    pools = find_pool_ranges(ws_src)
    sum_cache = {}
    n_data = ws_src.max_row - 1

    pool_count = len(set(p.get('sum_range') for p in pools.values() if p.get('sum_range')))
    print(f'  Sheet {name}: {n_data}条数据, {pool_count}个池文库组')

    # 收集数据
    rows_data = []
    for row in range(2, ws_src.max_row + 1):
        lib_id = resolve_lib_id(row, ws_src, pools)
        data_amt = resolve_data_amount(row, ws_src, pools, sum_cache)
        lib_type = ws_src.cell(row=row, column=COL_H).value
        customer = ws_src.cell(row=row, column=COL_L).value
        remark = ws_src.cell(row=row, column=COL_M).value
        rows_data.append({
            'lib_id': lib_id,
            'data_amt': data_amt,
            'lib_type': lib_type,
            'customer': customer,
            'remark': remark,
        })

    # 去重: 文库编号相同 → 保留第一行
    seen = set()
    deduped = []
    for rd in rows_data:
        key = str(rd['lib_id'] or '').strip()
        if not key:
            continue
        if key not in seen:
            seen.add(key)
            deduped.append(rd)

    dup_count = len(rows_data) - len(deduped)
    if dup_count > 0:
        print(f'    去重: {len(rows_data)} -> {len(deduped)} 行 (去除 {dup_count} 条重复)')

    # B表写入
    for i, rd in enumerate(deduped):
        row = i + 2
        targets = [
            (COL_B, rd['lib_id']),     # 文库编号
            (COL_D, rd['data_amt']),   # 数据量(G)
            (COL_G, rd['lib_type']),   # 文库类型
            (COL_H, rd['customer']),   # 客户单位
            (COL_K, rd['remark']),     # 备注 (K列)
        ]
        for col, val in targets:
            cell = ws_dst.cell(row=row, column=col)
            cell.value = val
            cell.alignment = CENTER


def main():
    sheet_map = sheet_map_a()
    wb_src = openpyxl.load_workbook(get_src_a(), data_only=True)  # data_only: 读缓存值而非公式, 避免外部引用污染
    wb_dst = openpyxl.load_workbook(DST)

    # 动态同步sheet (根据A表)
    sync_b_table_sheets(wb_dst)

    for src_name, dst_name in sheet_map:
        print(f'\n{"─"*50}\n处理: A表[{src_name}] -> B表[{dst_name}面]')
        migrate_sheet(wb_src[src_name], wb_dst[dst_name], dst_name)

    wb_dst.save(DST)
    print(f'\n{"="*50}\n[DONE] 步骤一完成 -> {DST}')


if __name__ == '__main__':
    main()
