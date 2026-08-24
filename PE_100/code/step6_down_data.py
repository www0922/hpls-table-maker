"""
步骤六：填充"下机数据统计模版"子表
=====================================
从 T7+制备 sheet 读取数据，按 (Lane, 文库类型, 客户单位) 分组写入。
QC 相关列（C/H/I/J）留空待人工填写。
"""
import sys
from pathlib import Path
_sys_path = str(Path(__file__).resolve().parent.parent.parent)
if _sys_path not in sys.path:
    sys.path.insert(0, _sys_path)

import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from config import DST, get_target_sheets

T7_SHEET = 'T7+制备'
TEMPLATE_SHEET = '下机数据统计模版'

CENTER = Alignment(horizontal='center', vertical='center')
ALL_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
LIGHT_RED = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
LIGHT_GREEN = PatternFill(start_color='FF92D04F', end_color='FF92D04F', fill_type='solid')

# T7+制备 列索引 (1-based)
T7_A = 1    # laneID          → 提取 lane 字母
T7_B = 2    # 文库ssDNA编号    → 计数用
T7_H = 8    # 数据量(G)        → G列求和
T7_N = 14   # 文库类型         → E列
T7_O = 15   # 客户单位         → F列

# 下机数据统计模版 列索引 (1-based)
DST_A = 1   # FlowCell_ID      → 留空
DST_B = 2   # Lane_Number      → lane 字母
DST_C = 3   # 单lane产出(M)     → 留空
DST_D = 4   # 文库ssDNA编号     → T7 B列
DST_E = 5   # 文库数量          → count
DST_F = 6   # 文库类型          → T7 N列
DST_G = 7   # 客户单位          → T7 O列
DST_H = 8   # 客户需求数据量(G)  → T7 H列 sum
DST_I = 9   # 实际数据量         → 留空
DST_J = 10  # 欠缺数据量(G)      → =I-H
DST_K = 11  # 补充数据比例       → =J/H
DST_L = 12  # 上机模式           → "T7+100"
DST_N = 14  # 操作人员           → 留空
DST_O = 15  # 异常原因           → 留空


def read_group_data(wb):
    """从 T7+制备 逐行取数据, 每条 T7 行对应下机表一行(不聚合)。
    文库条数从数据 sheet 按 lane名 匹配。
    返回 [{lane, lib_type, customer, count, data_sum}, ...]
    """
    data_sheets = get_target_sheets()

    # ── 从数据 sheet 汇总行 B 列: laneID → 文库条数 ──
    lane_counts = {}
    for sn in data_sheets:
        ws = wb[sn]
        gi = 0
        for row in range(2, ws.max_row + 1):
            b_val = ws.cell(row=row, column=2).value
            if isinstance(b_val, (int, float)):
                gi += 1
                lane_id = f'{sn}{gi}'
                lane_counts[lane_id] = int(b_val)

    # ── 从 T7+制备: 逐行读取, 不聚合 ──
    ws_t7 = wb[T7_SHEET]
    rows = []
    for row in range(2, ws_t7.max_row + 1):
        b_val = ws_t7.cell(row=row, column=T7_B).value
        if not b_val or not isinstance(b_val, str) or str(b_val).startswith('='):
            continue
        lane_name = str(b_val).strip()
        lib_type = str(ws_t7.cell(row=row, column=T7_N).value or '').strip()
        customer = str(ws_t7.cell(row=row, column=T7_O).value or '').strip()
        data_g = ws_t7.cell(row=row, column=T7_H).value or 0
        lane_letter = lane_name[0] if lane_name else ''

        rows.append({
            'lane': lane_letter,
            'ssdna_id': lane_name,
            'lib_type': lib_type,
            'customer': customer,
            'count': lane_counts.get(lane_name, 1),
            'data_sum': float(data_g) if isinstance(data_g, (int, float)) else 0.0,
        })

    return rows


def fill_down_data_sheet(wb, rows):
    """填充下机数据统计模版"""
    ws = wb[TEMPLATE_SHEET]

    # 清空数据行和旧合并
    if ws.max_row > 1:
        for mr in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(mr))
        ws.delete_rows(2, ws.max_row - 1)

    current_row = 2
    lane_groups = {}  # {lane_letter: [row_nums]}
    prev_lane = None

    for rd in rows:
        # A: FlowCell_ID (留空)
        c = ws.cell(row=current_row, column=DST_A)
        c.alignment = CENTER

        # B: Lane_Number
        c = ws.cell(row=current_row, column=DST_B)
        c.value = rd['lane']
        c.alignment = CENTER

        # C: 单lane产出(M) (留空)
        c = ws.cell(row=current_row, column=DST_C)
        c.alignment = CENTER

        # D: 文库ssDNA编号
        c = ws.cell(row=current_row, column=DST_D)
        c.value = rd['ssdna_id']
        c.alignment = CENTER

        # E: 文库数量
        c = ws.cell(row=current_row, column=DST_E)
        c.value = rd['count']
        c.alignment = CENTER
        c.number_format = '0'

        # F: 文库类型
        c = ws.cell(row=current_row, column=DST_F)
        c.value = rd['lib_type']
        c.alignment = CENTER

        # G: 客户单位
        c = ws.cell(row=current_row, column=DST_G)
        c.value = rd['customer']
        c.alignment = CENTER

        # H: 客户需求数据量(G)
        c = ws.cell(row=current_row, column=DST_H)
        c.value = round(rd['data_sum'], 2)
        c.alignment = CENTER
        c.number_format = '0.00'

        # I: 实际数据量 (留空, 待QC数据填入)
        c = ws.cell(row=current_row, column=DST_I)
        c.alignment = CENTER
        c.number_format = '0.00'

        # J: =I-H (实际-需求, <0表示欠缺)
        c = ws.cell(row=current_row, column=DST_J)
        c.value = f'=I{current_row}-H{current_row}'
        c.alignment = CENTER
        c.number_format = '0.00'

        # K: =IF(H=0,"",J/H) (欠缺比例)
        c = ws.cell(row=current_row, column=DST_K)
        c.value = f'=IF(H{current_row}=0,"",J{current_row}/H{current_row})'
        c.alignment = CENTER
        c.number_format = '0.00%'

        # L: 上机模式
        c = ws.cell(row=current_row, column=DST_L)
        c.value = 'T7+100'
        c.alignment = CENTER
        c.fill = LIGHT_GREEN

        # N: 操作人员 (留空)
        c = ws.cell(row=current_row, column=DST_N)
        c.alignment = CENTER

        # O: 异常原因 (留空)
        c = ws.cell(row=current_row, column=DST_O)
        c.alignment = CENTER

        # 记录 lane 分组用于合并
        lane = rd['lane']
        if lane not in lane_groups:
            lane_groups[lane] = []
        lane_groups[lane].append(current_row)

        current_row += 1

    # ── 合并单元格 ──
    # A列、B列、C列: 按 lane 合并
    for lane, rows in lane_groups.items():
        if len(rows) > 1:
            ws.merge_cells(start_row=rows[0], start_column=DST_A,
                           end_row=rows[-1], end_column=DST_A)
            ws.merge_cells(start_row=rows[0], start_column=DST_B,
                           end_row=rows[-1], end_column=DST_B)
            ws.merge_cells(start_row=rows[0], start_column=DST_C,
                           end_row=rows[-1], end_column=DST_C)

    # L列: 全表合并 (同一个run)
    total_rows = current_row - 2
    if total_rows > 1:
        ws.merge_cells(start_row=2, start_column=DST_L,
                       end_row=current_row - 1, end_column=DST_L)

    # ── 框线: 数据行 A~O列 ──
    last_data_row = current_row - 1
    for row in range(2, last_data_row + 1):
        for col in range(DST_A, DST_O + 1):
            ws.cell(row=row, column=col).border = ALL_BORDER

    # ── 条件格式: J<0 或 K<0 → 浅红色 ──
    ws.conditional_formatting.add(
        f'J2:J{last_data_row}',
        CellIsRule(operator='lessThan', formula=['0'], fill=LIGHT_RED)
    )
    ws.conditional_formatting.add(
        f'K2:K{last_data_row}',
        CellIsRule(operator='lessThan', formula=['0'], fill=LIGHT_RED)
    )

    print(f'  下机数据统计模版: {len(rows)} 行, {len(lane_groups)} 个 lane')


def main(pool_wb=None):
    wb = pool_wb if pool_wb is not None else openpyxl.load_workbook(DST)

    rows = read_group_data(wb)
    total_libs = sum(r['count'] for r in rows)
    print(f'下机数据: {len(rows)} 个分组, 共 {total_libs} 个文库')

    print(f'\n{"─"*50}')
    print(f'填充 [{TEMPLATE_SHEET}]')
    print(f'{"─"*50}')
    fill_down_data_sheet(wb, rows)

    if pool_wb is None:
        wb.save(DST)
    print(f'\n{"="*50}')
    print(f'[DONE] 步骤六完成 -> {DST}')


if __name__ == '__main__':
    main()
