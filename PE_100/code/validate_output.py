"""PE100 输出校验。

检查项：
  1. 输出文件可正常打开
  2. 数据 sheet 存在且与 A 表匹配
  3. 每个数据 sheet 的分组结构正确
  4. Lane 编号连续
  5. T7+制备 有数据
  6. Sheet 排列顺序正确
"""
import sys
from pathlib import Path
_sys_path = str(Path(__file__).resolve().parent.parent.parent)
if _sys_path not in sys.path:
    sys.path.insert(0, _sys_path)

import openpyxl

from config import DST, get_target_sheets, is_data_sheet
from common.validate_utils import (
    validate_data_sheet,
    check_sheet_order,
    check_t7_sheet,
)


def main(pool_wb=None):
    errors = []
    warnings = []
    total_groups = 0

    # ── 1. 打开输出文件 ──
    try:
        wb = pool_wb if pool_wb is not None else openpyxl.load_workbook(DST)
    except Exception as e:
        print(f'[FAILED] 无法打开输出文件: {e}')
        raise SystemExit(1)

    target_sheets = get_target_sheets()

    # ── 2. 校验每个数据 sheet ──
    for sn in target_sheets:
        if sn not in wb.sheetnames:
            errors.append(f'缺少数据 sheet: [{sn}]')
            continue
        ws = wb[sn]
        sheet_errors, sheet_warnings, groups = validate_data_sheet(ws, sn)
        errors.extend(sheet_errors)
        warnings.extend(sheet_warnings)
        total_groups += groups
        status = 'OK' if not sheet_errors else f'{len(sheet_errors)} errors'
        print(f'  Sheet [{sn}]: {groups} 组  ({status})')

    # ── 3. T7+制备 ──
    if 'T7+制备' in wb.sheetnames:
        t7_errors = check_t7_sheet(wb['T7+制备'])
        errors.extend(t7_errors)
        if not t7_errors:
            print(f'  Sheet [T7+制备]: OK')
    else:
        errors.append('缺少 sheet: [T7+制备]')

    # ── 4. Sheet 顺序 ──
    expected_order = target_sheets + ['T7+制备']
    order_errors = check_sheet_order(wb, expected_order)
    errors.extend(order_errors)

    if pool_wb is None:
        wb.close()

    # ── 5. 报告 ──
    print(f'\n校验统计: {len(target_sheets)} 个数据 sheet, 共 {total_groups} 组')
    if errors:
        print(f'[FAILED] {len(errors)} 项错误:')
        for e in errors[:30]:
            print(f'  - {e}')
        if len(errors) > 30:
            print(f'  ... 还有 {len(errors) - 30} 项')
        raise SystemExit(1)
    if warnings:
        print(f'[WARNING] {len(warnings)} 项提示:')
        for w in warnings[:10]:
            print(f'  - {w}')
    print('[PASSED] 业务校验通过')


if __name__ == '__main__':
    main()
