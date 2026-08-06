"""
清空 Pooling 表内容，只保留表头
操作对象: output_result/ 中的工作副本 (原始模板不会被修改)
根据A表动态调整B表sheet: 缺失的创建, 多余的删除
"""
import sys
from pathlib import Path
_sys_path = str(Path(__file__).resolve().parent.parent.parent)
if _sys_path not in sys.path:
    sys.path.insert(0, _sys_path)

from config import DST, prepare_output, sync_b_table_sheets, is_data_sheet
import openpyxl


def main():
    # 确保模板已复制到 output_result (首次运行复制, 之后跳过)
    prepare_output()

    POOLING = DST

    wb = openpyxl.load_workbook(POOLING)

    # ── 动态同步sheet (根据A表) ──
    print('同步B表sheet (根据A表)...')
    sync_b_table_sheets(wb)

    for sn in wb.sheetnames:
        ws = wb[sn]

        # 解除第2行起的所有合并单元格（表头行1不动）
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row >= 2:
                ws.unmerge_cells(str(mr))

        # 删除第2行起的所有数据行（表头行1不动）
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        print(f'Sheet [{sn}]: 已清空')

    wb.save(POOLING)
    print(f'\n[DONE] 完成 -> {POOLING}')


if __name__ == '__main__':
    main()
