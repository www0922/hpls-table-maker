#!/usr/bin/env python3
"""改造 XP/ZM 项目：所有步骤共享同一个 openpyxl Workbook 实例。"""
import re
import subprocess
import sys
from pathlib import Path


def transform_step(filepath: Path):
    """改造单步骤：main() 接受可选 pool_wb / pcr_wb。"""
    code = filepath.read_text(encoding="utf-8")
    name = filepath.name

    # --- clear_pooling: 不改造 ---
    if name == "clear_pooling.py":
        return

    # --- 1. 函数签名 ---
    if name == "step9_qpcr.py":
        code = code.replace("def main():", "def main(pool_wb=None, pcr_wb=None):")
        # pool_workbook
        code = code.replace(
            "pool_workbook = openpyxl.load_workbook(DST_POOL, data_only=False)",
            "pool_wb_local = pool_wb if pool_wb is not None else openpyxl.load_workbook(DST_POOL, data_only=False)",
        )
        code = code.replace(
            "pool_workbook.close()",
            "if pool_wb is None:\n    pool_wb_local.close()",
        )
        code = code.replace("pool_workbook", "pool_wb_local")
        # pcr_workbook
        code = code.replace(
            "pcr_workbook = openpyxl.load_workbook(DST_PCR)",
            "pcr_wb_local = pcr_wb if pcr_wb is not None else openpyxl.load_workbook(DST_PCR)",
        )
        code = code.replace(
            "pcr_workbook.save(DST_PCR)",
            "if pcr_wb is None:\n    pcr_wb_local.save(DST_PCR)",
        )
        code = code.replace("pcr_workbook", "pcr_wb_local")
    elif name == "format_font.py":
        code = code.replace("def main():", "def main(pool_wb=None, pcr_wb=None):")

        # 找第一个 load_workbook 变量（pool）
        m1 = re.search(r'(\w+)\s*=\s*openpyxl\.load_workbook\(DST_POOL\)', code)
        if m1:
            v = m1.group(1)
            code = code.replace(
                f'{v} = openpyxl.load_workbook(DST_POOL)',
                f'{v}_local = pool_wb if pool_wb is not None else openpyxl.load_workbook(DST_POOL)',
            )
            code = code.replace(f'{v}.save(DST_POOL)', f'if pool_wb is None:\n    {v}_local.save(DST_POOL)')
            code = re.sub(rf'(?<![_\w]){v}(?![_\w])', f'{v}_local', code)

        # 找第二个 load_workbook 变量（pcr）
        m2 = re.search(r'(\w+)\s*=\s*openpyxl\.load_workbook\(DST_PCR\)', code)
        if m2:
            v = m2.group(1)
            code = code.replace(
                f'{v} = openpyxl.load_workbook(DST_PCR)',
                f'{v}_local = pcr_wb if pcr_wb is not None else openpyxl.load_workbook(DST_PCR)',
            )
            code = code.replace(f'{v}.save(DST_PCR)', f'if pcr_wb is None:\n    {v}_local.save(DST_PCR)')
            code = re.sub(rf'(?<![_\w]){v}(?![_\w])', f'{v}_local', code)
    elif name == "validate_output.py":
        code = code.replace("def main():", "def main(pool_wb=None):")
        code = code.replace(
            "pool_workbook = openpyxl.load_workbook(DST_POOL)",
            "pool_workbook = pool_wb if pool_wb is not None else openpyxl.load_workbook(DST_POOL)",
        )
        code = code.replace(
            "pool_workbook.close()",
            "if pool_wb is None:\n        pool_workbook.close()",
        )
    else:
        # 通用步骤：只操作 DST_POOL
        code = code.replace("def main():", "def main(pool_wb=None):")
        m = re.search(r'(\w+)\s*=\s*openpyxl\.load_workbook\(DST_POOL\)', code)
        if m:
            v = m.group(1)
            code = code.replace(
                f'{v} = openpyxl.load_workbook(DST_POOL)',
                f'{v} = pool_wb if pool_wb is not None else openpyxl.load_workbook(DST_POOL)',
            )
            code = code.replace(f'{v}.save(DST_POOL)', f'if pool_wb is None:\n    {v}.save(DST_POOL)')
            if f'{v}.close()' in code:
                code = code.replace(f'{v}.close()', f'if pool_wb is None:\n    {v}.close()')

    filepath.write_text(code, encoding="utf-8")


def transform_run_all(code_dir: Path):
    """改造 run_all.py：统一加载/保存。"""
    run_all = code_dir / "run_all.py"
    code = run_all.read_text(encoding="utf-8")

    # 检测 import 别名
    alias_map = {}
    for line in code.split("\n"):
        if line.startswith("from ") and "import main as " in line:
            parts = line.split("import main as ")
            if len(parts) == 2:
                module = parts[0].split("from ")[1].strip()
                alias = parts[1].strip()
                alias_map[module] = alias

    clear_fn = alias_map.get("clear_pooling", "clear_outputs")
    step1_fn = alias_map.get("step1_migrate", "step1")
    step2_fn = alias_map.get("step2_group_sort", "step2")
    step3_fn = alias_map.get("step3_lookup", "step3")
    step4_fn = alias_map.get("step4_formula", "step4")
    step5_fn = alias_map.get("step5_summary", "step5")
    step6_fn = alias_map.get("step6_lane_name", "step6")
    step7_fn = alias_map.get("step7_mark", "step7")
    step8_fn = alias_map.get("step8_dilution", "step8")
    step_stats_fn = alias_map.get("step_stats_template", "step_stats")
    step9_fn = alias_map.get("step9_qpcr", "step9")
    fmt_fn = alias_map.get("format_font", "format_outputs")
    validate_fn = alias_map.get("validate_output", "validate_outputs")

    # 检测变量名（XP vs ZM）
    has_pool = "DST_POOL" in code or "pool_dst" in code
    pool_var = "DST_POOL" if "DST_POOL" in code else "pool_dst"
    pcr_var = "DST_PCR" if "DST_PCR" in code else "pcr_dst"

    new_main = f'''def main():
    """优化版：统一加载，各步骤共享同一个 Workbook 实例。"""
    import openpyxl

    # 步骤0：清空输出副本
    {clear_fn}()

    pool_wb = openpyxl.load_workbook({pool_var})

    {step1_fn}(pool_wb=pool_wb)
    {step2_fn}(pool_wb=pool_wb)
    {step3_fn}(pool_wb=pool_wb)
    {step4_fn}(pool_wb=pool_wb)
    {step5_fn}(pool_wb=pool_wb)
    {step6_fn}(pool_wb=pool_wb)
    {step7_fn}(pool_wb=pool_wb)
    {step8_fn}(pool_wb=pool_wb)
    pool_wb.save({pool_var})
    pool_wb.close()
    print("\\n[OK] 步骤1~8完成，已保存")

    pool2 = openpyxl.load_workbook({pool_var})
    {step_stats_fn}(pool_wb=pool2)
    pool2.save({pool_var})
    pool2.close()
    print("\\n[OK] 下机数据统计模版完成，已保存")

    pool3 = openpyxl.load_workbook({pool_var})
    pcr_wb = openpyxl.load_workbook({pcr_var})
    {step9_fn}(pool_wb=pool3, pcr_wb=pcr_wb)
    pool3.close()
    pcr_wb.save({pcr_var})
    pcr_wb.close()
    print("\\n[OK] qPCR定量表完成，已保存")

    pool4 = openpyxl.load_workbook({pool_var})
    pcr2 = openpyxl.load_workbook({pcr_var})
    {fmt_fn}(pool_wb=pool4, pcr_wb=pcr2)
    pool4.save({pool_var})
    pcr2.save({pcr_var})
    pool4.close()
    pcr2.close()
    print("\\n[OK] 格式化完成，已保存")

    pool5 = openpyxl.load_workbook({pool_var})
    {validate_fn}(pool_wb=pool5)
    pool5.close()
    print("\\n[OK] 校验通过")

    print("\\n[DONE] 全流程执行完成")'''

    code = re.sub(r'def main\(\).*?(?=\n\nif __name__|\nif __name__|\Z)', new_main, code, flags=re.DOTALL)
    run_all.write_text(code, encoding="utf-8")


def main():
    if len(sys.argv) < 2:
        print("用法: python transform_optimize.py <XP|ZM>")
        sys.exit(1)

    project = sys.argv[1].upper()
    code_dir = Path(f"E:/HP_Project/{project}/code")
    if not code_dir.is_dir():
        print(f"目录不存在: {code_dir}")
        sys.exit(1)

    print(f"\n=== 改造 {project} ===")

    for f in sorted(code_dir.glob("*.py")):
        if f.name != "run_all.py":
            transform_step(f)
            print(f"  ✓ {f.name}")

    transform_run_all(code_dir)

    # 语法检查
    print("\n--- 语法检查 ---")
    errors = []
    py_exe = sys.executable
    for f in sorted(code_dir.glob("*.py")):
        r = subprocess.run([py_exe, "-m", "py_compile", str(f)], capture_output=True, text=True)
        if r.returncode == 0:
            print(f"  ✓ {f.name}")
        else:
            print(f"  ✗ {f.name}")
            errors.append(f"{f.name}: {r.stderr.strip()}")
    if errors:
        print(f"\n{len(errors)} 个文件有语法错误:")
        for e in errors:
            print(f"  - {e}")
    else:
        print("\n全部通过")


if __name__ == "__main__":
    main()
