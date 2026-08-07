#!/usr/bin/env python3
"""批量修复 ZM 项目改造后的缩进问题。"""
import re
import sys
from pathlib import Path


def fix_indent(filepath: Path):
    code = filepath.read_text(encoding="utf-8")
    # 修复 if pool_wb is None: 后面的 save/close 缺少 4 空格缩进
    code = re.sub(r'(if pool_wb is None:)\n(\S)', r'\1\n    \2', code)
    code = re.sub(r'(if pcr_wb is None:)\n(\S)', r'\1\n    \2', code)
    filepath.write_text(code, encoding="utf-8")
    return True


def main():
    code_dir = Path("E:/HP_Project/ZM/code")
    for f in sorted(code_dir.glob("*.py")):
        if f.name not in ("config.py", "pooling_utils.py", "source_lookups.py", "clear_pooling.py", "run_all.py"):
            fix_indent(f)
            print(f"  ✓ {f.name}")

    # 修复 format_font / step9_qpcr 特殊变量名
    for name in ["format_font.py", "step9_qpcr.py"]:
        f = code_dir / name
        if f.is_file():
            fix_indent(f)
            print(f"  ✓ {name} (fixed)")

    # 写 ZM run_all.py
    run_all = code_dir / "run_all.py"
    run_all.write_text('''"""按操作手册要求依次执行完整 ZM 流程（优化版）。"""

from clear_pooling import main as clear_outputs
from format_font import main as format_outputs
from step1_migrate import main as step1
from step2_group_sort import main as step2
from step3_lookup import main as step3
from step4_formula import main as step4
from step5_summary import main as step5
from step6_lane_name import main as step6
from step7_mark import main as step7
from step8_dilution import main as step8
from step_stats_template import main as step_stats
from step9_qpcr import main as step9
from validate_output import main as validate_outputs
from config import DST_POOL, DST_PCR
import openpyxl


def main():
    clear_outputs()

    pool_wb = openpyxl.load_workbook(DST_POOL)
    step1(pool_wb=pool_wb)
    step2(pool_wb=pool_wb)
    step3(pool_wb=pool_wb)
    step4(pool_wb=pool_wb)
    step5(pool_wb=pool_wb)
    step6(pool_wb=pool_wb)
    step7(pool_wb=pool_wb)
    step8(pool_wb=pool_wb)
    pool_wb.save(DST_POOL)
    pool_wb.close()
    print("\\n[OK] 步骤1~8 完成")

    pool2 = openpyxl.load_workbook(DST_POOL)
    step_stats(pool_wb=pool2)
    pool2.save(DST_POOL)
    pool2.close()
    print("[OK] 下机数据统计模版 完成")

    pool3 = openpyxl.load_workbook(DST_POOL)
    pcr_wb = openpyxl.load_workbook(DST_PCR)
    step9(pool_wb=pool3, pcr_wb=pcr_wb)
    pool3.close()
    pcr_wb.save(DST_PCR)
    pcr_wb.close()
    print("[OK] qPCR 定量表 完成")

    pool4 = openpyxl.load_workbook(DST_POOL)
    pcr2 = openpyxl.load_workbook(DST_PCR)
    format_outputs(pool_wb=pool4, pcr_wb=pcr2)
    pool4.save(DST_POOL)
    pcr2.save(DST_PCR)
    pool4.close()
    pcr2.close()
    print("[OK] 格式化 完成")

    pool5 = openpyxl.load_workbook(DST_POOL)
    validate_outputs(pool_wb=pool5)
    pool5.close()
    print("[OK] 校验通过")

    print("\\n[DONE] ZM 全流程执行完成")


if __name__ == "__main__":
    main()
''', encoding="utf-8")
    print("  ✓ run_all.py")

    # 语法检查
    print("\n--- 语法检查 ---")
    import subprocess
    py = sys.executable
    errs = 0
    for f in sorted(code_dir.glob("*.py")):
        r = subprocess.run([py, "-m", "py_compile", str(f)], capture_output=True, text=True)
        if r.returncode == 0:
            print(f"  ✓ {f.name}")
        else:
            print(f"  ✗ {f.name}")
            errs += 1
    if errs:
        print(f"\n{errs} 个文件有语法错误")


if __name__ == "__main__":
    main()
