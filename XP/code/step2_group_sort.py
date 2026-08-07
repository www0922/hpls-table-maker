"""步骤二：按 lane 和业务条件分组排序。

规则优先级：

1. qPCR 历史结果已存在或类似 ``260712P1-126`` 的整体，强制单独成组；
2. HGC-数字、数字E、数字X：同 lane 内仅按文库类型分组；
3. 普通文库：同 lane 内按文库类型、客户单位、备注分组；
4. 备注均含“补测”时，以“补测”作为统一备注键；
5. 多文库组在前，单文库组集中放在该 lane 最后。
"""

from __future__ import annotations

from collections import OrderedDict
import re

import openpyxl
from openpyxl.styles import Border

from config import DST_POOL
from pooling_utils import (
    CENTER,
    POOL_MAX_COL,
    clear_business_rows,
    normalized,
    safe_float,
    snapshot_row,
    write_snapshot,
)
from source_lookups import build_qpcr_lookup


from config import get_target_sheets
SHEETS = get_target_sheets()

COL_B = 2
COL_D = 4
COL_H = 8
COL_P = 16
COL_R = 18
COL_S = 19
COL_T = 20


def is_supplementary(remark):
    return "补测" in normalized(remark)


def is_bare_library(lib_id):
    value = normalized(lib_id)
    return bool(
        re.match(r"^HGC-\d", value, re.IGNORECASE)
        or re.search(r"\d+E$", value, re.IGNORECASE)
        or re.search(r"\d+X$", value, re.IGNORECASE)
    )


def is_merged_whole(lib_id):
    return bool(re.search(r"\d+P\d+-\d+$", normalized(lib_id), re.IGNORECASE))


def collect_rows_by_lane(ws, qpcr_lookup):
    lanes = OrderedDict()
    for row in range(2, ws.max_row + 1):
        lib_id = normalized(ws.cell(row=row, column=COL_B).value)
        if not lib_id:
            continue
        values = snapshot_row(ws, row)
        qpcr_value = qpcr_lookup.get(lib_id)
        if qpcr_value is not None:
            values[COL_P] = qpcr_value
        lane_id = normalized(values.get(COL_T))
        lanes.setdefault(lane_id, []).append(values)
    return lanes


def group_one_lane(rows):
    grouped = OrderedDict()
    forced_singles = []

    for values in rows:
        lib_id = normalized(values.get(COL_B))
        if values.get(COL_P) is not None or is_merged_whole(lib_id):
            forced_singles.append([values])
            continue

        library_type = normalized(values.get(COL_H))
        remark = normalized(values.get(COL_S))
        if is_supplementary(remark):
            supplement_key = "补测"
        elif "加急" in remark or "加测" in remark:
            supplement_key = ""  # 加急/加测视为空备注
        else:
            supplement_key = remark

        if is_bare_library(lib_id):
            key = ("special", library_type, "补测" if is_supplementary(remark) else "")
        else:
            key = ("normal", library_type, normalized(values.get(COL_R)), supplement_key)
        grouped.setdefault(key, []).append(values)

    ordered_groups = [grouped[key] for key in sorted(grouped, key=lambda item: tuple(map(str, item)))]
    for group in ordered_groups:
        group.sort(key=lambda values: normalized(values.get(COL_B)))
    multi_groups = [group for group in ordered_groups if len(group) > 1]
    single_groups = [group for group in ordered_groups if len(group) == 1]
    single_groups.extend(forced_singles)
    return multi_groups + single_groups


def write_lane_groups(ws, lanes):
    clear_business_rows(ws)
    current_row = 2
    lane_items = list(lanes.items())

    for lane_index, (_, groups) in enumerate(lane_items):
        for group in groups:
            for values in group:
                write_snapshot(ws, current_row, values)
                current_row += 1

            if len(group) > 1:
                ws.cell(row=current_row, column=COL_B).value = str(len(group))
                ws.cell(row=current_row, column=COL_D).value = sum(
                    safe_float(values.get(COL_D)) for values in group
                )
                for col in (COL_B, COL_D):
                    ws.cell(row=current_row, column=col).alignment = CENTER
                current_row += 1

        if lane_index < len(lane_items) - 1:
            # lane 间仅保留一行，并移除业务区框线。
            for col in range(1, POOL_MAX_COL + 1):
                ws.cell(row=current_row, column=col).value = None
                ws.cell(row=current_row, column=col).border = Border()
            current_row += 1

    return current_row - 2


def process_sheet(ws, qpcr_lookup, name):
    lanes = collect_rows_by_lane(ws, qpcr_lookup)
    grouped_lanes = OrderedDict(
        (lane_id, group_one_lane(rows)) for lane_id, rows in sorted(lanes.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 0)
    )
    group_count = sum(len(groups) for groups in grouped_lanes.values())
    written = write_lane_groups(ws, grouped_lanes)
    print(f"  Sheet {name}: {len(lanes)}个lane, {group_count}组, 写入{written}行")


def main(pool_wb=None):
    qpcr_lookup = build_qpcr_lookup()
    workbook = pool_wb if pool_wb is not None else openpyxl.load_workbook(DST_POOL)
    for name in SHEETS:
        print(f"\n处理 [{name}]")
        process_sheet(workbook[name], qpcr_lookup, name)
    if pool_wb is None:
        workbook.save(DST_POOL)
    print(f"\n[DONE] 步骤二完成 -> {DST_POOL}")


if __name__ == "__main__":
    main()
