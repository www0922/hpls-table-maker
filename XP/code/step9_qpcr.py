"""步骤九：填充 qPCR 定量表。"""

from __future__ import annotations

from copy import copy
import re

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.styles import Border, Font, Side

from config import DST_PCR, DST_POOL, get_src_a
from pooling_utils import CENTER, normalized, read_groups, safe_float
from step3_lookup import build_external_lookup, build_self_lookup, is_external_library
from step5_summary import is_bare_library, calculated_e_f


RED_BOLD = Font(bold=True, color="FF0000")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
DATA_START = 8
DATA_END = 46  # 8-46 = 39条
ROWS_PER_SHEET = DATA_END - DATA_START + 1


def is_group_name(sample_id):
    return bool(re.match(r"^[AB]\d+(?:-\d+(?:\.\d+)?(?:-\d+)?)?$", normalized(sample_id), re.IGNORECASE))


def is_hgc_prefix(sample_id):
    s = normalized(sample_id).upper()
    return s.startswith("HGC-") or s.startswith("HGC-LIB") or s.startswith("HGC-POOL")


def build_a_table_c_lookup():
    """构建 {HaploX编号(C列) → 样本名称(D列)} 从A表"""
    wb = openpyxl.load_workbook(get_src_a(), data_only=True)
    lookup = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in range(2, ws.max_row + 1):
            c = normalized(ws.cell(row=row, column=3).value)
            d = normalized(ws.cell(row=row, column=4).value)
            if c and d:
                lookup[c] = d
    wb.close()
    print(f"  A表C→D查找: {len(lookup)}条")
    return lookup


def collect_pooling_groups(workbook):
    result = {}
    from config import get_target_sheets
    for face in get_target_sheets():
        ws = workbook[face]
        for group in read_groups(ws):
            rows = group["data_rows"]
            if not rows:
                continue
            first = rows[0]
            sample_id = normalized(ws.cell(row=first, column=1).value)
            if not sample_id:
                continue
            # 从E/F公式反算M/N静态值(稀释后浓度)
            evaluated = [calculated_e_f(ws, row) for row in rows]
            e_sum = sum(e for e, f in evaluated)
            f_sum = sum(f for e, f in evaluated)
            conc = round(e_sum / f_sum, 3) if f_sum > 0 else None
            # M列: 稀释后(>15→12.5)
            m_raw = ws.cell(row=first, column=13).value
            if isinstance(m_raw, str) and m_raw.startswith("="):
                m_val = 12.5 if (conc and conc > 15) else conc
            else:
                m_val = safe_float(m_raw, default=None)
            # N列: 稀释后(>50→45)
            n_raw = ws.cell(row=first, column=14).value
            if isinstance(n_raw, str) and n_raw.startswith("="):
                n_val = 45.0 if (conc and conc > 50) else conc
            else:
                n_val = safe_float(n_raw, default=None)
            result[sample_id] = {
                "o": ws.cell(row=first, column=15).value,
                "k": ws.cell(row=first, column=11).value,
                "m": m_val,
                "n": n_val,
                "data_amount": sum(
                    safe_float(ws.cell(row=row, column=4).value) for row in rows
                ),
            }
    return result


def collect_pending_dilution_rows(workbook):
    ws = workbook['文库稀释计算表']
    rows = []
    for row in range(3, ws.max_row + 1):
        status = normalized(ws.cell(row=row, column=1).value)
        sample_id = normalized(ws.cell(row=row, column=2).value)
        if not sample_id or status == "已定量":
            continue
        rows.append({
            "status": status,
            "sample_id": sample_id,
            "data_amount": ws.cell(row=row, column=8).value,
            "library_type": ws.cell(row=row, column=7).value,
            "contract": ws.cell(row=row, column=18).value,
            "remark": ws.cell(row=row, column=17).value,
            "customer": ws.cell(row=row, column=19).value,
        })
    return rows


def source_values(sample_id, pool_record, external_lookup, self_hybrid, self_sample):
    n_value = None; o_value = None
    if is_group_name(sample_id):
        n_value = pool_record.get("m")
        o_value = pool_record.get("n")
    elif is_external_library(sample_id):
        external = external_lookup.get(sample_id)
        if external:
            n_value = external.get("conc_l")
    else:
        self_record = self_hybrid.get(sample_id) or self_sample.get(sample_id)
        if self_record:
            n_value = self_record.get("qubit")
    return n_value, o_value


def clear_pcr_targets(ws):
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= DATA_START:
            # 保护第 48-50 行的合并区域
            if merged.min_row <= 50 and merged.max_row >= 48:
                continue
            ws.unmerge_cells(str(merged))
    max_row = max(ws.max_row, DATA_END)
    for row in range(DATA_START, max_row + 1):
        # 第 48-50 行整行保留，不清除任何内容
        if 48 <= row <= 50:
            continue
        for col in range(1, 19):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            cell.value = None


def write_record(ws, row, seq, record, pool_record, a_lookup, external_lookup, self_hybrid, self_sample):
    sample_id = record["sample_id"]
    n_value, o_value = source_values(sample_id, pool_record, external_lookup, self_hybrid, self_sample)

    # A列: 序号
    cell = ws.cell(row=row, column=1)
    cell.value = seq
    cell.alignment = CENTER

    # B列: SampleID
    cell = ws.cell(row=row, column=2)
    cell.value = sample_id
    cell.alignment = CENTER

    # C列: HGC开头→A表D列(样本名称); 非HGC→复制B列
    c_value = a_lookup.get(sample_id, sample_id) if is_hgc_prefix(sample_id) else sample_id
    cell = ws.cell(row=row, column=3)
    cell.value = c_value
    cell.alignment = CENTER

    # D列: 数据量
    cell = ws.cell(row=row, column=4)
    cell.value = record["data_amount"] or pool_record.get("data_amount")
    cell.alignment = CENTER

    # E列: 文库类型 (稀释表G列)
    cell = ws.cell(row=row, column=5)
    cell.value = record["library_type"]
    cell.alignment = CENTER

    # F列: 合同编号
    cell = ws.cell(row=row, column=6)
    cell.value = record["contract"]
    cell.alignment = CENTER

    # H列: 片段
    if is_group_name(sample_id):
        h_value = pool_record.get("o")
    elif is_external_library(sample_id) or is_bare_library(sample_id):
        h_value = pool_record.get("k")
    else:
        h_value = pool_record.get("o")
    cell = ws.cell(row=row, column=8)
    cell.value = h_value
    cell.alignment = CENTER

    # I列: =G/H/650×1000000
    cell = ws.cell(row=row, column=9)
    cell.value = f"=G{row}/H{row}/650*1000000"
    cell.alignment = CENTER

    # J列: 纯化标记
    if record["status"] == "纯化":
        cell = ws.cell(row=row, column=10)
        cell.value = "纯化"
        cell.alignment = CENTER
        cell.font = RED_BOLD

    # M列: =N/G
    cell = ws.cell(row=row, column=13)
    cell.value = f"=N{row}/G{row}"
    cell.alignment = CENTER

    # N列: 浓度
    cell = ws.cell(row=row, column=14)
    cell.value = n_value
    cell.alignment = CENTER

    # O列
    cell = ws.cell(row=row, column=15)
    cell.value = o_value
    cell.alignment = CENTER

    # P列: =O/I
    cell = ws.cell(row=row, column=16)
    cell.value = f"=O{row}/I{row}"
    cell.alignment = CENTER

    # Q列: 备注
    cell = ws.cell(row=row, column=17)
    cell.value = record["remark"]
    cell.alignment = CENTER

    # R列: 客户单位
    cell = ws.cell(row=row, column=18)
    cell.value = record["customer"]
    cell.alignment = CENTER

    # 框线 A-I
    for col in range(1, 10):
        ws.cell(row=row, column=col).border = THIN_BORDER


def main():
    pool_workbook = openpyxl.load_workbook(DST_POOL, data_only=False)
    pool_groups = collect_pooling_groups(pool_workbook)
    pending = collect_pending_dilution_rows(pool_workbook)
    a_lookup = build_a_table_c_lookup()
    external_lookup = build_external_lookup()
    self_hybrid, self_sample = build_self_lookup()

    from datetime import datetime
    pcr_workbook = openpyxl.load_workbook(DST_PCR)
    base_title = datetime.now().strftime("%m%d")
    # 第一个子表重命名
    pcr_workbook.worksheets[0].title = base_title
    while len(pending) > len(pcr_workbook.sheetnames) * ROWS_PER_SHEET:
        new_sheet = pcr_workbook.copy_worksheet(pcr_workbook.worksheets[-1])
        new_sheet.title = f"{base_title}-{len(pcr_workbook.sheetnames)}"
        print(f"  已增加sheet：{new_sheet.title}")

    for ws in pcr_workbook.worksheets:
        clear_pcr_targets(ws)

    for index, record in enumerate(pending):
        sheet_index = index // ROWS_PER_SHEET
        row = DATA_START + index % ROWS_PER_SHEET
        seq = 3 + index % ROWS_PER_SHEET  # 每个子表3→41
        ws = pcr_workbook.worksheets[sheet_index]
        pool_record = pool_groups.get(record["sample_id"], {})
        write_record(ws, row, seq, record, pool_record, a_lookup, external_lookup, self_hybrid, self_sample)

    # G3: 本子表最后序号+1
    for idx, ws in enumerate(pcr_workbook.worksheets):
        count = min(len(pending) - idx * ROWS_PER_SHEET, ROWS_PER_SHEET)
        last_seq = 3 + count - 1 if count > 0 else 2  # 3→41
        ws.cell(row=3, column=7).value = last_seq + 1

    pool_workbook.close()
    pcr_workbook.save(DST_PCR)
    print(f"  qPCR定量表: 写入{len(pending)}条")
    print(f"\n[DONE] 步骤九完成 -> {DST_PCR}")


if __name__ == "__main__":
    main()
