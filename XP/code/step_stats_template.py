"""下机数据统计模版填充：从文库稀释计算表按标题匹配抓取数据。

上机前阶段 —— 从稀释表逐行复制已填列，写入 J/K 公式。
QC 相关列（I/M/O 等）及人工列留空，等下机后再补。
"""

from __future__ import annotations

import openpyxl

from config import DST_POOL
from pooling_utils import CENTER

SRC_SHEET = "文库稀释计算表"
DST_SHEET = "下机数据统计模版"

SRC_HEADER_ROW = 2   # 稀释表表头行
SRC_DATA_START = 3    # 稀释表数据起始行
DST_HEADER_ROW = 1    # 统计模版表头行
DST_DATA_START = 2    # 统计模版数据起始行


def _contains(val, keywords):
    """所有 keyword 均在 val 中出现（不区分大小写）。"""
    if val is None:
        return False
    s = str(val).lower()
    return all(kw.lower() in s for kw in keywords)


def find_src_col(ws, keywords, exclude=()):
    """在 src sheet 表头行中按关键词查找列号，排除含 exclude 关键词的列。"""
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=SRC_HEADER_ROW, column=col).value
        if _contains(val, keywords):
            if exclude and _contains(val, exclude):
                continue
            return col
    return None


def build_column_map(src_ws, dst_ws):
    """构建 {dst_col: src_col} 映射，基于标题关键词匹配。"""
    # (模版列号, 关键词列表, 排除关键词)
    rules = [
        (2,  ["sampleid"],           ()),
        (3,  ["qpcr", "浓度"],        ()),
        (5,  ["浓度"],               ("qpcr",)),
        (6,  ["文库类型"],            ()),   # F ← 稀释表G(文库类型)
        (7,  ["数据量"],              ()),   # G ← 稀释表H(数据量G)
        (8,  ["lane"],               ()),
        (12, ["客户单位"],            ()),
    ]

    mapping = {}
    for dst_col, keywords, exclude in rules:
        dst_header = dst_ws.cell(row=DST_HEADER_ROW, column=dst_col).value or ""
        src_col = find_src_col(src_ws, keywords, exclude)
        if src_col is None:
            print(f"  [WARNING] 模版列 {dst_col}({dst_header}) 未在稀释表中找到匹配列")
        else:
            src_header = src_ws.cell(row=SRC_HEADER_ROW, column=src_col).value or ""
            print(f"  {dst_col}: {dst_header} ← [{src_col}] {src_header}")
        mapping[dst_col] = src_col
    return mapping


def main():
    workbook = openpyxl.load_workbook(DST_POOL)
    src_ws = workbook[SRC_SHEET]
    dst_ws = workbook[DST_SHEET]

    # 解除目标区域已有合并格
    for merged in list(dst_ws.merged_cells.ranges):
        if merged.min_row >= DST_DATA_START:
            dst_ws.unmerge_cells(str(merged))

    print(f"从 [{SRC_SHEET}] → [{DST_SHEET}]")
    print("列映射:")
    col_map = build_column_map(src_ws, dst_ws)

    # 统计稀释表数据行数
    src_rows = 0
    for row in range(SRC_DATA_START, src_ws.max_row + 1):
        if src_ws.cell(row=row, column=2).value is not None:
            src_rows += 1
    # 从pooling表构建 SampleID → (组内条数, lane_label) 映射
    from config import get_target_sheets
    from pooling_utils import read_groups
    group_info = {}
    for face in get_target_sheets():
        ws_pool = workbook[face]
        for group in read_groups(ws_pool):
            first = group["data_rows"][0]
            sid = str(ws_pool.cell(row=first, column=1).value or "")
            if sid and sid != "None":
                lane_id = str(ws_pool.cell(row=first, column=20).value or "")
                group_info[sid] = {
                    "count": len(group["data_rows"]),
                    "lane_label": f"{face}{lane_id}",
                }

    print(f"\n稀释表数据行: {src_rows}")

    # 逐行复制
    dst_row = DST_DATA_START
    for src_row in range(SRC_DATA_START, src_ws.max_row + 1):
        sample_id = src_ws.cell(row=src_row, column=2).value
        if sample_id is None:
            continue

        info = group_info.get(str(sample_id or ""), {"count": 1, "lane_label": ""})

        for dst_col, src_col in col_map.items():
            if src_col is None:
                continue
            # H列改为从pooling表取lane_label，不依赖稀释表合并格
            if dst_col == 8:
                src_val = info["lane_label"]
            else:
                src_val = src_ws.cell(row=src_row, column=src_col).value
            cell = dst_ws.cell(row=dst_row, column=dst_col)
            cell.value = src_val
            cell.alignment = CENTER

        # A列: 当天日期
        from datetime import datetime as _dt
        a_cell = dst_ws.cell(row=dst_row, column=1)
        a_cell.value = _dt.now().strftime('%Y/%m/%d')
        a_cell.alignment = CENTER

        # D列: 该组数据条数
        count = info["count"]
        d_cell = dst_ws.cell(row=dst_row, column=4)
        d_cell.value = count
        d_cell.alignment = CENTER

        # E/G/I/J 保留两位小数
        for col in (5, 7, 9, 10):
            dst_ws.cell(row=dst_row, column=col).number_format = '0.00'

        # J列: =I-G (I列人工填写, G列人工填写)
        j_cell = dst_ws.cell(row=dst_row, column=10)
        j_cell.value = f"=I{dst_row}-G{dst_row}"
        j_cell.alignment = CENTER

        # K列: =J/G, 百分比格式
        k_cell = dst_ws.cell(row=dst_row, column=11)
        k_cell.value = f"=J{dst_row}/G{dst_row}"
        k_cell.number_format = '0.00%'
        k_cell.alignment = CENTER

        dst_row += 1

    # H列同lane合并单元格
    merge_start = DST_DATA_START
    prev_label = dst_ws.cell(row=DST_DATA_START, column=8).value
    for r in range(DST_DATA_START + 1, dst_row):
        cur_label = dst_ws.cell(row=r, column=8).value
        if cur_label != prev_label:
            if r - 1 > merge_start:
                dst_ws.merge_cells(start_row=merge_start, start_column=8,
                                   end_row=r - 1, end_column=8)
            merge_start = r
            prev_label = cur_label
    if dst_row - 1 > merge_start:
        dst_ws.merge_cells(start_row=merge_start, start_column=8,
                           end_row=dst_row - 1, end_column=8)

    # 同lane组 A-R 列加框线
    from openpyxl.styles import Border, Side
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    lane_start = DST_DATA_START
    prev = dst_ws.cell(row=DST_DATA_START, column=8).value
    for r in range(DST_DATA_START, dst_row):
        cur = dst_ws.cell(row=r, column=8).value
        if cur != prev:
            for rr in range(lane_start, r):
                for cc in range(1, 19):  # A-R
                    dst_ws.cell(row=rr, column=cc).border = border
            lane_start = r
            prev = cur
    for rr in range(lane_start, dst_row):
        for cc in range(1, 19):
            dst_ws.cell(row=rr, column=cc).border = border

    workbook.save(DST_POOL)
    print(f"\n[DONE] 下机数据统计模版填充完成 -> 写入 {dst_row - DST_DATA_START} 行")


if __name__ == "__main__":
    main()
