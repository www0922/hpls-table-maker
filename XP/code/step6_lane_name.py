"""步骤六：按仪器面、laneID 和组数据量命名。

多文库组：``面号 + laneID + "-" + 数据量合计``；
单文库组：直接使用文库编号；
同一 sheet 内重名时从第一个开始依次增加 ``-1``、``-2``。
"""

from __future__ import annotations

from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Border, Side, PatternFill

from config import DST_POOL
from pooling_utils import CENTER, POOL_MAX_COL, normalized, read_groups, safe_float


from config import get_target_sheets
SHEETS = get_target_sheets()
COL_A = 1
COL_B = 2
COL_D = 4
COL_L = 12
COL_T = 20

ALL_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def format_amount(value):
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.3f}".rstrip("0").rstrip(".")


def _alt_fill_plate(ws, data_rows):
    """板号交替填充：同一组内不同板号前缀交替浅蓝（L列）。"""
    blue = PatternFill(start_color="FFADD8E6", end_color="FFADD8E6", fill_type="solid")

    def prefix_of(plate):
        if "_" in plate:
            return plate.split("_")[0]
        if "-" in plate:
            return plate.rsplit("-", 1)[0]
        return plate

    seen = []
    for row in data_rows:
        plate = str(ws.cell(row=row, column=COL_L).value or "").strip()
        if not plate:
            continue
        prefix = prefix_of(plate)
        if prefix not in seen:
            seen.append(prefix)
    if len(seen) < 2:
        return

    fill_prefixes = {p for i, p in enumerate(seen) if i % 2 == 1}
    for row in data_rows:
        plate = str(ws.cell(row=row, column=COL_L).value or "").strip()
        if not plate:
            continue
        if prefix_of(plate) in fill_prefixes:
            ws.cell(row=row, column=COL_L).fill = blue


def process_sheet(ws, face):
    for merged in list(ws.merged_cells.ranges):
        if merged.min_col <= COL_A <= merged.max_col and merged.min_row >= 2:
            ws.unmerge_cells(str(merged))

    groups = read_groups(ws)
    bases = []
    for group in groups:
        data_rows = group["data_rows"]
        if len(data_rows) == 1:
            bases.append(normalized(ws.cell(row=data_rows[0], column=COL_B).value))
            continue
        lane_id = normalized(ws.cell(row=data_rows[0], column=COL_T).value)
        data_sum = sum(safe_float(ws.cell(row=row, column=COL_D).value) for row in data_rows)
        bases.append(f"{face}{lane_id}-{format_amount(data_sum)}")

    total_counts = Counter(bases)
    used_counts = defaultdict(int)

    for group, base in zip(groups, bases):
        data_rows = group["data_rows"]
        if len(data_rows) > 1 and total_counts[base] > 1:
            used_counts[base] += 1
            group_name = f"{base}-{used_counts[base]}"
        else:
            group_name = base

        first_row, last_row = data_rows[0], data_rows[-1]
        for row in data_rows:
            cell = ws.cell(row=row, column=COL_A)
            cell.value = group_name
            cell.alignment = CENTER
            for col in range(1, 7):  # A-F
                ws.cell(row=row, column=col).border = ALL_BORDER

        if len(data_rows) > 1:
            ws.merge_cells(
                start_row=first_row,
                start_column=COL_A,
                end_row=last_row,
                end_column=COL_A,
            )

        summary_row = group["summary_row"]
        if summary_row is not None:
            for col in range(1, 7):
                ws.cell(row=summary_row, column=col).border = Border()

        _alt_fill_plate(ws, data_rows)

    print(f"  Sheet {face}: {len(groups)}组")


def main(pool_wb=None):
    workbook = pool_wb if pool_wb is not None else openpyxl.load_workbook(DST_POOL)
    for face in SHEETS:
        print(f"\n处理 [{face}]")
        process_sheet(workbook[face], face)
    if pool_wb is None:
        workbook.save(DST_POOL)
    print(f"\n[DONE] 步骤六完成 -> {DST_POOL}")


if __name__ == "__main__":
    main()
