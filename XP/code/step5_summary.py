"""步骤五：汇总、平均片段、组浓度和补水体积。

* 汇总行 D/E/F 求和，G=E/F；
* 组内 K 列平均片段写到组内第一行 O 列；
* 普通组 K 有非零值：组浓度写 N 第一行，否则写 M 第一行；
* HGC-数字、数字E、数字X组固定写 M 第一行；
* N>50 时稀释：N第一行=G/D（40≤值<50），M第二行=F×(D-1)；
* M>15 时稀释：M第一行=G/D（10≤值<15），M第二行=F×(D-1)；
* 不需要稀释时 M/N 第一行直接引用 G 列，M 第二行清空。
"""

from __future__ import annotations

import re

import openpyxl

from config import DST_POOL
from openpyxl.styles import PatternFill
from pooling_utils import CENTER, normalized, read_groups, safe_float


from config import get_target_sheets
SHEETS = get_target_sheets()

COL_B = 2
COL_D = 4
COL_E = 5
COL_F = 6
COL_G = 7
COL_K = 11
COL_M = 13
COL_N = 14
COL_O = 15

TARGET_N = 45.0       # N列稀释目标浓度
TARGET_M = 12.5      # M列稀释目标浓度
THRESHOLD_N = 50     # N列稀释阈值
THRESHOLD_M = 15     # M列稀释阈值


def is_bare_library(lib_id):
    value = normalized(lib_id)
    if value.upper().startswith("HGC-LIB") or value.upper().startswith("HGC-POOL"):
        return False
    return bool(
        re.match(r"^HGC-\d", value, re.IGNORECASE)
        or re.search(r"\d+E$", value, re.IGNORECASE)
        or re.search(r"\d+X$", value, re.IGNORECASE)
    )


def write_formula(ws, row, col, formula):
    cell = ws.cell(row=row, column=col)
    cell.value = formula
    cell.alignment = CENTER


def calculated_e_f(ws, row):
    concentration = safe_float(ws.cell(row=row, column=3).value)
    data_amount = safe_float(ws.cell(row=row, column=COL_D).value)
    e_value = ws.cell(row=row, column=COL_E).value
    if isinstance(e_value, (int, float)):
        calculated_e = float(e_value)
    elif isinstance(e_value, str):
        match = re.search(r"\*([0-9.]+),3\)", e_value)
        calculated_e = round(data_amount * float(match.group(1)), 3) if match else 0.0
    else:
        calculated_e = 0.0
    calculated_f = round(calculated_e / concentration, 3) if concentration > 0 else 0.0
    return calculated_e, calculated_f


def process_sheet(ws, name):
    groups = read_groups(ws)
    summary_count = 0
    dilution_count = 0

    for group in groups:
        data_rows = group["data_rows"]
        summary_row = group["summary_row"]
        if not data_rows:
            continue

        first_row = data_rows[0]
        last_row = data_rows[-1]

        fragment_values = [
            safe_float(ws.cell(row=row, column=COL_K).value, default=None)
            for row in data_rows
        ]
        fragment_values = [value for value in fragment_values if value is not None and value > 0]
        if fragment_values:
            average_fragment = round(sum(fragment_values) / len(fragment_values), 2)
            cell = ws.cell(row=first_row, column=COL_O)
            cell.value = average_fragment
            cell.alignment = CENTER

        # 单文库组没有间隔汇总行，因此不执行组浓度及补水计算。
        if summary_row is None:
            continue

        summary_count += 1
        write_formula(ws, summary_row, COL_D, f"=SUM(D{first_row}:D{last_row})")
        write_formula(ws, summary_row, COL_E, f"=SUM(E{first_row}:E{last_row})")
        write_formula(ws, summary_row, COL_F, f"=SUM(F{first_row}:F{last_row})")
        write_formula(ws, summary_row, COL_G, f"=ROUND(E{summary_row}/F{summary_row},3)")

        evaluated = [calculated_e_f(ws, row) for row in data_rows]
        e_sum = sum(item[0] for item in evaluated)
        f_sum = sum(item[1] for item in evaluated)
        group_conc = round(e_sum / f_sum, 3) if f_sum > 0 else 0

        has_fragment = bool(fragment_values)
        special_group = any(
            is_bare_library(ws.cell(row=row, column=COL_B).value) for row in data_rows
        )
        target_col = COL_N if has_fragment and not special_group else COL_M

        # 稀释判定与公式写入
        if target_col == COL_N and group_conc > THRESHOLD_N:
            # N列稀释: G/D, 40 ≤ G/D < 50
            d_value = round(group_conc / TARGET_N, 2)
            ws.cell(row=first_row, column=COL_M).value = None
            cell = ws.cell(row=first_row, column=COL_N)
            cell.value = f"=G{summary_row}/{d_value}"
            cell.alignment = CENTER
            cell.number_format = '0.00'
            if len(data_rows) >= 2:
                second_row = data_rows[1]
                water_cell = ws.cell(row=second_row, column=COL_M)
                water_cell.value = f"=F{summary_row}*({d_value}-1)"
                water_cell.alignment = CENTER
                water_cell.number_format = '0.00'
                water_cell.fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")
            dilution_count += 1
        elif target_col == COL_M and group_conc > THRESHOLD_M:
            # M列稀释: G/D, 10 ≤ G/D < 15
            d_value = round(group_conc / TARGET_M, 2)
            ws.cell(row=first_row, column=COL_N).value = None
            cell = ws.cell(row=first_row, column=COL_M)
            cell.value = f"=G{summary_row}/{d_value}"
            cell.alignment = CENTER
            cell.number_format = '0.00'
            if len(data_rows) >= 2:
                second_row = data_rows[1]
                water_cell = ws.cell(row=second_row, column=COL_M)
                water_cell.value = f"=F{summary_row}*({d_value}-1)"
                water_cell.alignment = CENTER
                water_cell.number_format = '0.00'
                water_cell.fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")
            dilution_count += 1
        else:
            # 不稀释，保持原公式
            if target_col == COL_N:
                ws.cell(row=first_row, column=COL_M).value = None
                cell = ws.cell(row=first_row, column=COL_N)
                cell.value = f"=G{summary_row}"
                cell.alignment = CENTER
                cell.number_format = '0.00'
            else:
                ws.cell(row=first_row, column=COL_N).value = None
                cell = ws.cell(row=first_row, column=COL_M)
                cell.value = f"=G{summary_row}"
                cell.alignment = CENTER
                cell.number_format = '0.00'
            # 清空M列第二行
            if len(data_rows) >= 2:
                ws.cell(row=data_rows[1], column=COL_M).value = None

    print(
        f"  Sheet {name}: {len(groups)}组, {summary_count}个多文库汇总, "
        f"{dilution_count}组写入稀释公式"
    )


def main(pool_wb=None):
    workbook = pool_wb if pool_wb is not None else openpyxl.load_workbook(DST_POOL)
    for name in SHEETS:
        print(f"\n处理 [{name}]")
        process_sheet(workbook[name], name)
    if pool_wb is None:
        workbook.save(DST_POOL)
    print(f"\n[DONE] 步骤五完成 -> {DST_POOL}")


if __name__ == "__main__":
    main()
