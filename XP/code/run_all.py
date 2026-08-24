"""按操作手册要求依次执行完整 XP 流程（优化版：减少 Excel 读写次数）。"""

from clear_pooling import main as clear_outputs
from format_font import main as format_outputs
from step1_migrate import main as step1
from step2_group_sort import main as step2
from step3_lookup import main as step3
from step4_formula import main as step4
from step5_summary import main as step5
from step6_lane_name import main as step6
from step7_mark import main as step7
from step8_dilution import main as step8
from step_stats_template import main as step_stats
from step9_qpcr import main as step9
from validate_output import main as validate_outputs
from config import DST_POOL, DST_PCR
import openpyxl


def main():
    """优化版：统一加载，各步骤共享同一个 Workbook 实例。"""

    # 步骤0：清空输出副本（独立运行，负责文件创建）
    clear_outputs()

    # ── 加载 Pooling 工作簿，步骤1~8 共享 ──
    pool_wb = openpyxl.load_workbook(DST_POOL)
    step1(pool_wb=pool_wb)
    step2(pool_wb=pool_wb)
    step3(pool_wb=pool_wb)
    step4(pool_wb=pool_wb)
    step5(pool_wb=pool_wb)
    step6(pool_wb=pool_wb)
    step7(pool_wb=pool_wb)
    step8(pool_wb=pool_wb)
    step_stats(pool_wb=pool_wb)
    pool_wb.save(DST_POOL)
    pool_wb.close()
    print("\n[OK] 步骤1~8 + 下机统计 完成")

    # 步骤9: qPCR 定量表
    pool3 = openpyxl.load_workbook(DST_POOL)
    pcr_wb = openpyxl.load_workbook(DST_PCR)
    step9(pool_wb=pool3, pcr_wb=pcr_wb)
    pool3.close()
    pcr_wb.save(DST_PCR)
    pcr_wb.close()
    print("[OK] qPCR 定量表 完成")

    # 格式化
    pool4 = openpyxl.load_workbook(DST_POOL)
    pcr2 = openpyxl.load_workbook(DST_PCR)
    format_outputs(pool_wb=pool4, pcr_wb=pcr2)
    pool4.save(DST_POOL)
    pcr2.save(DST_PCR)
    pool4.close()
    pcr2.close()
    print("[OK] 格式化 完成")

    # 校验
    pool5 = openpyxl.load_workbook(DST_POOL)
    validate_outputs(pool_wb=pool5)
    pool5.close()
    print("[OK] 校验通过")

    print("\n[DONE] XP 全流程执行完成")


if __name__ == "__main__":
    main()
