"""步骤四：计算 E/F 列，使每条取样体积落在 0.5～5 μL。

优先为同组文库使用统一调整数；如果同组不存在共同可行调整数，
则按行选择调整数，使该行取样体积为 2 μL。
"""

from __future__ import annotations

import openpyxl

from config import DST_POOL
from pooling_utils import CENTER, read_groups, safe_float


from config import get_target_sheets
SHEETS = get_target_sheets()
COL_C = 3
COL_D = 4
COL_E = 5
COL_F = 6


def common_multiplier(records):
    lower = 0.0
    upper = float("inf")
    valid_count = 0
    for concentration, data_amount in records:
        if concentration <= 0 or data_amount <= 0:
            continue
        valid_count += 1
        lower = max(lower, 0.5 * concentration / data_amount)
        upper = min(upper, 5.0 * concentration / data_amount)
    # ΣF > 7 → M × Σ(D/C) > 7 (仅多记录组)
    if len(records) > 1:
        d_over_c = sum(d / c for c, d in records if c > 0 and d > 0)
        if d_over_c > 0:
            lower = max(lower, 7.0 / d_over_c)
    if not valid_count or lower > upper:
        return None
    return round(lower, 6)


def process_sheet(ws, name):
    groups = read_groups(ws)
    max_col = ws.max_column

    # 快照所有行原始数据
    snapshots = {r: {c: ws.cell(row=r, column=c).value for c in range(1, max_col + 1)} for g in groups for r in g["data_rows"]}

    # 计算每组multiplier, 无统一M则拆组
    plan = []  # [(is_single, [snapshot_keys])]
    singles_count = 0
    for group in groups:
        data_rows = group["data_rows"]
        records = [(safe_float(snapshots[r].get(COL_C)), safe_float(snapshots[r].get(COL_D))) for r in data_rows]
        m = common_multiplier(records)
        if m is None and len(data_rows) > 1:
            for r in data_rows:
                c = safe_float(snapshots[r].get(COL_C))
                d = safe_float(snapshots[r].get(COL_D))
                mm = common_multiplier([(c, d)])
                plan.append((True, [r], mm, None))
            singles_count += len(data_rows)
        else:
            plan.append((False, data_rows, m, group["summary_row"]))

    # 尝试将拆出的单样本重组成多文库组
    multi_plan = [p for p in plan if not p[0]]  # 已有的多文库组
    single_plan = [p for p in plan if p[0]]     # 拆出的单样本

    # 按(H,R)分桶
    singles_by_key = {}
    for p in single_plan:
        r = p[1][0]
        sn = snapshots[r]
        key = (str(sn.get(8) or ''), str(sn.get(18) or ''), str(sn.get(20) or ''))
        singles_by_key.setdefault(key, []).append(p)

    new_plan = list(multi_plan)
    merged_count = 0
    for key, singles in singles_by_key.items():
        used = [False] * len(singles)
        for i in range(len(singles)):
            if used[i]:
                continue
            current_rows = list(singles[i][1])
            current_recs = [(safe_float(snapshots[r].get(COL_C)), safe_float(snapshots[r].get(COL_D))) for r in current_rows]
            used[i] = True
            # 全量扫描同key内所有未用的单样本
            changed = True
            while changed:
                changed = False
                for j in range(len(singles)):
                    if used[j]:
                        continue
                    nxt_rows = singles[j][1]
                    cand_recs = current_recs + [(safe_float(snapshots[r].get(COL_C)), safe_float(snapshots[r].get(COL_D))) for r in nxt_rows]
                    if common_multiplier(cand_recs) is not None:
                        current_rows.extend(nxt_rows)
                        current_recs = cand_recs
                        used[j] = True
                        changed = True
            if len(current_rows) > 1:
                m = common_multiplier(current_recs)
                new_plan.append((False, current_rows, m, None))
                merged_count += len(current_rows) - 1
            else:
                new_plan.append(singles[i])

    plan = new_plan
    # 按 laneID 排序，小的在前；同 lane 内多文库组优先，单文库组在后
    plan.sort(key=lambda p: (int(str(snapshots[p[1][0]].get(20) or '0')), len(p[1]) == 1))
    if merged_count:
        print(f"    单样本重合并: {merged_count}个")

    # 清空并重写
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    current_row = 2

    for gi, (is_single, rows, mm, summary_row) in enumerate(plan):
        d_sum = 0
        for r in rows:
            sn = snapshots[r]
            for col in range(1, max_col + 1):
                cell = ws.cell(row=current_row, column=col)
                if col == COL_E:
                    cell.value = f"=ROUND(D{current_row}*{mm:.6f},3)" if mm else None
                elif col == COL_F:
                    cell.value = f"=ROUND(E{current_row}/C{current_row},3)" if mm else None
                else:
                    cell.value = sn.get(col)
                cell.alignment = CENTER
            d_sum += safe_float(sn.get(COL_D))
            current_row += 1

        # 多文库组写汇总行（汇总行即组尾标记，无需空行）
        if not is_single and len(rows) > 1:
            ws.cell(row=current_row, column=2).value = str(len(rows))
            ws.cell(row=current_row, column=2).alignment = CENTER
            ws.cell(row=current_row, column=4).value = d_sum
            ws.cell(row=current_row, column=4).alignment = CENTER
            current_row += 1
        elif gi < len(plan) - 1:
            # 单文库组无汇总行，必须靠空行界定组尾（read_groups 分支①）；
            # 末组由 flush_single_tail 兜底，不留尾部空行。
            current_row += 1

    print(f"  Sheet {name}: {len(plan)}组, {singles_count}个拆分为单样本")


def main(pool_wb=None):
    workbook = pool_wb if pool_wb is not None else openpyxl.load_workbook(DST_POOL)
    for name in SHEETS:
        print(f"\n处理 [{name}]")
        process_sheet(workbook[name], name)
    if pool_wb is None:
        workbook.save(DST_POOL)
    print(f"\n[DONE] 步骤四完成 -> {DST_POOL}")


if __name__ == "__main__":
    main()
