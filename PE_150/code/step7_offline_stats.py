"""
步骤七：生成下机数据统计模版
============================
从 T7+制备 sheet 聚合数据，填入「下机数据统计模版」sheet。

数据来源（T7+制备）：
  A列(laneID)   → 提取末字母 → B列(Lane_Number)
  B列(ssDNA编号) → 同组计数    → D列(文库总数)
  N列(文库类型)  → 去重取值    → E列(文库类型)
  O列(客户单位)  → 取第一个    → F列(客户单位)
  H列(数据量G)   → 同组求和    → G列(需求数据量)

聚合维度：lane + 文库类型 + 客户单位
跳过 T7+制备的合计行（A列为空）。
A/C/H/L/M/N 列留空（下机后/人工填写）。
I列公式 =H-G, J列公式 =I/G, K列填 PE150。
"""
import openpyxl
from openpyxl.styles import Alignment, PatternFill

from config import DST

CENTER = Alignment(horizontal='center', vertical='center')
LIGHT_GREEN = PatternFill(start_color='FF92D04F', end_color='FF92D04F', fill_type='solid')

# ── 下机数据统计模版 列索引 (1-based) ──
T_A = 1   # FlowCell_ID（留空）
T_B = 2   # Lane_Number
T_C = 3   # 单lane产出(M)（留空）
T_D = 4   # 文库ssDNA编号
T_E = 5   # 文库总数
T_F = 6   # 文库类型
T_G = 7   # 客户单位
T_H = 8   # 需求数据量(G)
T_I = 9   # 产出数据量（留空）
T_J = 10  # 欠缺数据量(G) =I-H
T_K = 11  # 数据量补测率 =J/H
T_L = 12  # 上机模式
T_M = 13  # 调整比例（留空）
T_N = 14  # 操作人员（留空）
T_O = 15  # 分析原因（留空）

# ── T7+制备 列索引 ──
T7_A = 1   # laneID
T7_B = 2   # 文库ssDNA编号
T7_H = 8   # 数据量G
T7_N = 14  # 文库类型
T7_O = 15  # 客户单位


def build_lib_count_lookup(wb):
    """从数据 sheet 的汇总行 B 列读取每组文库条数, 返回 {laneID: count}"""
    from config import is_data_sheet
    lookup = {}
    for sn in wb.sheetnames:
        if not is_data_sheet(sn):
            continue
        ws = wb[sn]
        # 按顺序枚举 lane: B1, B2, B3... / C1, C2... / D1, D2...
        gi = 0
        for row in range(2, ws.max_row + 1):
            b_val = ws.cell(row=row, column=2).value
            if isinstance(b_val, (int, float)):
                # 汇总行: B列为数字(文库条数)
                gi += 1
                lane_id = f'{sn}{gi}'
                lookup[lane_id] = int(b_val)
    return lookup


def main(pool_wb=None):
    wb = pool_wb if pool_wb is not None else openpyxl.load_workbook(DST)

    ws_t7 = wb['T7+制备']
    ws_dst = wb['下机数据统计模版']

    # ── 1. 从数据sheet获取文库条数 ──
    lib_count = build_lib_count_lookup(wb)
    print(f'  数据sheet文库计数: {len(lib_count)} 个 lane')

    # ── 2. 从 T7+制备 逐行复制（不去重） ──
    rows = []
    for row in range(2, ws_t7.max_row + 1):
        lane_id = ws_t7.cell(row=row, column=T7_A).value
        if lane_id is None or str(lane_id).strip() == '':
            continue  # 跳过合计行
        lane_id = str(lane_id).strip()
        lane_letter = lane_id[-1]  # 末字母
        ssdna_id = ws_t7.cell(row=row, column=T7_B).value  # 如 B1
        lib_name = ws_t7.cell(row=row, column=T7_N).value or ''
        customer = ws_t7.cell(row=row, column=T7_O).value or ''
        data_g = ws_t7.cell(row=row, column=T7_H).value
        try:
            data_g = float(data_g) if data_g is not None else 0.0
        except (ValueError, TypeError):
            data_g = 0.0

        count = lib_count.get(str(ssdna_id).strip() if ssdna_id else '', 1)

        rows.append({
            'lane': lane_letter,
            'ssdna_id': str(ssdna_id).strip() if ssdna_id else '',
            'lib_type': str(lib_name).strip(),
            'customer': str(customer).strip(),
            'data_g': data_g,
            'count': count,
        })

    print(f'  T7+制备: {len(rows)} 条数据')
    for lane in sorted(set(r['lane'] for r in rows)):
        print(f'    Lane {lane}: {sum(1 for r in rows if r["lane"] == lane)} 条')

    # ── 3. 清空并写入（不去重，逐行复制） ──
    if ws_dst.max_row > 1:
        ws_dst.delete_rows(2, ws_dst.max_row - 1)

    # 解除旧合并
    for mr in list(ws_dst.merged_cells.ranges):
        try:
            ws_dst.unmerge_cells(str(mr))
        except (KeyError, AttributeError):
            pass

    current_row = 2
    prev_lane = None
    lane_start_rows = {}

    for r in rows:
        write_cell(ws_dst, current_row, T_B, r['lane'])
        write_cell(ws_dst, current_row, T_D, r['ssdna_id'])
        write_cell(ws_dst, current_row, T_E, r['count'], '0')
        write_cell(ws_dst, current_row, T_F, r['lib_type'])
        write_cell(ws_dst, current_row, T_G, r['customer'])
        write_cell(ws_dst, current_row, T_H, round(r['data_g'], 2), '0.00')
        write_cell(ws_dst, current_row, T_I, None, '0.00')
        write_cell(ws_dst, current_row, T_J, f'=I{current_row}-H{current_row}', '0.00')
        write_cell(ws_dst, current_row, T_K, f'=IF(H{current_row}=0,"",J{current_row}/H{current_row})', '0.00%')
        write_cell(ws_dst, current_row, T_L, 'T7+150', fill=LIGHT_GREEN)

        # 追踪 lane 范围用于合并
        if r['lane'] != prev_lane:
            if prev_lane is not None and prev_lane in lane_start_rows:
                lane_start_rows[prev_lane] = (lane_start_rows[prev_lane], current_row - 1)
            lane_start_rows[r['lane']] = current_row
            prev_lane = r['lane']

        current_row += 1

    # 最后一个 lane
    if prev_lane is not None and prev_lane in lane_start_rows:
        lane_start_rows[prev_lane] = (lane_start_rows[prev_lane], current_row - 1)

    # ── 4. 合并 A/B/C 列（按lane合并），L列全表合并 ──
    for lane, (start, end) in lane_start_rows.items():
        if end > start:
            for col in (T_A, T_B, T_C):
                ws_dst.merge_cells(
                    start_row=start, start_column=col,
                    end_row=end, end_column=col
                )
        ws_dst.cell(row=start, column=T_A).alignment = CENTER
        ws_dst.cell(row=start, column=T_B).alignment = CENTER
        ws_dst.cell(row=start, column=T_C).alignment = CENTER

    # L列: 全表合并（同一个run）
    total_rows = current_row - 2
    if total_rows > 1:
        ws_dst.merge_cells(start_row=2, start_column=T_L,
                           end_row=current_row - 1, end_column=T_L)
        ws_dst.cell(row=2, column=T_L).alignment = CENTER

    if pool_wb is None:
        wb.save(DST)
    print(f'  写入完成: {current_row - 2} 行')
    print(f'\n[DONE] 步骤七完成 → {DST}')


def write_cell(ws, row, col, value, number_format=None, fill=None):
    """写入单元格并设置居中及可选数字格式/填充"""
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.alignment = CENTER
    if number_format:
        cell.number_format = number_format
    if fill:
        cell.fill = fill


if __name__ == '__main__':
    main()
