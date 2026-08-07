"""
统一格式: 字体、行高、数字格式、列宽
"""
import sys
from pathlib import Path
_sys_path = str(Path(__file__).resolve().parent.parent.parent)
if _sys_path not in sys.path:
    sys.path.insert(0, _sys_path)

import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from config import DST, get_target_sheets, is_data_sheet
from common.format_utils import apply_standard_format

FONT = Font(name='Times New Roman', size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
LIGHT_RED = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')


def format_offline_stats(ws):
    """下机数据统计模版 专属格式"""
    max_row = ws.max_row
    if max_row < 2:
        return

    # ── 数字格式 ──
    for row in range(2, max_row + 1):
        for col in (7, 8, 9):  # G, H, I: 两位小数
            try:
                ws.cell(row=row, column=col).number_format = '0.00'
            except AttributeError:
                pass
        # D列: 整数
        try:
            ws.cell(row=row, column=4).number_format = '0'
        except AttributeError:
            pass
        # J列: 百分比
        try:
            ws.cell(row=row, column=10).number_format = '0.00%'
        except AttributeError:
            pass

    # ── 框线 A-N, 字体 ──
    for row in range(1, max_row + 1):
        for col in range(1, 15):
            try:
                cell = ws.cell(row=row, column=col)
                cell.border = THIN_BORDER
                cell.font = FONT
            except AttributeError:
                pass

    # ── 条件格式: I/J < 0 → 浅红色填充 ──
    ws.conditional_formatting.add(
        f'I2:I{max_row}',
        CellIsRule(operator='lessThan', formula=['0'], fill=LIGHT_RED)
    )
    ws.conditional_formatting.add(
        f'J2:J{max_row}',
        CellIsRule(operator='lessThan', formula=['0'], fill=LIGHT_RED)
    )

    # ── 行高 ──
    ws.row_dimensions[1].height = 40
    for row in range(2, max_row + 1):
        ws.row_dimensions[row].height = 30


def main(pool_wb=None):
    wb = pool_wb if pool_wb is not None else openpyxl.load_workbook(DST)

    apply_standard_format(
        wb,
        sheet_order=get_target_sheets() + ['文库环化', 'T7+制备', 'B1-B3'],
        data_sheet_predicate=is_data_sheet,
        header_height_map={'文库环化': 40, 'T7+制备': 40},
        data_header_height=90,
        number_formats={'D': '0.00', 'E': '0.000', 'F': '0.000', 'I': '0.000'},
    )

    # T7+制备 H列(数据量G): 保留两位小数
    if 'T7+制备' in wb.sheetnames:
        ws_t7 = wb['T7+制备']
        for row in range(2, ws_t7.max_row + 1):
            cell = ws_t7.cell(row=row, column=8)
            try:
                cell.number_format = '0.00'
            except AttributeError:
                pass

    # 下机数据统计模版 专属格式
    if '下机数据统计模版' in wb.sheetnames:
        format_offline_stats(wb['下机数据统计模版'])

    if pool_wb is None:
        wb.save(DST)
    print(f'\n[DONE] 完成 -> {DST}')


if __name__ == '__main__':
    main()
