"""
步骤二：从C表/D表查找并填充 B表(Pooling表)
=============================================
匹配规则:
  B列以 HGC-Lib- / HGC-POOL- 开头 → C表 (搜所有sheet, F列匹配)
  其他 → D表 (D列匹配)

HGC类型 (C表 → B表):
  L列(Qubit浓度) → C列 + N列
  V列(文库结构)  → O列
  W列(磷酸化*)   → P列
  X列(环化*)     → Q列
  O列(板号)      → K列
  J列(孔号)      → 不填

非HGC类型 (D表 → B表):
  K列(Qubit浓度) → C列 + N列
  O列(结果评价)  → O列
  T列(版号)      → K列
  S列(孔位)      → J列
"""
import openpyxl
from openpyxl.styles import Alignment
from collections import defaultdict

from config import DST, get_src_c, get_src_d, get_target_sheets

CENTER = Alignment(horizontal='center', vertical='center')

# ── B表目标列 (1-based) ──
B_C = 3     # Qubit浓度
B_D = 4     # 数据量
B_J = 10    # 孔号
B_K = 11    # 板号
B_M = 13    # 平均片段
B_N = 14    # Qubit浓度 (第二处)
B_O = 15    # 文库结构
B_P = 16    # 磷酸化*
B_Q = 17    # 环化*
B_R = 18    # 人工检测结果
B_S = 19    # 人工备注


def build_c_lookup(src):
    """构建C表查找字典: {HGC编号: row_data} + {文库名称: row_data}，搜所有sheet"""
    wb = openpyxl.load_workbook(src, data_only=True)
    lookup = {}
    lookup_by_name = {}

    for sn in wb.sheetnames:
        ws = wb[sn]
        header_row = None
        col_f = col_g = col_l = col_o = col_v = col_w = col_x = col_m = None
        for row in range(1, min(5, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                v = str(ws.cell(row=row, column=col).value or '')
                if 'HGC编号' in v:
                    col_f = col
                if v == '文库名称':
                    col_g = col
                if 'Qubit浓度' in v:
                    col_l = col
                if v == '板号':
                    col_o = col
                if '文库结构' in v:
                    col_v = col
                if '磷酸化' in v:
                    col_w = col
                if '环化' in v:
                    col_x = col
                if '平均片段' in v:
                    col_m = col
            if col_f:
                header_row = row
                break
        # R/S列: 固定取Z(26)=人工检测结果, AA(27)=人工备注
        col_z = 26
        col_aa = 27

        if not col_f:
            continue

        print(f'  C表[{sn}]: F={col_f}, G={col_g}, L={col_l}, O={col_o}, V={col_v}, W={col_w}, X={col_x}, M={col_m}, Z=26, AA=27')

        for row in range(header_row + 1, ws.max_row + 1):
            key = ws.cell(row=row, column=col_f).value
            if key is None:
                continue
            key = str(key).strip()
            if not key:
                continue

            def get_val(c):
                return ws.cell(row=row, column=c).value if c else None

            rd = {
                'qubit':  get_val(col_l),
                'plate':  get_val(col_o),
                'struct': get_val(col_v),
                'phos':   get_val(col_w),
                'circ':   get_val(col_x),
                'frag':   get_val(col_m),
                'result': get_val(col_z),   # Z列 人工检测结果
                'remark': get_val(col_aa),  # AA列 人工备注
            }
            lookup[key] = rd
            # 文库名称备用索引
            name_key = get_val(col_g)
            if name_key:
                lookup_by_name[str(name_key).strip()] = rd

    wb.close()
    print(f'  C表共加载 {len(lookup)} 条(HGC编号) + {len(lookup_by_name)} 条(文库名称)')
    return lookup, lookup_by_name


def build_d_lookup(src):
    """构建D表查找字典: {样本编号: row_data}"""
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb['自建库出库报告总表']

    col_d = col_k = col_o = col_s = col_t = col_m = None
    for col in range(1, ws.max_column + 1):
        v = str(ws.cell(row=1, column=col).value or '')
        if v == '样本编号':
            col_d = col
        elif 'Qubit浓度' in v:
            col_k = col
        elif '结果评价' in v:
            col_o = col
        elif v == '孔位':
            col_s = col
        elif v == '版号':
            col_t = col
        elif '平均片段' in v:
            col_m = col

    print(f'  D表: D={col_d}, K={col_k}, O={col_o}, S={col_s}, T={col_t}, M={col_m}')

    lookup = {}
    for row in range(2, ws.max_row + 1):
        key = ws.cell(row=row, column=col_d).value
        if key is None:
            continue
        key = str(key).strip()
        if not key:
            continue

        def get_val(c):
            return ws.cell(row=row, column=c).value if c else None

        lookup[key] = {
            'qubit':  get_val(col_k),
            'eval':   get_val(col_o),
            'hole':   get_val(col_s),
            'plate':  get_val(col_t),
            'frag':   get_val(col_m),
        }

    wb.close()
    print(f'  D表共加载 {len(lookup)} 条记录')
    return lookup


def is_hgc_type(lib_id):
    """判断是否为HGC类型"""
    if lib_id is None:
        return False
    s = str(lib_id).strip()
    return s.startswith('HGC-Lib-') or s.startswith('HGC-POOL-')


def fill_sheet(ws_b, c_lookup, c_by_name, d_lookup, name):
    """填充单个工作表"""
    filled = 0
    missed = 0

    for row in range(2, ws_b.max_row + 1):
        lib_id = ws_b.cell(row=row, column=2).value
        if lib_id is None:
            continue
        lib_id = str(lib_id).strip()

        if is_hgc_type(lib_id):
            data = c_lookup.get(lib_id)
            if data:
                write_cell(ws_b, row, B_C, data.get('qubit'))
                write_cell(ws_b, row, B_N, data.get('qubit'))
                write_cell(ws_b, row, B_O, data.get('struct'))
                write_cell(ws_b, row, B_P, data.get('phos'))
                write_cell(ws_b, row, B_Q, data.get('circ'))
                write_cell(ws_b, row, B_K, data.get('plate'))
                write_cell(ws_b, row, B_M, data.get('frag'))
                filled += 1
            else:
                missed += 1
        else:
            # 非HGC: 先查D表，再查C表文库名称
            data = d_lookup.get(lib_id)
            if data:
                write_cell(ws_b, row, B_C, data.get('qubit'))
                write_cell(ws_b, row, B_N, data.get('qubit'))
                write_cell(ws_b, row, B_O, data.get('eval'))
                write_cell(ws_b, row, B_K, data.get('plate'))
                write_cell(ws_b, row, B_J, data.get('hole'))
                write_cell(ws_b, row, B_M, data.get('frag'))
                filled += 1
            elif c_by_name and lib_id in c_by_name:
                data = c_by_name[lib_id]
                write_cell(ws_b, row, B_C, data.get('qubit'))
                write_cell(ws_b, row, B_N, data.get('qubit'))
                write_cell(ws_b, row, B_O, data.get('struct'))
                write_cell(ws_b, row, B_P, data.get('phos'))
                write_cell(ws_b, row, B_Q, data.get('circ'))
                write_cell(ws_b, row, B_K, data.get('plate'))
                write_cell(ws_b, row, B_M, data.get('frag'))
                filled += 1
            else:
                missed += 1

    # ── R/S列: 所有行根据B列查C表F列 ──
    rs_filled = 0
    for row in range(2, ws_b.max_row + 1):
        lib_id = ws_b.cell(row=row, column=2).value
        if lib_id is None:
            continue
        lib_id = str(lib_id).strip()
        c_data = c_lookup.get(lib_id)
        if c_data:
            write_cell(ws_b, row, B_R, c_data.get('result'))
            write_cell(ws_b, row, B_S, c_data.get('remark'))
            rs_filled += 1

    print(f'  Sheet {name}: 匹配={filled}, 未匹配={missed}, R/S填充={rs_filled}')


def write_cell(ws, row, col, value):
    """写入单元格并设置居中"""
    if col is None:
        return
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.alignment = CENTER


def main():
    print('构建C表查找字典...')
    c_lookup, c_by_name = build_c_lookup(get_src_c())

    print('构建D表查找字典...')
    d_lookup = build_d_lookup(get_src_d())

    wb_dst = openpyxl.load_workbook(DST)

    for name in get_target_sheets():
        print(f'\n处理工作表 [{name}]')
        fill_sheet(wb_dst[name], c_lookup, c_by_name, d_lookup, name)

    wb_dst.save(DST)
    print(f'\n{"="*50}\n[DONE] 步骤二完成 → {DST}')


if __name__ == '__main__':
    main()
