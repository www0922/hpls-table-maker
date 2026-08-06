"""
步骤三：B表 分组与排序
=======================
对B表(20260711文库pooling表T7+PE150-zss.xlsx)的A/B/C工作表:

规则:
  1. 判大类:
     直接环化文库 → G列/O列含"华大" 或 P列="已磷酸化"
     转化文库     → 其他
     转化文库在前, 直接环化文库在后, 之间空1行

  2. 池文库/杂交文库:
     池文库编号:   HGC-POOL- 开头 → 单独成一组
     杂交文库编号: 无"HGC-"前缀 (如 260709P1-60) → 单独成一组

  3. 非池文库分组:
     a. 先按客户单位(H列)分组 (应用特殊合并: 两家广州公司合并, XXX-T7合并)
     b. 再按文库类型(G列)细分组

  4. 组间空1行 (不加框线/条件格式)

  5. 组内按文库编号（B列）升序排列
"""
import sys
from pathlib import Path
_sys_path = str(Path(__file__).resolve().parent.parent.parent)
if _sys_path not in sys.path:
    sys.path.insert(0, _sys_path)

import openpyxl
from openpyxl.styles import Alignment

from config import DST, get_target_sheets
from common.d_split import split_rows_by_d_capacity
CENTER = Alignment(horizontal='center', vertical='center')

# ── 列索引 (1-based) ──
COL_B = 2    # 文库编号
COL_D = 4    # 数据量
COL_G = 7    # 文库类型
COL_H = 8    # 客户单位
COL_O = 15   # 文库结构
COL_P = 16   # 磷酸化*

# ── 客户单位特殊合并 ──
CUSTOMER_MERGE = {
    '广州基迪奥科技服务有限公司': '_MERGED_GZ_GROUP_',
    '广州奥智生物科技有限公司': '_MERGED_GZ_GROUP_',
}


def is_direct_circulation(lib_type, lib_struct, phosphorylation):
    """
    判断是否为「直接环化文库」
    条件: 文库类型(G列)含"华大" OR 文库结构(O列)含"华大" OR 磷酸化*(P列)="已磷酸化"
    """
    lt = str(lib_type or '')
    ls = str(lib_struct or '')
    phos = str(phosphorylation or '')
    return ('华大' in lt) or ('华大' in ls) or (phos == '已磷酸化')


def is_pool_library(lib_id):
    """判断B列文库编号是否为杂交文库编号或池文库编号
    杂交文库编号: 无"HGC-"前缀 (如 260709P1-60)
    池文库编号:   HGC-POOL- 开头 (如 HGC-POOL-xxxx)
    """
    if lib_id is None:
        return False
    s = str(lib_id).strip()
    # 池文库编号
    if s.startswith('HGC-POOL-'):
        return True
    # 杂交文库编号: 不含 HGC- 前缀
    if 'HGC-' not in s:
        return True
    return False


def get_customer_key(lib_type, customer):
    """
    获取客户单位分组键
    - XXX-T7 → 视为同一客户单位
    - 广州基迪奥 + 广州奥智 → 视为同一客户单位
    """
    lt = str(lib_type or '').strip()
    cust = str(customer or '').strip()

    if lt == 'XXX-T7':
        return '_XXX_T7_GROUP_'

    return CUSTOMER_MERGE.get(cust, cust)


def _is_hgc(lib_id):
    """判断文库编号是否以 HGC- 开头"""
    return str(lib_id or '').startswith('HGC-')

def _is_bare_groupable(lib_id):
    """非HGC文库：后缀为 X 或 E 时可以成组，否则单独成组"""
    s = str(lib_id or '').strip()
    if not s:
        return False
    last = s[-1].upper()
    return last in ('X', 'E')


def group_rows(rows):
    """
    对同一大类内的行进行分组。

    - HGC-POOL- 池文库: 同客户+同类型合并, 不与 HGC-Lib- 混合
    - HGC-Lib- 普通库: 同客户+同类型合并
    - 非 HGC 杂交文库: 后缀为 X 或 E 的可按客户+类型成组；其余后缀单独成组

    返回: [[row, ...], [row, ...], ...]
    """
    pools = [r for r in rows if str(r.get('lib_id') or '').startswith('HGC-POOL-')]
    hgc_normal = [r for r in rows if _is_hgc(r.get('lib_id'))
                  and not str(r.get('lib_id') or '').startswith('HGC-POOL-')]
    bare = [r for r in rows if not _is_hgc(r.get('lib_id'))]

    sort_key = lambda r: (
        get_customer_key(r.get('lib_type'), r.get('customer')),
        str(r.get('lib_type') or ''),
    )

    def scan_group(lst):
        lst.sort(key=sort_key)
        groups = []
        cur = []
        for r in lst:
            if cur:
                first = cur[0]
                if (get_customer_key(r.get('lib_type'), r.get('customer')),
                    str(r.get('lib_type') or '')) == \
                   (get_customer_key(first.get('lib_type'), first.get('customer')),
                    str(first.get('lib_type') or '')):
                    cur.append(r)
                else:
                    groups.append(cur)
                    cur = [r]
            else:
                cur = [r]
        if cur:
            groups.append(cur)
        return groups

    # 非HGC文库: X/E后缀可成组, 其余单独成组
    bare_groupable = [r for r in bare if _is_bare_groupable(r.get('lib_id'))]
    bare_solo = [r for r in bare if not _is_bare_groupable(r.get('lib_id'))]

    result = scan_group(hgc_normal) + scan_group(pools)
    result += scan_group(bare_groupable)
    result += [[r] for r in bare_solo]  # 每个单独一组

    return result


def split_oversized_groups(groups):
    """拆分 D_sum > 1000 的组, 使用 Best Fit Decreasing 容量装箱, 保证每组 ≤1000G。

    单条 D > 1000 的记录单独成组并打印警告。
    """
    result = []
    for g in groups:
        d_sum = sum(r['cells'].get(COL_D, 0) or 0 for r in g)
        if d_sum <= 1000:
            result.append(g)
            continue

        sub_groups, oversized = split_rows_by_d_capacity(
            g,
            get_d=lambda r: r['cells'].get(COL_D, 0) or 0,
            limit=1000,
        )

        # 超限单条记录: 各自成组并警告
        for r in oversized:
            d_val = r['cells'].get(COL_D, 0) or 0
            print(f'  [WARN] 单条 D={d_val}G > 1000G，无法拆分，单独成组')
            result.append([r])

        result.extend(sub_groups)
    return result


def process_sheet(ws, name):
    """处理单个工作表: 读取→分类→分组→排序→写回"""
    max_col = ws.max_column

    # ── 1. 读取所有数据行 ──
    all_rows = []
    for row in range(2, ws.max_row + 1):
        cells = {}
        for col in range(1, max_col + 1):
            cells[col] = ws.cell(row=row, column=col).value
        all_rows.append({
            'lib_id':         cells.get(COL_B),
            'lib_type':       cells.get(COL_G),
            'customer':       cells.get(COL_H),
            'lib_struct':     cells.get(COL_O),
            'phosphorylation': cells.get(COL_P),
            'cells': cells,
        })

    # ── 2. 分类: 转化文库 / 直接环化文库 ──
    conv_rows = []
    direct_rows = []

    for r in all_rows:
        if is_direct_circulation(r['lib_type'], r['lib_struct'], r['phosphorylation']):
            r['category'] = '直接环化文库'
            direct_rows.append(r)
        else:
            r['category'] = '转化文库'
            conv_rows.append(r)

    print(f'  Sheet {name}: 总行数={len(all_rows)}')
    print(f'    转化文库={len(conv_rows)}, 直接环化文库={len(direct_rows)}')

    # ── 3. 各大类内部分组 ──
    conv_groups   = group_rows(conv_rows)
    direct_groups = group_rows(direct_rows)

    print(f'    转化文库分组:     {len(conv_groups)} 组')
    print(f'    直接环化文库分组: {len(direct_groups)} 组')

    # ── 3.5 D_sum>1000 拆分 ──
    conv_groups   = split_oversized_groups(conv_groups)
    direct_groups = split_oversized_groups(direct_groups)
    print(f'    拆分后转化文库分组:     {len(conv_groups)} 组')
    print(f'    拆分后直接环化文库分组: {len(direct_groups)} 组')

    # ── 4. 组内按文库编号(B列)排序 ──
    for g in conv_groups + direct_groups:
        g.sort(key=lambda r: str(r.get('lib_id') or ''))

    # ── 5. 拼接: 转化(前) + 直接环化(后) ──
    all_groups = conv_groups + direct_groups

    # ── 6. 清空数据行 ──
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # ── 6. 按新顺序写回 ──
    current_row = 2

    for gi, group in enumerate(all_groups):
        # 组间空1行 (第一组前不空)
        if gi > 0:
            current_row += 1  # 留空行

        for row_data in group:
            for col in range(1, max_col + 1):
                cell = ws.cell(row=current_row, column=col)
                if col == 5:  # E列: 标注分类, 供步骤四使用
                    cell.value = row_data.get('category', '')
                else:
                    cell.value = row_data['cells'].get(col)
                cell.alignment = CENTER
            current_row += 1

    print(f'    写入完成: 共 {current_row - 2} 行 (含间隔空行)')


def main():
    wb = openpyxl.load_workbook(DST)

    for name in get_target_sheets():
        print(f'\n{"─"*50}')
        print(f'处理工作表 [{name}]')
        print(f'{"─"*50}')
        process_sheet(wb[name], name)

    wb.save(DST)
    print(f'\n{"="*50}')
    print(f'[DONE] 步骤三完成 → {DST}')


if __name__ == '__main__':
    main()
