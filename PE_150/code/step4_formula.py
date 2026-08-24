"""
步骤四：B表 公式计算、重分组、体积调整、汇总
=============================================
对B表(20260711文库pooling表T7+PE150-zss.xlsx)的A/B/C工作表:

任务:
  1. 计算 E列(文库摩尔质量)、F列(最终取样体积)
  2. 检查重分组条件(总数据量>1000G 或 max/min>5)
  3. 重分组后组内按K列(板号)排序
  4. 体积调整(F<0.250 → 放大)
  5. 填写汇总行
"""
import sys
from pathlib import Path
_sys_path = str(Path(__file__).resolve().parent.parent.parent)
if _sys_path not in sys.path:
    sys.path.insert(0, _sys_path)

import openpyxl
from openpyxl.styles import Alignment, PatternFill
from config import DST, get_target_sheets
from common.d_split import split_rows_by_d_capacity
CENTER = Alignment(horizontal='center', vertical='center')
LIGHT_BLUE = PatternFill(start_color='99CCFF', end_color='99CCFF', fill_type='solid')
YELLOW = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# ── 列索引 (1-based) ──
COL_A = 1    # Lane编号
COL_B = 2    # 文库编号
COL_C = 3    # Qubit浓度
COL_D = 4    # 数据量
COL_E = 5    # 文库摩尔质量
COL_F = 6    # 最终取样体积
COL_G = 7    # 文库类型
COL_H = 8    # 客户单位
COL_I = 9    # TE Buffer补充体积
COL_J = 10   # 孔号
COL_K = 11   # 板号
COL_L = 12   # 备注
COL_N = 14   # Qubit浓度(二)
COL_O = 15   # 文库结构
COL_P = 16   # 磷酸化*
COL_M = 13   # 平均片段
COL_Q = 17   # 环化*


def safe_float(v):
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _try_float(v):
    """尝试将值转为 float，失败或为空返回 None（区别于 safe_float 返回 0）。"""
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def read_groups(ws):
    """读取工作表数据行, 按空白行分隔为组"""
    max_col = ws.max_column
    groups = []
    current = []

    for row in range(2, ws.max_row + 1):
        values = [ws.cell(row=row, column=c).value for c in range(1, max_col + 1)]
        if all(v is None for v in values):
            if current:
                groups.append(current)
                current = []
            continue

        rd = {
            'B': ws.cell(row=row, column=COL_B).value,
            'C': safe_float(ws.cell(row=row, column=COL_C).value),
            'D': safe_float(ws.cell(row=row, column=COL_D).value),
            'N': ws.cell(row=row, column=COL_N).value,  # 保留原始值，空值即 None，由 _try_float 区分
            'G': ws.cell(row=row, column=COL_G).value,
            'K': ws.cell(row=row, column=COL_K).value,
            'O': ws.cell(row=row, column=COL_O).value,
            'P': ws.cell(row=row, column=COL_P).value,
            'category': str(ws.cell(row=row, column=COL_E).value or ''),  # 步骤三写入的分类标记
            'cells': {c: ws.cell(row=row, column=c).value for c in range(1, max_col + 1)},
        }
        current.append(rd)

    if current:
        groups.append(current)
    return groups


def calc_EF(rows, molar_mass):
    """计算组内每行的 E列 和 F列, 返回实际使用的摩尔质量
    转化文库: 组内任一行 N列 < 2.6 → 整组 ×10, 否则 ×50
    直接环化: 固定 ×300
    """
    d_sum = sum(r['D'] for r in rows)
    if d_sum == 0:
        for r in rows:
            r['E'] = 0.0
            r['F'] = 0.0
        return molar_mass

    is_conv = (molar_mass == 50.0)
    if is_conv:
        # 检查组内是否有 N < 2.6（仅检查有有效数值的行，空值跳过不触发）
        n_vals = [
            v for r in rows
            if (v := _try_float(r.get('N'))) is not None
        ]
        has_low_n = any(v < 2.6 for v in n_vals)
        mm = 10.0 if has_low_n else 50.0
    else:
        mm = molar_mass

    for r in rows:
        r['E'] = round(r['D'] / d_sum * mm, 3)
        r['F'] = round(r['E'] / r['C'], 3) if r['C'] > 0 else 0.0

    return mm


def regroup(rows, molar_mass):
    """
    检查重分组条件, 必要时拆分并重新计算。
    重分组条件: D列总和 > 1000G, 或 F列 max/min > 5。
    切分时同时约束 F 比值 ≤ 5 和 D合计 ≤ 1000, 双重保险。
    返回: [(sub_rows, actual_molar_mass), ...]
    """
    actual_mm = calc_EF(rows, molar_mass)
    d_sum = sum(r['D'] for r in rows)
    f_vals = [r['F'] for r in rows if r['F'] > 0]

    if not f_vals:
        # 无有效 F 值时仍须检查 D_sum 上限
        if d_sum > 1000:
            sub_groups, oversized = split_rows_by_d_capacity(
                rows, get_d=lambda r: r['D'], limit=1000
            )
            for r in oversized:
                print(f'  [WARN] 单条 D={r["D"]}G > 1000G，无法拆分，单独成组')
                sub_groups.append([r])
        else:
            sub_groups = [rows]

        result = []
        for sg in sub_groups:
            sg.sort(key=lambda r: str(r.get('B') or ''))
            sg_mm = calc_EF(sg, molar_mass)
            result.append((sg, sg_mm))
        return result

    need = (d_sum > 1000) or (max(f_vals) / min(f_vals) > 5 if min(f_vals) > 0 else False)
    if not need:
        rows.sort(key=lambda r: str(r.get('B') or ''))  # B列升序
        return [(rows, actual_mm)]

    # ── 重分组: 按F列排序, 同时约束F比值≤5和D合计≤1000 ──
    sorted_by_f = sorted(rows, key=lambda r: r['F'])

    sub_groups = []
    cur = [sorted_by_f[0]]
    cur_min = cur_max = sorted_by_f[0]['F']
    cur_d_sum = sorted_by_f[0]['D']

    for r in sorted_by_f[1:]:
        new_min = min(cur_min, r['F'])
        new_max = max(cur_max, r['F'])
        new_d_sum = cur_d_sum + r['D']
        # 切分条件: F比值超过5, 或 D合计超过1000
        f_exceed = (new_min > 0 and new_max / new_min > 5)
        d_exceed = (new_d_sum > 1000)
        if f_exceed or d_exceed:
            sub_groups.append(cur)
            cur = [r]
            cur_min = cur_max = r['F']
            cur_d_sum = r['D']
        else:
            cur.append(r)
            cur_min = new_min
            cur_max = new_max
            cur_d_sum = new_d_sum

    if cur:
        sub_groups.append(cur)

    # 每个子组: B列升序 → 重新计算, 再板号排序
    result = []
    for sg in sub_groups:
        sg.sort(key=lambda r: str(r['B'] or ''))
        sg_mm = calc_EF(sg, molar_mass)
        result.append((sg, sg_mm))

    return result


def sort_by_plate(rows):
    """组内按K列(板号)排序(稳定排序, 保持B列顺序), 无板号排末尾"""
    def key(r):
        k = r.get('K')
        return (0, str(k)) if (k is not None and str(k).strip()) else (1, '')
    rows.sort(key=key)  # Python sort is stable, preserves prior B-column order


def apply_scale(rows):
    """
    检查F列最小值, 确定缩放倍数N
    有板号: F < 1 → 放大到 > 1
    无板号: F < 0.250 → 放大到 > 0.500
    返回: scale_factor (1 表示未放大)
    """
    f_vals = [r['F'] for r in rows if r['F'] > 0]
    if not f_vals:
        return 1

    # 是否有板号
    has_plate = any(
        r.get('K') is not None and str(r.get('K')).strip()
        for r in rows
    )

    if has_plate:
        threshold = 1.0
        if min(f_vals) >= threshold:
            return 1
        n = 1
        while min(f_vals) * n <= threshold:
            n += 1
    else:
        if min(f_vals) >= 0.250:
            return 1
        n = 1
        while min(f_vals) * n <= 0.500:
            n += 1

    for r in rows:
        r['F'] = round(r['F'] * n, 3)
    return n


def process_sheet(ws, name):
    """处理单个工作表"""
    max_col = ws.max_column
    groups = read_groups(ws)
    print(f'  Sheet {name}: 读取 {len(groups)} 组')

    # ── 处理每个组 ──
    all_final = []  # [(rows, molar_mass, scale, is_direct), ...]

    for g in groups:
        # 从步骤三的标记读取分类
        is_direct = (g[0].get('category', '') == '直接环化文库') if g else False
        molar_mass = 300.0 if is_direct else 50.0

        # 重分组 → [(sub_rows, sub_molar_mass), ...]
        sub_results = regroup(g, molar_mass)

        for sg_rows, sg_mm in sub_results:
            scale = apply_scale(sg_rows)
            all_final.append((sg_rows, sg_mm, scale, is_direct))

    regroup_count = len(all_final) - len(groups)
    print(f'    处理: {len(groups)} → {len(all_final)} 组 (新增 {regroup_count})')

    # ── 排序: 10nM组在前 → 50nM → 300nM(直接环化) ──
    mm_order = {10.0: 0, 50.0: 1, 300.0: 2}
    all_final.sort(key=lambda x: mm_order.get(x[1], 99))

    # ── 清空并写回 ──
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    current_row = 2
    for sg_rows, sg_mm, scale, is_direct in all_final:
        d_first = current_row
        d_last = current_row + len(sg_rows) - 1
        summary_row = d_last + 1
        is_low_n = (sg_mm == 10.0)  # 10nM组(因Qubit浓度<2.6触发)

        # 写入数据行
        for rd in sg_rows:
            for col in range(1, max_col + 1):
                cell = ws.cell(row=current_row, column=col)
                if col == COL_E:
                    # E列公式: IF(C为空, 强制用50, 否则用组MM)
                    cell.value = f'=ROUND(D{current_row}/SUM(D{d_first}:D{d_last})*IF(C{current_row}="",50,{sg_mm}),3)'
                elif col == COL_F:
                    cell.value = f'=ROUND(E{current_row}/C{current_row}*G{summary_row},3)'
                else:
                    cell.value = rd['cells'].get(col)
                cell.alignment = CENTER
            current_row += 1

        # ── 汇总行 ──
        # B列: 数据条数
        c = ws.cell(row=summary_row, column=COL_B)
        c.value = len(sg_rows)
        c.alignment = CENTER

        # D列: =ROUND(SUM,2)
        c = ws.cell(row=summary_row, column=COL_D)
        c.value = f'=ROUND(SUM(D{d_first}:D{d_last}),2)'
        c.alignment = CENTER

        # E列: =ROUND(SUM,3)
        c = ws.cell(row=summary_row, column=COL_E)
        c.value = f'=ROUND(SUM(E{d_first}:E{d_last}),3)'
        c.alignment = CENTER

        # F列: =IF(G>1,ROUND(SUM/G*2,3),ROUND(SUM,3))
        c = ws.cell(row=summary_row, column=COL_F)
        c.value = f'=IF(G{summary_row}>1,ROUND(SUM(F{d_first}:F{d_last})/G{summary_row}*2,3),ROUND(SUM(F{d_first}:F{d_last}),3))'
        c.alignment = CENTER

        # G列: 倍数(静态)
        c = ws.cell(row=summary_row, column=COL_G)
        c.value = scale
        c.alignment = CENTER

        # I列: =ROUND(20/40-F,3), 浅蓝背景
        c = ws.cell(row=summary_row, column=COL_I)
        c.value = f'=ROUND(20-F{summary_row},3)' if not is_direct else f'=ROUND(40-F{summary_row},3)'
        c.alignment = CENTER
        c.fill = LIGHT_BLUE

        # M列: =AVERAGE(片段范围)
        c = ws.cell(row=summary_row, column=COL_M)
        c.value = f'=AVERAGE(M{d_first}:M{d_last})'
        c.alignment = CENTER

        # 10nM组: 汇总行不填色 (数据行已在上面填过)

        current_row = summary_row + 1

    print(f'    写入完成: {current_row - 2} 行 (含间隔和汇总行)')


def main(pool_wb=None):
    wb = pool_wb if pool_wb is not None else openpyxl.load_workbook(DST)

    for name in get_target_sheets():
        print(f'\n{"─"*50}')
        print(f'处理工作表 [{name}]')
        print(f'{"─"*50}')
        process_sheet(wb[name], name)

    if pool_wb is None:
        wb.save(DST)
    print(f'\n{"="*50}')
    print(f'步骤四完成 → {DST}')


if __name__ == '__main__':
    main()
