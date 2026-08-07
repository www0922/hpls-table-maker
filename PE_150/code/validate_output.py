"""PE150 输出校验。

检查项：
  1. 输出文件可正常打开
  2. 数据 sheet 存在且与 A 表匹配
  3. 每个数据 sheet 的分组结构正确
  4. Lane 编号连续
  5. 文库环化 sheet 有数据（引用所有数据 sheet 的 lane）
  6. T7+制备 有数据
  7. Sheet 排列顺序正确
"""
import sys
from pathlib import Path
_sys_path = str(Path(__file__).resolve().parent.parent.parent)
if _sys_path not in sys.path:
    sys.path.insert(0, _sys_path)

import openpyxl

from config import DST, get_target_sheets, is_data_sheet
from common.validate_utils import (
    validate_data_sheet,
    check_sheet_order,
    check_t7_sheet,
)


def check_huanhua_sheet(ws, target_sheets, wb):
    """校验文库环化 sheet：
    - 有数据
    - 包含所有数据 sheet 的首字母（如 C 面和 D 面各至少一行）
    - 字母变化处有空行分隔或 lane 编号能区分
    """
    errors = []
    if ws.max_row < 2:
        errors.append('文库环化: sheet 为空')
        return errors

    # 读取 A 列所有 lane 编号
    lanes = []
    for row in range(2, ws.max_row + 1):
        val = str(ws.cell(row=row, column=1).value or '').strip()
        if val:
            lanes.append(val)

    if not lanes:
        errors.append('文库环化: 无 lane 数据（A列全空）')
        return errors

    # 检查每个数据 sheet 的首字母是否出现
    for sn in target_sheets:
        if not any(lane.startswith(sn) for lane in lanes):
            errors.append(f'文库环化: 缺少 {sn} 面的 lane')

    # Lane 数应与数据 sheet 总组数一致
    total_expected = 0
    for sn in target_sheets:
        if sn in wb.sheetnames:
            from common.validate_utils import read_groups
            groups = read_groups(wb[sn])
            total_expected += len(groups)
    if len(lanes) != total_expected:
        errors.append(
            f'文库环化: lane 数({len(lanes)})与总组数({total_expected})不一致'
        )

    return errors


def main(pool_wb=None):
    errors = []
    warnings = []
    total_groups = 0

    # ── 1. 打开输出文件 ──
    try:
        wb = pool_wb if pool_wb is not None else openpyxl.load_workbook(DST)
    except Exception as e:
        print(f'[FAILED] 无法打开输出文件: {e}')
        raise SystemExit(1)

    target_sheets = get_target_sheets()

    # ── 2. 校验每个数据 sheet ──
    for sn in target_sheets:
        if sn not in wb.sheetnames:
            errors.append(f'缺少数据 sheet: [{sn}]')
            continue
        ws = wb[sn]
        sheet_errors, sheet_warnings, groups = validate_data_sheet(ws, sn)
        errors.extend(sheet_errors)
        warnings.extend(sheet_warnings)
        total_groups += groups

        # D 合计 ≤ 1000G 硬校验
        from common.validate_utils import read_groups
        for g in read_groups(ws):
            d_vals = []
            for r in g['data_rows']:
                v = ws.cell(row=r, column=4).value
                try:
                    d_vals.append(float(v) if v is not None else 0)
                except (ValueError, TypeError):
                    d_vals.append(0)
            d_sum = sum(d_vals)
            if d_sum > 1000:
                errors.append(
                    f'{sn}!行{g["data_rows"][0]}: '
                    f'组 D 合计 {d_sum}G 超过 1000G 上限'
                )

        status = 'OK' if not sheet_errors else f'{len(sheet_errors)} errors'
        print(f'  Sheet [{sn}]: {groups} 组  ({status})')

    # ── 3. 文库环化 ──
    if '文库环化' in wb.sheetnames:
        hh_errors = check_huanhua_sheet(wb['文库环化'], target_sheets, wb)
        errors.extend(hh_errors)
        if not hh_errors:
            print(f'  Sheet [文库环化]: OK')
    else:
        errors.append('缺少 sheet: [文库环化]')

    # ── 4. T7+制备 ──
    if 'T7+制备' in wb.sheetnames:
        t7_errors = check_t7_sheet(wb['T7+制备'])
        errors.extend(t7_errors)
        if not t7_errors:
            print(f'  Sheet [T7+制备]: OK')
    else:
        errors.append('缺少 sheet: [T7+制备]')

    # ── 5. Sheet 顺序 ──
    expected_order = target_sheets + ['文库环化', 'T7+制备', 'B1-B3']
    order_errors = check_sheet_order(wb, expected_order)
    errors.extend(order_errors)

    if pool_wb is None:
        wb.close()

    # ── 6. 报告 ──
    print(f'\n校验统计: {len(target_sheets)} 个数据 sheet, 共 {total_groups} 组')
    if errors:
        print(f'[FAILED] {len(errors)} 项错误:')
        for e in errors[:30]:
            print(f'  - {e}')
        if len(errors) > 30:
            print(f'  ... 还有 {len(errors) - 30} 项')
        raise SystemExit(1)
    if warnings:
        print(f'[WARNING] {len(warnings)} 项提示:')
        for w in warnings[:10]:
            print(f'  - {w}')
    print('[PASSED] 业务校验通过')


if __name__ == '__main__':
    main()
